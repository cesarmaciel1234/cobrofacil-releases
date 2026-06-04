"""
Módulo de Importación / Exportación de Productos
Formato Excel Flexible: Soportando precios, ofertas y lectura robusta de columnas sin importar su posición.
Optimizado para cargas masivas instantáneas (17,000+ registros en lote/WAL).
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

try:
    from src.base_de_datos.database import db_manager
except ImportError:
    from database import db_manager

# ── Columnas por Defecto para Exportación ───────
COLS = [
    ("Código",        "id"),
    ("Producto",      "nombre"),
    ("P. Costo",      "costo"),
    ("P. Venta",      "precio"),
    ("P. Mayoreo",    "precio_mayoreo"),
    ("Cant. Oferta",  "cant_oferta"),
    ("P. Oferta",     "precio_oferta"),
    ("Departamento",  "departamento"),
    ("Existencia",    "stock"),
    ("Inv. Mínimo",   "stock_minimo"),
    ("Inv. Máximo",   "stock_maximo"),
    ("Tipo de Venta", "unidad"),      # KG→GRANEL, UN→UNIDAD
]

HEADER_FILL  = PatternFill("solid", fgColor="FFD966")   # amarillo dorado
HEADER_FONT  = Font(bold=True, size=10)
BORDER_THIN  = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

# ─────────────────────────────────────────────────────────
def exportar_excel(filepath: str) -> tuple[bool, str]:
    """Exporta todos los productos al archivo Excel indicado con las columnas completas de ofertas."""
    try:
        rows = db_manager.execute_query(
            "SELECT id, nombre, costo, precio, precio_mayoreo, cant_oferta, precio_oferta, departamento, stock, stock_minimo, stock_maximo, unidad "
            "FROM productos ORDER BY departamento, nombre"
        ) or []

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"

        # ── Encabezado ────────────────────────────────────
        headers = [c[0] for c in COLS]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = BORDER_THIN

        # ── Datos ─────────────────────────────────────────
        for row_idx, r in enumerate(rows, 2):
            tipo = "GRANEL" if (r['unidad'] or '').upper() == 'KG' else "UNIDAD"
            valores = [
                r['id'],
                r['nombre'] or '',
                r['costo'] or 0.0,
                r['precio'] or 0.0,
                r['precio_mayoreo'] or 0.0,
                r['cant_oferta'] or 0.0,
                r['precio_oferta'] or 0.0,
                r['departamento'] or '',
                r['stock'] or 0.0,
                r['stock_minimo'] or 0.0,
                r['stock_maximo'] or 0.0,
                tipo,
            ]
            for col_idx, val in enumerate(valores, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border    = BORDER_THIN
                cell.alignment = Alignment(vertical="center")
                # Formato moneda para precios
                if col_idx in (3, 4, 5, 7):
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx in (6, 9, 10, 11):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        # ── Anchos de columna ────────────────────────────
        anchos = [12, 28, 12, 12, 12, 12, 12, 16, 12, 12, 12, 14]
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        ws.freeze_panes = "A2"
        wb.save(filepath)
        return True, f"Se exportaron {len(rows)} productos a:\n{filepath}"

    except Exception as e:
        return False, f"Error al exportar: {e}"


# ─────────────────────────────────────────────────────────
def importar_excel(filepath: str) -> tuple[bool, str]:
    """
    Importa productos desde un Excel de forma 'universal' e 'inteligente'.
    Detecta las columnas automáticamente analizando sus nombres y utilizando
    palabras clave (ej: 'p. venta', 'precio', 'venta' -> Precio Venta),
    sin importar su posición.
    Optimizado para 17,000+ productos en lote/WAL.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        # 1. Definir aliases (palabras clave) para cada campo lógico
        alias_map = {
            'codigo': ['código', 'codigo', 'cod', 'id', 'referencia', 'sku', 'barra'],
            'nombre': ['producto', 'nombre', 'descripcion', 'descripción', 'articulo', 'artículo', 'detalle'],
            'costo': ['costo', 'compra', 'p. costo', 'p.costo', 'precio costo', 'precio de compra', 'p costo'],
            'precio': ['venta', 'p. venta', 'p.venta', 'precio', 'precio de venta', 'precio publico', 'p venta'],
            'mayoreo': ['mayoreo', 'p. mayoreo', 'mayorista', 'precio por mayor', 'p.mayoreo'],
            'cant_oferta': ['cant. oferta', 'cant oferta', 'cantidad oferta', 'oferta_cant'],
            'precio_oferta': ['p. oferta', 'precio oferta', 'precio_oferta', 'oferta'],
            'stock': ['existencia', 'stock', 'cantidad', 'inventario', 'disponible', 'cant'],
            'minimo': ['minimo', 'mínimo', 'inv. minimo', 'inv. mínimo'],
            'maximo': ['maximo', 'máximo', 'inv. maximo', 'inv. máximo'],
            'depto': ['departamento', 'depto', 'categoria', 'rubro', 'familia', 'línea', 'linea'],
            'tipo': ['tipo de venta', 'tipo', 'unidad', 'medida', 'formato']
        }

        # 2. Buscar la fila de encabezado (buscando en las primeras 20 filas)
        header_row = -1
        col_indices = {} # mapea campo_logico -> indice_columna (0-based)

        for row in ws.iter_rows(min_row=1, max_row=20):
            row_vals = [str(c.value or '').strip().lower() for c in row]
            
            # Intento de mapeo para esta fila
            temp_map = {}
            for col_idx, cell_val in enumerate(row_vals):
                if not cell_val: continue
                # Limpiar saltos de linea y espacios multiples
                cell_val = ' '.join(cell_val.split())
                
                # Buscar a qué campo lógico pertenece esta columna
                for campo, keywords in alias_map.items():
                    if campo not in temp_map: # Solo tomar el primero que coincida por campo
                        # Si alguna keyword está dentro del nombre de la columna o viceversa
                        if any(kw == cell_val or kw in cell_val or cell_val.startswith(kw) for kw in keywords):
                            temp_map[campo] = col_idx
                            break
            
            # Consideramos que es la fila de encabezado si encontramos al menos 'nombre' o 'codigo' y 'precio'
            if ('nombre' in temp_map or 'codigo' in temp_map) and ('precio' in temp_map or 'costo' in temp_map or 'stock' in temp_map):
                header_row = row[0].row
                col_indices = temp_map
                break

        if header_row == -1 or not col_indices:
            return False, "No se pudo detectar automáticamente la estructura del Excel. Asegúrate de que las columnas tengan nombres descriptivos como 'Código', 'Descripción', 'P. Venta'."

        # Si encontramos el encabezado, validar que tengamos al menos la columna nombre o código para poder importar
        if 'nombre' not in col_indices and 'codigo' not in col_indices:
            return False, "Falta la columna principal de 'Producto/Descripción' o 'Código'."

        insertados = 0; actualizados = 0; errores = 0

        # MODO ALTO RENDIMIENTO: Usar 1 sola conexión y 1 sola transacción en memoria
        conn = db_manager.get_connection()
        cursor = conn.cursor()

        # Pre-cargar identificadores existentes para búsqueda O(1) en memoria
        cursor.execute("SELECT id, codigo FROM productos")
        existentes_id = {}
        existentes_cod = {}
        for r in cursor.fetchall():
            existentes_id[str(r['id'])] = r['id']
            if r['codigo']:
                existentes_cod[str(r['codigo'])] = r['id']

        import re

        # Extraer filas de valores, omitiendo el encabezado y las anteriores
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(row): continue  # fila vacía o nula

            def get_val(campo_logico, default=''):
                if campo_logico in col_indices:
                    idx = col_indices[campo_logico]
                    if idx < len(row):
                        v = row[idx]
                        return v if v is not None else default
                return default

            def parse_float(val):
                if val is None or val == '': return 0.0
                if isinstance(val, (int, float)): return float(val)
                s = str(val).strip()
                # Extraer solo números, comas, puntos y signo negativo
                s = re.sub(r'[^\d.,-]', '', s)
                if not s: return 0.0
                if ',' in s and '.' not in s: s = s.replace(',', '.')
                elif ',' in s and '.' in s: s = s.replace('.', '').replace(',', '.')
                try: return float(s)
                except: return 0.0

            codigo_val  = str(get_val('codigo', '')).strip()
            if codigo_val.endswith('.0'):
                codigo_val = codigo_val[:-2]

            nombre = str(get_val('nombre', '')).strip()
            
            # Lógica de Descripción e ID (código) variado
            # Si no hay nombre pero hay código, el código funciona como nombre
            if not nombre and codigo_val:
                nombre = "Producto " + codigo_val
            # Si no hay código pero hay nombre, intentamos generar un código o dejamos que la base de datos lo asigne
            elif not codigo_val and nombre:
                # Opcional: Podríamos intentar extraer algún número del nombre, pero mejor dejamos que SQLite auto-asigne el ID
                pass
            
            if not nombre and not codigo_val: continue

            costo       = parse_float(get_val('costo', 0))
            precio      = parse_float(get_val('precio', 0))
            mayoreo     = parse_float(get_val('mayoreo', 0))
            cant_of     = parse_float(get_val('cant_oferta', 0))
            precio_of   = parse_float(get_val('precio_oferta', 0))
            stock       = parse_float(get_val('stock', 0))
            minimo      = parse_float(get_val('minimo', 0))
            maximo      = parse_float(get_val('maximo', 0))

            depto  = str(get_val('depto', '')).strip() or None
            tipo   = str(get_val('tipo', 'UNIDAD')).strip().upper()
            unidad = 'KG' if 'GRANEL' in tipo or 'KG' in tipo else 'UN'
            es_pes = 1 if unidad == 'KG' else 0
            cat    = depto.upper() if depto else 'GENERAL'

            # Búsqueda instantánea O(1) en los diccionarios pre-cargados
            existe_id = None
            if codigo_val and codigo_val.isdigit():
                existe_id = existentes_id.get(codigo_val)
            if not existe_id and codigo_val:
                existe_id = existentes_cod.get(codigo_val)

            try:
                if existe_id:
                    # Actualizar solo si hay nombre, de lo contrario mantener el nombre anterior no es posible fácilmente aquí
                    # Si 'nombre' está vacío pero la columna sí existe (el usuario la vació), podríamos ignorar, pero mejor actualizar
                    # Para seguridad, si nombre viene vacío pero detectó columna, y existe, podemos dejar el de la BD (omitimos por brevedad)
                    cursor.execute(
                        "UPDATE productos SET nombre=COALESCE(NULLIF(?, ''), nombre),precio=?,costo=?,stock=?,precio_mayoreo=?,stock_minimo=?,stock_maximo=?,"
                        "unidad=?,es_pesable=?,departamento=?,categoria=?,cant_oferta=?,precio_oferta=? WHERE id=?",
                        (nombre, precio, costo, stock, mayoreo, minimo, maximo, unidad, es_pes, depto, cat, cant_of, precio_of, existe_id))
                    actualizados += 1
                else:
                    if not nombre: 
                        nombre = "Producto " + codigo_val
                    params = (codigo_val if not codigo_val.isdigit() else None,
                              nombre, precio, costo, stock, mayoreo, minimo, maximo, unidad, es_pes, depto, cat, cant_of, precio_of)
                    if codigo_val and codigo_val.isdigit():
                        cursor.execute(
                            "INSERT OR IGNORE INTO productos "
                            "(id,codigo,nombre,precio,costo,stock,precio_mayoreo,stock_minimo,stock_maximo,unidad,es_pesable,departamento,categoria,cant_oferta,precio_oferta) "
                            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                            (int(codigo_val), None, nombre, precio, costo, stock, mayoreo, minimo, maximo, unidad, es_pes, depto, cat, cant_of, precio_of))
                        existentes_id[codigo_val] = int(codigo_val)
                    else:
                        cursor.execute(
                            "INSERT INTO productos "
                            "(codigo,nombre,precio,costo,stock,precio_mayoreo,stock_minimo,stock_maximo,unidad,es_pesable,departamento,categoria,cant_oferta,precio_oferta) "
                            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                            params)
                        if codigo_val:
                            existentes_cod[codigo_val] = cursor.lastrowid
                    insertados += 1
            except Exception as ex:
                errores += 1

        # Confirmar todo el lote masivo en una sola ráfaga atómica
        conn.commit()
        conn.close()

        msg = (f"Importación masiva completada exitosamente:\n"
               f"  ✔ Insertados:   {insertados}\n"
               f"  ✔ Actualizados: {actualizados}\n"
               f"  ✖ Errores:      {errores}")
        return True, msg

    except Exception as e:
        return False, f"Error al importar: {e}"
