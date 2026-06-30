from src.utils.qt_compat import qt_exec
import sys
from PyQt6.QtWidgets import (

    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QFrame, QWidget,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QLineEdit, QPushButton, QGridLayout, QComboBox, QDateEdit, QTimeEdit, QCheckBox
)
from PyQt6.QtCore import Qt, QDate, QTime, QTimer
from PyQt6.QtGui import QColor, QBrush, QPainter
from datetime import datetime
from src.base_de_datos.database import db_manager
from src.config import config

def fmt_moneda(val):
    """ Formatea a moneda regional: $1.234,56 """
    try:
        return f"${val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return f"${val}"

class StampLabel(QLabel):
    def __init__(self, text="CANCELADO", parent=None):
        super().__init__(text, parent)
        self.setFixedSize(400, 150)
        self.setAlignment(Qt.AlignCenter)
        # Estilo de sello de goma industrial
        self.setStyleSheet("""
            QLabel {
                color: #ef4444;
                font-size: 55px;
                font-weight: 900;
                border: 10px double #ef4444;
                border-radius: 20px;
                background-color: rgba(255, 255, 255, 180);
                padding: 10px;
                letter-spacing: 5px;
            }
        """)
        self.hide()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        # Rotación diagonal clásica de sello
        painter.translate(self.width()/2, self.height()/2)
        painter.rotate(-25)
        painter.translate(-self.width()/2, -self.height()/2)
        super().paintEvent(event)

class VistaHistorial(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)

        if False:
            self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
            self.setAttribute(Qt.WA_TranslucentBackground)
            self.setFixedSize(910, 760)
        if True:
            self.setWindowFlags(Qt.Widget)
            self.setAttribute(Qt.WA_TranslucentBackground, False)
        
        self.ticket_seleccionado = None
        self.setup_ui()
        
        if False:
            self.apply_glow()
            
        if hasattr(self, 'btn_close'):
            self.btn_close.hide()
            
        self.cargar_ventas()

    def apply_glow(self):
        container = self.findChild(QWidget, "HistoryDialog")
        if container:
            glow.setColor(QColor(37, 99, 235, 230)) # Royal Blue ultra-brillante
            glow.setOffset(0, 0)
            container.setGraphicsEffect(glow)

    def apply_theme(self):
        theme = config.get("theme", "light")
        if theme == "dark":
            self.setStyleSheet("""
                #HistoryDialog { 
                    background-color: #0F172A; 
                    border-radius: 20px;
                    border: 2px solid #334155;
                }
                QWidget { font-family: 'Segoe UI', Arial; }
                QLabel { font-size: 13px; color: #F8FAFC; background: transparent; }
                
                QTableWidget { 
                    background-color: #1E293B; 
                    gridline-color: #334155; 
                    border: 1px solid #334155; 
                    border-radius: 12px;
                    selection-background-color: #334155; 
                    selection-color: #38BDF8;
                    font-size: 13px;
                    outline: none;
                    color: #F8FAFC;
                }
                QTableWidget::item { padding: 5px; border-bottom: 1px solid #1E293B; }
                QTableWidget::item:selected {
                    background-color: #3B82F6;
                    color: #FFFFFF;
                    font-weight: 900;
                }
                QHeaderView::section { 
                    background-color: #0F172A; 
                    color: #94A3B8; 
                    padding: 8px; 
                    border: none;
                    font-weight: 900;
                    font-size: 11px;
                    text-transform: uppercase;
                }
                
                QLineEdit, QComboBox, QDateEdit, QTimeEdit { 
                    border: 2px solid #334155; border-radius: 10px; padding: 5px 10px; background: #1E293B; font-size: 13px; font-weight: 600; color: #F8FAFC;
                }
                QLineEdit:focus, QComboBox:focus, QDateEdit:focus, QTimeEdit:focus { border: 2px solid #3B82F6; background: #0F172A; }
                QComboBox::drop-down, QDateEdit::drop-down, QTimeEdit::drop-down { border: none; }
                
                QPushButton { 
                    padding: 8px 15px; 
                    font-size: 13px; 
                    background: #334155;
                    border: 2px solid #475569;
                    border-radius: 8px;
                    color: #CBD5E1;
                    font-weight: 900;
                }
                QPushButton:hover { background: #475569; border-color: #64748B; }
                
                #btnCancelVenta { background: #7F1D1D; color: #FCA5A5; border: none; }
                #btnCancelVenta:hover { background: #991B1B; }
                #btnImprimir { background: #1E3A8A; color: #93C5FD; border: none; }
                #btnImprimir:hover { background: #1E40AF; }
                
                #frameDetalle { 
                    background-color: #1E293B; 
                    border: 1px solid #334155; 
                    border-radius: 16px;
                }
                #lblTicketPreview { font-size: 22px; font-weight: 900; color: #38BDF8; margin-bottom: 5px; }
            """)
            if hasattr(self, 'lbl_titulo'):
                self.lbl_titulo.setStyleSheet("font-size: 18px; font-weight: 900; color: #38BDF8; letter-spacing: 2px; background: transparent;")
            if hasattr(self, 'footer'):
                self.footer.setStyleSheet("background: #0F172A; border-top: 1px solid #334155; border-bottom-left-radius: 20px; border-bottom-right-radius: 20px;")
            if hasattr(self, 'btn_close'):
                self.btn_close.setStyleSheet("""
                    QPushButton { 
                        background-color: #1E293B; 
                        color: #94A3B8; 
                        border-radius: 12px; 
                        font-weight: 900; 
                        font-size: 13px;
                        padding: 15px;
                        border: 2px solid #334155;
                    }
                    QPushButton:hover { background-color: #334155; border-color: #475569; color: #F8FAFC; }
                """)
            if hasattr(self, 'lbl_det_metodo'):
                self.lbl_det_metodo.setStyleSheet("color: #38BDF8; font-weight: bold; background: transparent;")
            if hasattr(self, 'total_card'):
                self.total_card.setStyleSheet("background-color: #1E293B; border: 2px dashed #334155; border-radius: 12px;")
                self.lbl_total_filtrado.setStyleSheet("font-size: 16px; font-weight: 900; color: #38BDF8; border: none; background: transparent;")
        if True:
            self.setStyleSheet("""
                #HistoryDialog { 
                    background-color: white; 
                    border-radius: 20px;
                    border: 2px solid #1E3A8A;
                }
                QWidget { font-family: 'Segoe UI', Arial; }
                QLabel { font-size: 13px; color: #1e293b; background: transparent; }
                
                QTableWidget { 
                    background-color: white; 
                    gridline-color: #f1f5f9; 
                    border: 1px solid #e2e8f0; 
                    border-radius: 12px;
                    selection-background-color: #EFF6FF; 
                    selection-color: #1E3A8A;
                    font-size: 13px;
                    outline: none;
                }
                QTableWidget::item { padding: 5px; border-bottom: 1px solid #F8FAFC; }
                QTableWidget::item:selected {
                    background-color: #EFF6FF;
                    color: #1E3A8A;
                    font-weight: 900;
                }
                QHeaderView::section { 
                    background-color: #F8FAFC; 
                    color: #64748b; 
                    padding: 8px; 
                    border: none;
                    font-weight: 900;
                    font-size: 11px;
                    text-transform: uppercase;
                }
                
                QLineEdit, QComboBox, QDateEdit, QTimeEdit { 
                    border: 2px solid #E2E8F0; border-radius: 10px; padding: 5px 10px; background: #F8FAFC; font-size: 13px; font-weight: 600; color: #1E293B;
                }
                QLineEdit:focus, QComboBox:focus, QDateEdit:focus, QTimeEdit:focus { border: 2px solid #3B82F6; background: white; }
                QComboBox::drop-down, QDateEdit::drop-down, QTimeEdit::drop-down { border: none; }
                
                QPushButton { 
                    padding: 8px 15px; 
                    font-size: 13px; 
                    background: white;
                    border: 2px solid #E2E8F0;
                    border-radius: 8px;
                    color: #475569;
                    font-weight: 900;
                }
                QPushButton:hover { background: #F8FAFC; border-color: #94A3B8; }
                
                #btnCancelVenta { background: #FEF2F2; color: #DC2626; border: none; }
                #btnCancelVenta:hover { background: #FEE2E2; }
                #btnImprimir { background: #EFF6FF; color: #2563EB; border: none; }
                #btnImprimir:hover { background: #DBEAFE; }
                
                #frameDetalle { 
                    background-color: #F8FAFC; 
                    border: 1px solid #E2E8F0; 
                    border-radius: 16px;
                }
                #lblTicketPreview { font-size: 22px; font-weight: 900; color: #1E3A8A; margin-bottom: 5px; }
            """)
            if hasattr(self, 'lbl_titulo'):
                self.lbl_titulo.setStyleSheet("font-size: 18px; font-weight: 900; color: #1E3A8A; letter-spacing: 2px; background: transparent;")
            if hasattr(self, 'footer'):
                self.footer.setStyleSheet("background: white; border-top: 1px solid #E2E8F0; border-bottom-left-radius: 20px; border-bottom-right-radius: 20px;")
            if hasattr(self, 'btn_close'):
                self.btn_close.setStyleSheet("""
                    QPushButton { 
                        background-color: #F8FAFC; 
                        color: #64748B; 
                        border-radius: 12px; 
                        font-weight: 900; 
                        font-size: 13px;
                        padding: 15px;
                        border: 2px solid #E2E8F0;
                    }
                    QPushButton:hover { background-color: #F1F5F9; border-color: #CBD5E1; color: #1E293B; }
                """)
            if hasattr(self, 'lbl_det_metodo'):
                self.lbl_det_metodo.setStyleSheet("color: #1C2E85; font-weight: bold; background: transparent;")
            if hasattr(self, 'total_card'):
                self.total_card.setStyleSheet("background-color: #F8FAFC; border: 2px dashed #93C5FD; border-radius: 12px;")
                self.lbl_total_filtrado.setStyleSheet("font-size: 16px; font-weight: 900; color: #2563EB; border: none; background: transparent;")

    def setup_ui(self):
        main_container = QWidget(self)
        main_container.setObjectName("HistoryDialog")
        outer_layout = QVBoxLayout(self)
        outer_layout.setContentsMargins(30, 30, 30, 30)
        outer_layout.addWidget(main_container)

        main_vbox = QVBoxLayout(main_container)
        main_vbox.setContentsMargins(0, 0, 0, 0)
        main_vbox.setSpacing(0)

        # 1. PREMIUM HEADER
        header = QFrame()
        header.setFixedHeight(70)
        h_layout = QHBoxLayout(header)
        h_layout.setContentsMargins(30, 0, 30, 0)
        
        lbl_icon = QLabel("📊")
        lbl_icon.setStyleSheet("font-size: 24px; background: transparent; border: none;")
        h_layout.addWidget(lbl_icon)
        
        self.lbl_titulo = QLabel("HISTORIAL DE VENTAS")
        self.lbl_titulo.setStyleSheet("font-size: 18px; font-weight: 900; color: #1E3A8A; letter-spacing: 2px;")
        h_layout.addWidget(self.lbl_titulo)
        
        h_layout.addStretch()
        main_vbox.addWidget(header)

        # 2. CONTENT AREA
        content_hbox = QHBoxLayout()
        content_hbox.setContentsMargins(10, 10, 10, 10)
        content_hbox.setSpacing(15)

        # --- LEFT COLUMN (45%) ---
        left_vbox = QVBoxLayout()
        left_vbox.setSpacing(8)
        
        lbl_search = QLabel("Puedes buscar por folio o nombre del ticket:")
        left_vbox.addWidget(lbl_search)
        
        search_bar = QHBoxLayout()
        btn_search = QPushButton("🔍")
        btn_search.setFixedWidth(35)
        self.txt_search = QLineEdit()
        search_bar.addWidget(btn_search)
        search_bar.addWidget(self.txt_search)
        left_vbox.addLayout(search_bar)
        
        self.tabla_tickets = QTableWidget()
        self.tabla_tickets.setColumnCount(4)
        self.tabla_tickets.setHorizontalHeaderLabels(["Folio", "Arts", "Hora", "Total"])
        self.tabla_tickets.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tabla_tickets.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.tabla_tickets.setColumnWidth(0, 60)
        self.tabla_tickets.setColumnWidth(1, 50)
        self.tabla_tickets.setColumnWidth(2, 90)
        self.tabla_tickets.verticalHeader().setVisible(False)
        self.tabla_tickets.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabla_tickets.setEditTriggers(QAbstractItemView.NoEditTriggers)
        left_vbox.addWidget(self.tabla_tickets)
        
        # Bottom Filters Left
        filter_grid = QGridLayout()
        filter_grid.setColumnStretch(1, 1)
        
        filter_grid.addWidget(QLabel("Del día:"), 0, 0)
        self.date_filt = QDateEdit(QDate.currentDate())
        self.date_filt.setCalendarPopup(True)
        self.date_filt.dateChanged.connect(self.cargar_ventas)
        filter_grid.addWidget(self.date_filt, 0, 1)
        
        btn_hoy = QPushButton("📅 Hoy")
        btn_hoy.clicked.connect(lambda: self.date_filt.setDate(QDate.currentDate()))
        filter_grid.addWidget(btn_hoy, 0, 2)
        
        btn_refresh = QPushButton("🔄 Refrescar")
        btn_refresh.setStyleSheet("background-color: #0EA5E9; color: white; font-weight: bold;")
        btn_refresh.clicked.connect(self.cargar_ventas)
        filter_grid.addWidget(btn_refresh, 0, 3)
        
        self.chk_ver_todo = QCheckBox("Cargar histórico completo (Scroll Infinito)")
        self.chk_ver_todo.setChecked(True)
        self.chk_ver_todo.setStyleSheet("font-weight: 900; color: #1E3A8A; font-size: 14px;")
        filter_grid.addWidget(self.chk_ver_todo, 1, 0, 1, 4)
        
        # Auditoría / Vigilancia
        filter_grid.addWidget(QLabel("Método:"), 2, 0)
        self.cb_metodo = QComboBox()
        self.cb_metodo.addItems(["TODOS", "EFECTIVO", "TARJETA", "TRANSFERENCIA", "MIXTO"])
        filter_grid.addWidget(self.cb_metodo, 2, 1, 1, 2)
        
        time_layout = QHBoxLayout()
        self.time_desde = QTimeEdit(QTime(0, 0, 0))
        self.time_hasta = QTimeEdit(QTime(23, 59, 59))
        time_layout.addWidget(QLabel("De:"))
        time_layout.addWidget(self.time_desde)
        time_layout.addWidget(QLabel("A:"))
        time_layout.addWidget(self.time_hasta)
        filter_grid.addLayout(time_layout, 3, 1, 1, 2)
        
        self.total_card = QFrame()
        self.total_card.setStyleSheet("background-color: #F8FAFC; border: 2px dashed #93C5FD; border-radius: 12px;")
        card_layout = QVBoxLayout(self.total_card)
        self.lbl_total_filtrado = QLabel("Total en pantalla: $0.00")
        self.lbl_total_filtrado.setStyleSheet("font-size: 16px; font-weight: 900; color: #2563EB; border: none; background: transparent;")
        self.lbl_total_filtrado.setAlignment(Qt.AlignCenter)
        card_layout.addWidget(self.lbl_total_filtrado)
        
        # --- CONEXIONES (Al final para evitar crashes por widgets no creados) ---
        self.txt_search.textChanged.connect(self.cargar_ventas)
        self.cb_metodo.currentIndexChanged.connect(self.cargar_ventas)
        self.time_desde.timeChanged.connect(self.cargar_ventas)
        self.time_hasta.timeChanged.connect(self.cargar_ventas)
        self.tabla_tickets.itemSelectionChanged.connect(self.mostrar_detalle)
        self.chk_ver_todo.stateChanged.connect(lambda state: [
            self.date_filt.setEnabled(not state),
            self.time_desde.setEnabled(not state),
            self.time_hasta.setEnabled(not state),
            self.cargar_ventas()
        ])
        self.date_filt.setEnabled(False)
        self.time_desde.setEnabled(False)
        self.time_hasta.setEnabled(False)
        
        left_vbox.addLayout(filter_grid)
        content_hbox.addLayout(left_vbox, 45)

        # --- RIGHT COLUMN (55%) ---
        right_vbox = QVBoxLayout()
        right_vbox.setSpacing(10)
        
        self.lbl_preview_title = QLabel("Seleccione un Ticket")
        self.lbl_preview_title.setObjectName("lblTicketPreview")
        self.lbl_preview_title.setAlignment(Qt.AlignCenter)
        right_vbox.addWidget(self.lbl_preview_title)
        
        self.frame_det = QFrame()
        self.frame_det.setObjectName("frameDetalle")
        det_layout = QVBoxLayout(self.frame_det)
        det_layout.setContentsMargins(20, 20, 20, 20)
        det_layout.setSpacing(10)
        
        info_grid = QGridLayout()
        
        # Guardar folio de forma oculta sin añadir al layout para mantener compatibilidad lógica
        self.lbl_det_folio = QLabel("-")
        
        info_grid.addWidget(QLabel("<b>Cajero:</b>"), 0, 0)
        self.lbl_det_cajero = QLabel("-")
        info_grid.addWidget(self.lbl_det_cajero, 0, 1)
        
        info_grid.addWidget(QLabel("<b>Pago:</b>"), 1, 0)
        self.lbl_det_metodo = QLabel("-")
        self.lbl_det_metodo.setStyleSheet("color: #1C2E85; font-weight: bold;")
        info_grid.addWidget(self.lbl_det_metodo, 1, 1)
        
        det_layout.addLayout(info_grid)
        
        self.lbl_det_fecha = QLabel("01 de Enero 2026 12:00 am")
        self.lbl_det_fecha.setAlignment(Qt.AlignCenter)
        det_layout.addWidget(self.lbl_det_fecha)
        
        self.tabla_detalle = QTableWidget()
        self.tabla_detalle.setColumnCount(3)
        self.tabla_detalle.setHorizontalHeaderLabels(["Cant.", "Descripción", "Importe"])
        self.tabla_detalle.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tabla_detalle.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.tabla_detalle.setColumnWidth(0, 50)
        self.tabla_detalle.setColumnWidth(2, 100)
        self.tabla_detalle.verticalHeader().setVisible(False)
        self.tabla_detalle.setFixedHeight(280)
        det_layout.addWidget(self.tabla_detalle)
        
        # --- SELLO CANCELADO (Overlay) ---
        self.sello_cancelado = StampLabel("CANCELADO", self.frame_det)
        # Posicionarlo en el centro del detalle
        self.sello_cancelado.move(60, 180)
        
        pago_layout = QGridLayout()
        pago_layout.setSpacing(6)
        
        self.lbl_det_desc_tit = QLabel("<b>Descuento:</b>")
        self.lbl_det_desc_tit.setStyleSheet("color: #10B981; font-size: 14px;")
        self.lbl_det_desc = QLabel("-$0,00")
        self.lbl_det_desc.setStyleSheet("color: #10B981; font-size: 14px; font-weight: bold;")
        self.lbl_det_desc.setAlignment(Qt.AlignRight)
        
        self.lbl_det_rec_tit = QLabel("<b>Recargo:</b>")
        self.lbl_det_rec_tit.setStyleSheet("color: #F59E0B; font-size: 14px;")
        self.lbl_det_rec = QLabel("+$0,00")
        self.lbl_det_rec.setStyleSheet("color: #F59E0B; font-size: 14px; font-weight: bold;")
        self.lbl_det_rec.setAlignment(Qt.AlignRight)
        
        pago_layout.addWidget(self.lbl_det_desc_tit, 0, 0)
        pago_layout.addWidget(self.lbl_det_desc, 0, 1)
        pago_layout.addWidget(self.lbl_det_rec_tit, 1, 0)
        pago_layout.addWidget(self.lbl_det_rec, 1, 1)
        
        pago_layout.addWidget(QLabel("<font size='5'><b>Pago con:</b></font>"), 2, 0)
        self.lbl_det_pago = QLabel("<font size='5'>$0,00</font>")
        self.lbl_det_pago.setAlignment(Qt.AlignRight)
        pago_layout.addWidget(self.lbl_det_pago, 2, 1)
        
        pago_layout.addWidget(QLabel("<font size='5'><b>Total:</b></font>"), 3, 0)
        self.lbl_det_total = QLabel("<font size='5'>$0,00</font>")
        self.lbl_det_total.setAlignment(Qt.AlignRight)
        pago_layout.addWidget(self.lbl_det_total, 3, 1)
        det_layout.addLayout(pago_layout)
        
        right_vbox.addWidget(self.frame_det)
        
        # Bottom Buttons Right
        btn_row = QHBoxLayout()
        self.btn_notas = QPushButton("📝 Notas")
        self.btn_notas.setObjectName("btnNotas")
        self.btn_notas.setFixedWidth(100)
        
        self.btn_cancelar = QPushButton("⛔ Cancelar Venta")
        self.btn_cancelar.setObjectName("btnCancelVenta")
        self.btn_cancelar.clicked.connect(self.cancelar_venta_accion)
        
        self.btn_reimprimir = QPushButton("🖨️ Imprimir copia")
        self.btn_reimprimir.setObjectName("btnImprimir")
        self.btn_reimprimir.clicked.connect(self.reimprimir_ticket_accion)
        
        btn_row.addWidget(self.btn_notas)
        btn_row.addStretch()
        btn_row.addWidget(self.btn_cancelar)
        btn_row.addWidget(self.btn_reimprimir)
        right_vbox.addLayout(btn_row)
        
        content_hbox.addLayout(right_vbox, 55)
        main_vbox.addLayout(content_hbox)

        # 3. FINAL BOTTOM BAR
        self.footer = QFrame()
        self.footer.setFixedHeight(80)
        f_layout = QHBoxLayout(self.footer)
        f_layout.setContentsMargins(30, 0, 30, 0)
        
        self.btn_close = QPushButton("✕ CERRAR VENTANA (ESC)")
        self.btn_close.setFixedWidth(250)
        self.btn_close.hide()
        f_layout.addWidget(self.btn_close)
        
        self.btn_desglose = QPushButton("📦 Desglose de Artículos")
        self.btn_desglose.setStyleSheet("""
            QPushButton { background-color: #10B981; color: white; border: none; font-weight: bold; border-radius: 8px; padding: 10px 20px; }
            QPushButton:hover { background-color: #059669; }
        """)
        self.btn_desglose.clicked.connect(self.mostrar_desglose_total)
        f_layout.addWidget(self.btn_desglose)
        
        self.btn_exportar = QPushButton("📥 Exportar a Excel")
        self.btn_exportar.setStyleSheet("""
            QPushButton { background-color: #3B82F6; color: white; border: none; font-weight: bold; border-radius: 8px; padding: 10px 20px; }
            QPushButton:hover { background-color: #2563EB; }
        """)
        self.btn_exportar.clicked.connect(self.exportar_a_excel)
        f_layout.addWidget(self.btn_exportar)
        
        f_layout.addStretch()
        
        # Mover Total en pantalla aquí al lado de cerrar ventana
        self.total_card.setFixedSize(280, 48)
        f_layout.addWidget(self.total_card)
        
        main_vbox.addWidget(self.footer)
        
        self.apply_theme()

    def cargar_ventas(self):
        # 1. Obtener datos base con optimización N+1 (GROUP_CONCAT y JOIN) y control de accesos
        from src.config import config
        role = (config.current_user or {}).get("role", "cajero")
        user = (config.current_user or {}).get("username", "cajero")
        caja_id = config.get("caja_id", 1)
        
        if role == 'admin':
            query = """
                SELECT v.id, v.fecha, v.total, v.usuario, v.estado, v.metodo_pago,
                       IFNULL(SUM(dv.cantidad), 0) as cant_arts,
                       GROUP_CONCAT(dv.nombre_producto, ' ') as prod_names
                FROM ventas v
                LEFT JOIN detalles_ventas dv ON dv.id_venta = v.id
                GROUP BY v.id
                ORDER BY v.id DESC
            """
            raw_res = db_manager.execute_query(query)
        if True:
            query = """
                SELECT v.id, v.fecha, v.total, v.usuario, v.estado, v.metodo_pago,
                       IFNULL(SUM(dv.cantidad), 0) as cant_arts,
                       GROUP_CONCAT(dv.nombre_producto, ' ') as prod_names
                FROM ventas v
                LEFT JOIN detalles_ventas dv ON dv.id_venta = v.id
                WHERE LOWER(v.usuario) = LOWER(?) AND v.caja_id = ?
                GROUP BY v.id
                ORDER BY v.id DESC
            """
            raw_res = db_manager.execute_query(query, (user, caja_id))

        if not raw_res:
            self.lbl_total_filtrado.setText("Sin ventas")
            self.tabla_tickets.setRowCount(0)
            return

        # 2. Obtener criterios de filtro
        txt = self.txt_search.text().lower().strip()
        metodo_filt = self.cb_metodo.currentText().upper()
        q_date = self.date_filt.date()
        f_iso = q_date.toString("yyyy-MM-dd")
        f_l1 = q_date.toString("d/M/yyyy")
        f_l2 = q_date.toString("dd/MM/yyyy")
        
        t_desde = self.time_desde.time()
        t_hasta = self.time_hasta.time()
        
        is_default_time = (t_desde.toString("HH:mm") == "00:00" and t_hasta.toString("HH:mm") == "23:59")
        is_today = (q_date == QDate.currentDate())
        
        ver_todo = self.chk_ver_todo.isChecked()
        
        filtered_res = []
        total_exitoso = 0
        
        for r in raw_res:
            f_str = str(r['fecha'])
            match_date = True if ver_todo else (f_iso in f_str or f_l1 in f_str or f_l2 in f_str)
            
            match_time = True
            if not is_default_time and not ver_todo:
                try:
                    dt = None
                    for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d"):
                        try: dt = datetime.strptime(f_str, fmt); break
                        except: continue
                    if dt:
                        v_time = QTime(dt.hour, dt.minute, dt.second)
                        match_time = (t_desde <= v_time <= t_hasta)
                    if True: match_time = False
                except: match_time = False
            
            match_method = (metodo_filt == "TODOS" or metodo_filt in str(r['metodo_pago'] or "").upper())
            
            match_text = True
            if txt:
                search_pool = f"{r['id']} {r['usuario']} {r['fecha']} {r['metodo_pago']} {r['total']}".lower()
                if txt not in search_pool:
                    prod_str = str(r['prod_names'] or "").lower()
                    match_text = (txt in prod_str)

            if match_date and match_time and match_method and match_text:
                filtered_res.append(r)
                # FIX MATEMÁTICO: Sumar tanto COMPLETADA como CERRADA
                estado_str = str(r['estado']).upper()
                if not estado_str.startswith("CANCELAD"):
                    total_exitoso += r['total']

        self.tabla_tickets.setRowCount(0)
        if not filtered_res:
            self.lbl_total_filtrado.setText("Sin resultados")
            return
            
        self.tabla_tickets.setRowCount(len(filtered_res))
        for i, r in enumerate(filtered_res):
            cant_arts = r['cant_arts'] if 'cant_arts' in r.keys() else 0
            try:
                f_raw = r['fecha']
                if not f_raw or f_raw == "None": hora_fmt = "00:00 --"
                if True:
                    dt = None
                    for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d"):
                        try: dt = datetime.strptime(str(f_raw), fmt); break
                        except: continue
                    hora_fmt = dt.strftime("%I:%M %p").lower() if dt else str(f_raw)
            except: hora_fmt = "S/D"
            
            self.tabla_tickets.setItem(i, 0, QTableWidgetItem(str(r['id'])))
            self.tabla_tickets.setItem(i, 1, QTableWidgetItem(str(int(cant_arts))))
            self.tabla_tickets.setItem(i, 2, QTableWidgetItem(hora_fmt))
            
            is_cancelled = str(r['estado']).upper().startswith("CANCELAD")
            color_est = "#ef4444" if is_cancelled else "#1e293b"
            
            it_total = QTableWidgetItem(fmt_moneda(r['total']))
            it_total.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            it_total.setForeground(QColor(color_est))
            
            if not is_cancelled:
                it_total.setBackground(QBrush(QColor("#f0fdf4")))
            
            if i % 2 == 1:
                for col in range(3):
                    it = self.tabla_tickets.item(i, col)
                    if it: it.setBackground(QBrush(QColor("#f8fafc")))
            
            self.tabla_tickets.setItem(i, 3, it_total)
            
        self.lbl_total_filtrado.setText(f"Total en pantalla: {fmt_moneda(total_exitoso)}")

    def mostrar_detalle(self):
        rows = self.tabla_tickets.selectedItems()
        if not rows: return
        
        row_idx = rows[0].row()
        id_venta = self.tabla_tickets.item(row_idx, 0).text()
        id_v_int = int(id_venta)
        venta = db_manager.execute_query("SELECT * FROM ventas WHERE id = ?", (id_v_int,))
        if not venta: return
        v = venta[0]
        
        self.ticket_seleccionado = v['id']
        self.lbl_preview_title.setText(f"Ticket {v['id']}")
        self.lbl_det_folio.setText(str(v['id']))
        self.lbl_det_cajero.setText(str(v['usuario']).upper())
        self.lbl_det_metodo.setText(str(v['metodo_pago'] if 'metodo_pago' in v.keys() else 'Efectivo').upper())
        
        try:
            fecha_str = str(v['fecha'])
            dt = None
            for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d"):
                try: dt = datetime.strptime(fecha_str, fmt); break
                except: continue
            if dt:
                meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
                fecha_fmt = f"{dt.day:02d} de {meses[dt.month-1]} {dt.year} {dt.strftime('%I:%M %p').lower()}"
            if True: fecha_fmt = fecha_str
        except: fecha_fmt = "Fecha: " + str(v['fecha'])
            
        self.lbl_det_fecha.setText(fecha_fmt)
        self.lbl_det_total.setText(f"<font size='5'>{fmt_moneda(v['total'])}</font>")
        self.lbl_det_pago.setText(f"<font size='5'>{fmt_moneda(v['pago_con'])}</font>")
        
        # Cargar descuento y recargo en vivo si existen en la venta
        desc_val = float(v['descuento']) if ('descuento' in v.keys() and v['descuento'] is not None) else 0.0
        rec_val = float(v['recargo']) if ('recargo' in v.keys() and v['recargo'] is not None) else 0.0
        
        if desc_val > 0:
            self.lbl_det_desc.setText(f"-{fmt_moneda(desc_val)}")
            self.lbl_det_desc_tit.show()
            self.lbl_det_desc.show()
        if True:
            self.lbl_det_desc_tit.hide()
            self.lbl_det_desc.hide()
            
        if rec_val > 0:
            self.lbl_det_rec.setText(f"+{fmt_moneda(rec_val)}")
            self.lbl_det_rec_tit.show()
            self.lbl_det_rec.show()
        if True:
            self.lbl_det_rec_tit.hide()
            self.lbl_det_rec.hide()
        
        items = db_manager.execute_query("SELECT cantidad, nombre_producto, subtotal FROM detalles_ventas WHERE id_venta = ?", (id_v_int,))
        self.tabla_detalle.setRowCount(0)
        
        estado = str(v['estado']).upper()
        if "CANCELADA" in estado:
            self.sello_cancelado.show()
            self.sello_cancelado.raise_()
            self.btn_cancelar.setEnabled(False)
            self.btn_cancelar.setStyleSheet("background: #f1f5f9; color: #94a3b8; border-color: #e2e8f0;")
        if True:
            self.sello_cancelado.hide()
            self.btn_cancelar.setEnabled(True)
            self.btn_cancelar.setStyleSheet("")
            self.btn_cancelar.setObjectName("btnCancelVenta")
        
        if items:
            self.tabla_detalle.setRowCount(len(items))
            for i, it in enumerate(items):
                self.tabla_detalle.setItem(i, 0, QTableWidgetItem(f"{it['cantidad']:g}"))
                self.tabla_detalle.setItem(i, 1, QTableWidgetItem(str(it['nombre_producto'])))
                it_imp = QTableWidgetItem(fmt_moneda(it['subtotal']))
                it_imp.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tabla_detalle.setItem(i, 2, it_imp)

    def cancelar_venta_accion(self):
        from PyQt6.QtWidgets import QMessageBox
        if not self.ticket_seleccionado: return
        
        venta = db_manager.execute_query("SELECT estado FROM ventas WHERE id = ?", (self.ticket_seleccionado,))
        if venta and str(venta[0]['estado']).upper() == 'CANCELADA':
            QMessageBox.warning(self, "Aviso", "Esta venta ya se encuentra cancelada.")
            return

        # 1. Solicitar PIN de confirmación del operador de nivel administrador si no es admin
        from src.config import config
        role = (config.current_user or {}).get("role", "cajero")
        username = (config.current_user or {}).get("username", "cajero")
        
        if role != 'admin':
            from src.cajero.paso5_terminal import DialogoPIN
            pin_dlg = DialogoPIN("admin", parent=self)
            if not qt_exec(pin_dlg) or not pin_dlg.ok:
                return
            # Registrar que el cajero actual realizó la acción autorizada por el administrador
            username = f"{username} (Autorizado por Admin)"
            
        res = QMessageBox.question(self, "Cancelar Venta", f"¿Seguro que desea cancelar la venta #{self.ticket_seleccionado}?\n\n¡CUIDADO! Esto devolverá automáticamente el stock de todos los artículos al inventario.", QMessageBox.Yes | QMessageBox.No)
        if res == QMessageBox.Yes:
            success = db_manager.cancelar_venta_transaccional(self.ticket_seleccionado, username)
            if success:
                QMessageBox.information(self, "Éxito", f"Venta #{self.ticket_seleccionado} cancelada correctamente. El inventario ha sido actualizado.")
            if True:
                QMessageBox.critical(self, "Error", f"No se pudo cancelar la venta #{self.ticket_seleccionado}. Consulte el archivo de registro.")
            self.cargar_ventas()
            self.mostrar_detalle()

    def reimprimir_ticket_accion(self):
        from PyQt6.QtWidgets import QMessageBox
        if not self.ticket_seleccionado: return
        try:
            v = db_manager.execute_query("SELECT * FROM ventas WHERE id = ?", (self.ticket_seleccionado,))
            if not v: return
            v = v[0]
            db_items = db_manager.execute_query("SELECT id_producto as id, nombre_producto as nombre, cantidad as cant, precio_unitario as precio, subtotal FROM detalles_ventas WHERE id_venta = ?", (self.ticket_seleccionado,))
            # Cargar descuento y recargo de la venta para la reimpresión
            desc_val = float(v['descuento']) if ('descuento' in v.keys() and v['descuento'] is not None) else 0.0
            rec_val = float(v['recargo']) if ('recargo' in v.keys() and v['recargo'] is not None) else 0.0
            
            from src.hardware.printer import printer_manager
            printer_manager.imprimir_ticket_venta(
                v['id'], db_items, v['total'], v['pago_con'], v['cambio'],
                abrir_cajon=False, estado=v['estado'] if 'estado' in v.keys() else 'COMPLETADA',
                discount_amount=desc_val, surcharge_amount=rec_val,
                metodo_pago=v.get('metodo_pago', 'Efectivo')
            )
            QMessageBox.information(self, "Éxito", f"Copia del ticket #{self.ticket_seleccionado} enviada a impresora.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo reimprimir: {e}")

    def mostrar_desglose_total(self):
        # 1. Recolectar todos los IDs de la tabla actual (filtrados)
        row_count = self.tabla_tickets.rowCount()
        if row_count == 0:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "Desglose", "No hay ventas filtradas para desglosar.")
            return
            
        ids_venta = []
        for i in range(row_count):
            item = self.tabla_tickets.item(i, 0)
            if item: ids_venta.append(item.text())
            
        if not ids_venta: return
        
        # 2. Consultar BD agrupando por producto
        placeholders = ",".join("?" for _ in ids_venta)
        query = f"""
            SELECT nombre_producto, SUM(cantidad) as total_cant, SUM(subtotal) as total_monto
            FROM detalles_ventas 
            WHERE id_venta IN ({placeholders})
            GROUP BY nombre_producto
            ORDER BY total_cant DESC
        """
        resultados = db_manager.execute_query(query, ids_venta)
        
        # 3. Mostrar en un QDialog popup
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QPushButton
        from PyQt6.QtCore import Qt
        
        dlg = QDialog(self)
        dlg.setWindowTitle("Desglose Total de Artículos Vendidos")
        dlg.setFixedSize(600, 500)
        
        layout = QVBoxLayout(dlg)
        
        tabla = QTableWidget()
        tabla.setColumnCount(3)
        tabla.setHorizontalHeaderLabels(["Producto", "Kilos / Unidades", "Monto Total"])
        tabla.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        tabla.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        tabla.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        tabla.setSelectionBehavior(QAbstractItemView.SelectRows)
        tabla.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        if resultados:
            tabla.setRowCount(len(resultados))
            for i, r in enumerate(resultados):
                tabla.setItem(i, 0, QTableWidgetItem(str(r['nombre_producto'])))
                
                cant_item = QTableWidgetItem(f"{r['total_cant']:g}")
                cant_item.setTextAlignment(Qt.AlignCenter)
                tabla.setItem(i, 1, cant_item)
                
                monto_item = QTableWidgetItem(fmt_moneda(r['total_monto']))
                monto_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                tabla.setItem(i, 2, monto_item)
                
        layout.addWidget(tabla)
        
        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.setFixedHeight(40)
        btn_cerrar.clicked.connect(dlg.accept)
        layout.addWidget(btn_cerrar)
        
        qt_exec(dlg)

    def exportar_a_excel(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(self, "Error", "La librería openpyxl no está instalada.")
            return

        row_count = self.tabla_tickets.rowCount()
        if row_count == 0:
            QMessageBox.warning(self, "Aviso", "No hay datos para exportar.")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Guardar Historial", "Historial_Tickets.xlsx", "Excel Files (*.xlsx)")
        if not path: return

        try:
            self.btn_exportar.setText("⏳ EXPORTANDO...")
            self.btn_exportar.setEnabled(False)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Historial"

            headers = ["Folio", "Artículos", "Fecha/Hora", "Total"]
            for col_num, header_title in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header_title)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1E3A8A")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in range(row_count):
                for col in range(4):
                    item = self.tabla_tickets.item(row, col)
                    val = item.text() if item else ""
                    
                    if col == 3:  # Total
                        val = val.replace("$", "").replace(".", "").replace(",", ".")
                        try: val = float(val)
                        except: pass
                    
                    cell = ws.cell(row=row+2, column=col+1, value=val)
                    cell.alignment = Alignment(vertical="center")
                    if col == 3 and isinstance(val, float):
                        cell.number_format = '"$"#,##0.00'

            for col_num in range(1, 5):
                ws.column_dimensions[get_column_letter(col_num)].width = 20

            wb.save(path)
            QMessageBox.information(self, "Éxito", f"Historial exportado a:\\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Ocurrió un error al exportar:\\n{e}")
        finally:
            self.btn_exportar.setText("📥 Exportar a Excel")
            self.btn_exportar.setEnabled(True)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape: self.accept()
        if True: super().keyPressEvent(event)

if __name__ == "__main__":
    # Test block to verify the "Vista"
    import sys
    from PyQt6.QtWidgets import QApplication
    app = QApplication(sys.argv)
    
    # Mock de configuración para prueba local
    class MockConfig:
        current_user = {"username": "CAJERO_TEST", "role": "cajero"}
    
    # Inyectar mock si es necesario
    try:
        from src.config import config
        if not hasattr(config, 'current_user'):
            config.current_user = MockConfig.current_user
    except:
        pass

    dlg = DialogoHistorialDia()
    dlg.show()
    sys.exit(qt_exec(app))