from src.utils.qt_compat import qt_exec
from src.utils.theme_manager import theme_manager

import json
from PyQt6.QtWidgets import (

    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame, QPushButton, 
    QScrollArea, QGridLayout, QStackedWidget,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QComboBox, QLineEdit, QFileDialog, QMessageBox, QDialog
)
from PyQt6.QtCore import Qt, pyqtSignal, QRect, QPoint, QSize, QThread, QUrl
from PyQt6.QtGui import QColor, QFont, QPainter, QBrush, QPen, QLinearGradient, QPolygon, QPainterPath
import datetime

try:
    from src.base_de_datos.database import db_manager
except ImportError:
    from database import db_manager

# Paleta 2026 — claros, contraste marcado, profundidad 3D suave
_FIN = {
    "bg_page": "#F0F4F8",
    "card": "#FFFFFF",
    "card_border": "#E2E8F0",
    "row_white": "#FFFFFF",
    "row_gray": "#F1F5F9",
    "text_on_white": "#0F172A",
    "text_on_gray": "#1E293B",
    "muted_on_white": "#64748B",
    "muted_on_gray": "#475569",
    "text": "#0F172A",
    "text_soft": "#475569",
    "text_muted": "#94A3B8",
    "header_bg": "#E8EEF4",
    "header_text": "#334155",
    "accent": "#3B82F6",
    "accent_light": "#EFF6FF",
    "accent_hover": "#2563EB",
    "border_row": "#E2E8F0",
    "green": "#10B981",
    "red": "#EF4444",
}

_KEEP_FG = frozenset({"#10b981", "#059669", "#ef4444", "#dc2626"})


def _estilo_tabla_financiera() -> str:
    return f"""
        QTableWidget {{
            background-color: {_FIN['row_white']};
            border: 1px solid {_FIN['card_border']};
            border-radius: 12px;
            gridline-color: transparent;
            font-size: 13px;
            color: {_FIN['text_on_white']};
            outline: none;
        }}
        QTableWidget::item {{
            padding: 11px 10px;
            border-bottom: 1px solid {_FIN['border_row']};
        }}
        QTableWidget::item:selected {{
            background-color: {_FIN['accent_light']};
            color: #1E40AF;
        }}
        QHeaderView::section {{
            background-color: {_FIN['header_bg']};
            color: {_FIN['header_text']};
            font-weight: 800;
            font-size: 11px;
            padding: 12px 10px;
            border: none;
            border-bottom: 2px solid {_FIN['accent']};
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
    """


def _aplicar_paleta_tabla(tabla: QTableWidget):
    from PyQt6.QtGui import QPalette
    tabla.setAlternatingRowColors(False)
    pal = tabla.palette()
    pal.setColor(QPalette.ColorRole.Base, QColor(_FIN["row_white"]))
    pal.setColor(QPalette.ColorRole.AlternateBase, QColor(_FIN["row_gray"]))
    pal.setColor(QPalette.ColorRole.Text, QColor(_FIN["text_on_white"]))
    pal.setColor(QPalette.ColorRole.Window, QColor(_FIN["row_white"]))
    tabla.setPalette(pal)
    tabla.setStyleSheet(_estilo_tabla_financiera())
    tabla.viewport().setStyleSheet(f"background-color: {_FIN['row_white']};")


def _pintar_filas_tabla(tabla: QTableWidget):
    """Filas alternadas blanco / gris claro — texto oscuro bien marcado en ambas."""
    for row in range(tabla.rowCount()):
        es_gris = (row % 2) == 1
        bg = QColor(_FIN["row_gray"] if es_gris else _FIN["row_white"])
        fg = QColor(_FIN["text_on_gray"] if es_gris else _FIN["text_on_white"])
        fg_muted = QColor(_FIN["muted_on_gray"] if es_gris else _FIN["muted_on_white"])
        for col in range(tabla.columnCount()):
            item = tabla.item(row, col)
            if not item:
                continue
            item.setBackground(bg)
            cur = item.foreground().color().name().lower()
            if cur in _KEEP_FG:
                continue
            if cur in ("#94a3b8", "#64748b", "#9ca3af", "#cbd5e1"):
                item.setForeground(fg_muted)
            else:
                item.setForeground(fg)

def get_depto_icon(depto_name):
    if not depto_name:
        return "📦"
    name = depto_name.strip().upper()
    if "CARNE" in name or "VACUNO" in name or "RES" in name or "CERDO" in name or "VACUN" in name or "ASADO" in name or "TERNER" in name:
        return "🥩"
    if "AVE" in name or "POLLO" in name or "GRANJA" in name or "POLLER" in name:
        return "🍗"
    if "ACHURA" in name or "CHINCHU" in name or "MENUDE" in name or "RIÑON" in name or "MOLLEJ" in name or "INTESTI" in name:
        return "🍢"
    if "PREPARADO" in name or "ELABORADO" in name or "HAMBUR" in name or "MILANE" in name or "ROTIS" in name:
        return "🍳"
    if "EMBUTIDO" in name or "FIAMBRE" in name or "SALCHI" in name or "CHORI" in name or "JAMON" in name or "SALA" in name or "CHARCU" in name:
        return "🌭"
    if "ALMACEN" in name or "ALMACÉN" in name or "ABARRO" in name or "DESPEN" in name:
        return "🥫"
    if "BEBIDA" in name or "REFRES" in name or "GASEO" in name or "CERVE" in name or "VINO" in name or "TRAGO" in name:
        return "🥤"
    if "VERDU" in name or "FRUTA" in name or "VEGETA" in name or "HORTE" in name:
        return "🥦"
    if "PANAD" in name or "PAN" in name or "FACTU" in name or "FACTUR" in name or "BIZCO" in name:
        return "🍞"
    if "LACTEO" in name or "LÁCTEO" in name or "QUESO" in name or "LECHE" in name or "MANTE" in name or "YOGU" in name:
        return "🧀"
    if "LIMPIE" in name or "HIGIEN" in name or "JABON" in name or "DETER" in name or "PERFU" in name:
        return "🧼"
    if "CONGEL" in name or "HELA" in name:
        return "❄️"
    if "KIOS" in name or "GOLO" in name or "CARAME" in name or "CHOCO" in name:
        return "🍬"
    return "📦"

class ModernCard(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("card")
        self.setStyleSheet(f"""
            #card {{
                background: {_FIN['card']};
                border: 1px solid {_FIN['card_border']};
                border-radius: 18px;
            }}
        """)


class StockAreaChartWidget(QWidget):
    def __init__(self, data=None, parent=None):
        super().__init__(parent)
        self.data = data or {}
        self.data_prev = None
        self.setMinimumHeight(380)
        self.setAttribute(Qt.WA_Hover, True)
        self.hover_index = -1
        self.setMouseTracking(True)
        
    def update_data(self, data, data_prev=None):
        self.data = data
        self.data_prev = data_prev
        self.update()
        
    def mouseMoveEvent(self, event):
        if not self.data: return
        w = self.width()
        padding_l, padding_r = 60, 20
        chart_w = w - padding_l - padding_r
        keys = list(self.data.keys())
        if not keys: return
        step = chart_w / max(1, len(keys) - 1)
        x_pos = event.pos().x() - padding_l
        idx = int(round(x_pos / step))
        if 0 <= idx < len(keys):
            if self.hover_index != idx:
                self.hover_index = idx
                self.update()
        else:
            if self.hover_index != -1:
                self.hover_index = -1
                self.update()
                
    def leaveEvent(self, event):
        self.hover_index = -1
        self.update()

    def paintEvent(self, event):
        from PyQt6.QtGui import QPainter, QColor, QFont, QPen, QBrush, QPainterPath, QLinearGradient
        from PyQt6.QtCore import QPointF, QRect, Qt
        
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        w, h = self.width(), self.height()
        padding_l, padding_r, padding_t, padding_b = 60, 30, 30, 40
        chart_w, chart_h = w - padding_l - padding_r, h - padding_t - padding_b
        
        # Background lines
        painter.setPen(QPen(QColor("#EEF2F8"), 1, Qt.DashLine))
        painter.setFont(QFont("Segoe UI", 8))
        
        if not self.data:
            painter.drawText(self.rect(), Qt.AlignCenter, "Sin datos")
            return
            
        max_val = max(max([d.get('ventas', 0) for d in self.data.values()]), 1)
        if self.data_prev:
            max_prev = max(max([d.get('ventas', 0) for d in self.data_prev.values()]), 1)
            max_val = max(max_val, max_prev)
            
        magnitude = 10**(len(str(int(max_val))) - 1) if max_val >= 10 else 1
        max_val = ((int(max_val) // magnitude) + 1) * magnitude
        
        for i in range(6):
            y = h - padding_b - (i * chart_h / 5)
            painter.drawLine(padding_l, int(y), w - padding_r, int(y))
            val_text = f"${(max_val * i / 5):,.0f}"
            painter.setPen(QColor("#94A3B8"))
            painter.drawText(QRect(0, int(y) - 10, padding_l - 10, 20), Qt.AlignRight | Qt.AlignVCenter, val_text)
            painter.setPen(QPen(QColor("#EEF2F8"), 1, Qt.DashLine))
            
        keys = list(self.data.keys())
        if not keys:
            return
        step = chart_w / max(1, len(keys) - 1)
        
        # X labels
        painter.setPen(QColor("#94A3B8"))
        for i, key in enumerate(keys):
            if len(keys) > 15 and i % 2 != 0: continue
            x = padding_l + i * step
            painter.drawText(QRect(int(x) - 30, h - padding_b + 10, 60, 20), Qt.AlignCenter, str(key)[:5])
            
        # Draw Previous Ventas Area
        if self.data_prev:
            pts_prev = []
            for i, key in enumerate(keys):
                x = padding_l + i * step
                v = self.data_prev.get(key, {}).get('ventas', 0)
                y = h - padding_b - (v / max_val) * chart_h
                pts_prev.append(QPointF(x, y))
                
            if pts_prev:
                path_prev = QPainterPath()
                path_prev.moveTo(pts_prev[0])
                for i in range(1, len(pts_prev)):
                    p1 = pts_prev[i - 1]
                    p2 = pts_prev[i]
                    c1 = QPointF((p1.x() + p2.x()) / 2, p1.y())
                    c2 = QPointF((p1.x() + p2.x()) / 2, p2.y())
                    path_prev.cubicTo(c1, c2, p2)
                    
                pen_prev = QPen(QColor("#94A3B8"), 3, Qt.DashLine)
                painter.setPen(pen_prev)
                painter.setBrush(Qt.NoBrush)
                painter.drawPath(path_prev)

        # Draw Current Ventas Area
        pts_v = []
        for i, key in enumerate(keys):
            x = padding_l + i * step
            y = h - padding_b - (self.data[key].get('ventas', 0) / max_val) * chart_h
            pts_v.append(QPointF(x, y))

        if not pts_v:
            return
            
        path_v = QPainterPath()
        path_v.moveTo(pts_v[0])
        if len(pts_v) == 1:
            painter.setPen(QPen(QColor(59, 130, 246), 4))
            painter.setBrush(QColor(59, 130, 246))
            painter.drawEllipse(pts_v[0], 5, 5)
        else:
            for i in range(1, len(pts_v)):
                p1 = pts_v[i - 1]
                p2 = pts_v[i]
                c1 = QPointF((p1.x() + p2.x()) / 2, p1.y())
                c2 = QPointF((p1.x() + p2.x()) / 2, p2.y())
                path_v.cubicTo(c1, c2, p2)
                
            fill_path = QPainterPath(path_v)
            fill_path.lineTo(padding_l + (len(keys) - 1) * step, h - padding_b)
            fill_path.lineTo(padding_l, h - padding_b)
            fill_path.closeSubpath()
            
            grad_v = QLinearGradient(0, padding_t, 0, h - padding_b)
            grad_v.setColorAt(0, QColor(59, 130, 246, 100))
            grad_v.setColorAt(1, QColor(59, 130, 246, 0))
            painter.fillPath(fill_path, QBrush(grad_v))
            
            painter.setPen(QPen(QColor(59, 130, 246), 4))
            painter.drawPath(path_v)
        
        # Hover
        if self.hover_index != -1 and self.hover_index < len(keys):
            x = padding_l + self.hover_index * step
            painter.setPen(QPen(QColor("#CBD5E1"), 1, Qt.DashLine))
            painter.drawLine(int(x), padding_t, int(x), h - padding_b)
            
            v_val = self.data[keys[self.hover_index]].get('ventas', 0)
            y_v = h - padding_b - (v_val / max_val) * chart_h
            
            painter.setPen(QPen(QColor(255, 255, 255), 2))
            painter.setBrush(QColor(59, 130, 246))
            painter.drawEllipse(QPointF(x, y_v), 6, 6)
            
            if self.data_prev:
                v_prev = self.data_prev.get(keys[self.hover_index], {}).get('ventas', 0)
                y_prev = h - padding_b - (v_prev / max_val) * chart_h
                painter.setBrush(QColor("#94A3B8"))
                painter.drawEllipse(QPointF(x, y_prev), 5, 5)
                
            # Tooltip
            tt_w = 260
            tt_h = 100 if self.data_prev else 60
            tt_rect = QRect(int(x) - tt_w//2, padding_t - 20, tt_w, tt_h)
            if x + tt_w//2 > w: tt_rect.moveLeft(w - tt_w - 10)
            if x - tt_w//2 < padding_l: tt_rect.moveLeft(padding_l + 10)
            
            painter.setBrush(QColor(255, 255, 255, 250))
            painter.setPen(QPen(QColor("#E2E8F0"), 1))
            painter.drawRoundedRect(tt_rect, 8, 8)
            painter.setPen(QColor("#1E293B"))
            painter.setFont(QFont("Segoe UI", 10, QFont.Bold))
            painter.drawText(QRect(tt_rect.x(), tt_rect.y()+5, tt_rect.width(), 20), Qt.AlignCenter, f"{keys[self.hover_index]}")
            
            painter.setFont(QFont("Segoe UI", 9))
            painter.setPen(QColor(59, 130, 246))
            painter.drawText(QRect(tt_rect.x()+10, tt_rect.y()+30, tt_rect.width()-20, 20), Qt.AlignLeft, "Actual:")
            painter.drawText(QRect(tt_rect.x()+10, tt_rect.y()+30, tt_rect.width()-20, 20), Qt.AlignRight, f"${v_val:,.0f}")
            
            if self.data_prev:
                painter.setPen(QColor("#94A3B8"))
                painter.drawText(QRect(tt_rect.x()+10, tt_rect.y()+50, tt_rect.width()-20, 20), Qt.AlignLeft, "Anterior:")
                painter.drawText(QRect(tt_rect.x()+10, tt_rect.y()+50, tt_rect.width()-20, 20), Qt.AlignRight, f"${v_prev:,.0f}")
                
                diff = v_val - v_prev
                if v_prev > 0:
                    pct = (diff / v_prev) * 100
                else:
                    pct = 100.0 if diff > 0 else 0.0
                
                rect_diff = QRect(tt_rect.x()+10, tt_rect.y()+72, tt_rect.width()-20, 20)
                painter.setFont(QFont("Segoe UI", 9, QFont.Bold))
                if diff > 0:
                    painter.setPen(QColor("#10B981"))
                    painter.drawText(rect_diff, Qt.AlignCenter, f"▲ +${diff:,.0f} (+{pct:.1f}%)")
                elif diff < 0:
                    painter.setPen(QColor("#EF4444"))
                    painter.drawText(rect_diff, Qt.AlignCenter, f"▼ -${abs(diff):,.0f} ({pct:.1f}%)")
                else:
                    painter.setPen(QColor("#94A3B8"))
                    painter.drawText(rect_diff, Qt.AlignCenter, "Sin cambios")


class BarChartWidget(QWidget):
    def __init__(self, data=None, parent=None):
        super().__init__(parent)
        self.data = data or {}
        self.colors = ["#475569", "#64748B", "#94A3B8", "#CBD5E1", "#E2E8F0", "#F1F5F9"]
        self.setMinimumHeight(350)
        self.setAttribute(Qt.WA_Hover, True)
        self.hover_index = -1
        self.setMouseTracking(True)
        
    def update_data(self, data):
        self.data = data
        self.update()

    def mouseMoveEvent(self, event):
        if not self.data: return
        w = self.width()
        padding_l, padding_r = 80, 40
        chart_w = w - padding_l - padding_r
        days = list(self.data.keys())
        if not days: return
        spacing = chart_w / len(days)
        x_pos = event.pos().x() - padding_l
        idx = int(x_pos / spacing)
        if 0 <= idx < len(days):
            if self.hover_index != idx:
                self.hover_index = idx
                self.update()
        else:
            if self.hover_index != -1:
                self.hover_index = -1
                self.update()

    def leaveEvent(self, event):
        self.hover_index = -1
        self.update()

    def paintEvent(self, event):
        from PyQt6.QtGui import QPainter, QColor, QFont, QPen, QBrush
        from PyQt6.QtCore import QPoint, QRect, Qt
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        w = self.width()
        h = self.height()
        padding_l, padding_r, padding_t, padding_b = 80, 40, 40, 60
        chart_w, chart_h = w - padding_l - padding_r, h - padding_t - padding_b
        
        max_total = 0
        for day, methods in self.data.items():
            max_total = max(max_total, sum(methods.values()))
        if max_total == 0: max_total = 1000
        magnitude = 10**(len(str(int(max_total))) - 1) if max_total >= 10 else 1
        max_total = ((int(max_total) // magnitude) + 1) * magnitude
        
        # Grid lines
        painter.setPen(QPen(QColor("#F1F5F9"), 1, Qt.DashLine))
        painter.setFont(QFont("Segoe UI", 9))
        for i in range(6):
            y = h - padding_b - (i * chart_h / 5)
            painter.drawLine(padding_l, int(y), w - padding_r, int(y))
            val = (max_total / 5) * i
            painter.setPen(QColor("#94A3B8"))
            painter.drawText(QRect(0, int(y - 10), padding_l - 10, 20), Qt.AlignRight | Qt.AlignVCenter, f"${val:,.0f}")
            painter.setPen(QPen(QColor("#F1F5F9"), 1, Qt.DashLine))

        days = list(self.data.keys())
        if not days: return
        bar_w = min(40, (chart_w / len(days)) * 0.6)
        spacing = chart_w / len(days)
        
        methods_list = ["Efectivo", "Tarjeta", "Transferencia", "Vales", "Crédito", "Cheque"]
        
        for i, day in enumerate(days):
            x = padding_l + (i * spacing) + (spacing - bar_w) / 2
            current_y = h - padding_b
            
            day_values = []
            for j, m in enumerate(methods_list):
                val = self.data[day].get(m, 0)
                if val > 0: day_values.append((j, m, val))
                
            for j, m, value in day_values:
                bar_h = (value / max_total) * chart_h
                
                rect_y = current_y - bar_h
                # Gap between segments
                if current_y < h - padding_b:
                    rect_y -= 2
                    
                color = self.colors[j % len(self.colors)]
                if self.hover_index != -1 and self.hover_index != i:
                    # dim non-hovered
                    color = "#E2E8F0"
                
                painter.setBrush(QColor(color))
                painter.setPen(Qt.NoPen)
                painter.drawRoundedRect(int(x), int(rect_y), int(bar_w), int(bar_h), 4, 4)
                
                current_y = rect_y
                
            # X Label
            painter.setPen(QColor("#64748B"))
            painter.setFont(QFont("Segoe UI", 9, QFont.Bold))
            painter.drawText(QRect(int(x - 20), int(h - padding_b + 10), int(bar_w + 40), 30), Qt.AlignCenter, str(day)[:5])

        # Total floating label
        total_val = sum([sum(self.data[d].values()) for d in days])
        painter.setFont(QFont("Segoe UI", 16, QFont.Bold))
        painter.setPen(QColor("#64748B"))
        painter.drawText(QRect(w - 250, 10, 230, 30), Qt.AlignRight | Qt.AlignVCenter, f"${total_val:,.2f}")
        
        # Hover Tooltip
        if self.hover_index != -1 and self.hover_index < len(days):
            day = days[self.hover_index]
            tt_w, tt_h = 220, 30 + len(methods_list)*20
            x = padding_l + (self.hover_index * spacing) + spacing/2
            y = padding_t
            
            tt_rect = QRect(int(x) - tt_w//2, int(y), tt_w, tt_h)
            if tt_rect.right() > w: tt_rect.moveRight(w - 10)
            if tt_rect.left() < padding_l: tt_rect.moveLeft(padding_l + 10)
            
            painter.setBrush(QColor(255, 255, 255, 245))
            painter.setPen(QPen(QColor("#E2E8F0"), 1))
            painter.drawRoundedRect(tt_rect, 8, 8)
            
            painter.setPen(QColor("#1E293B"))
            painter.setFont(QFont("Segoe UI", 10, QFont.Bold))
            painter.drawText(QRect(tt_rect.x(), tt_rect.y()+5, tt_rect.width(), 20), Qt.AlignCenter, str(day))
            
            painter.setFont(QFont("Segoe UI", 9))
            cy = tt_rect.y() + 30
            for j, m in enumerate(methods_list):
                val = self.data[day].get(m, 0)
                if val > 0:
                    painter.setBrush(QColor(self.colors[j % len(self.colors)]))
                    painter.setPen(Qt.NoPen)
                    painter.drawEllipse(tt_rect.x()+10, cy+6, 8, 8)
                    painter.setPen(QColor("#475569"))
                    painter.drawText(QRect(tt_rect.x()+25, cy, tt_w-35, 20), Qt.AlignLeft, m)
                    painter.drawText(QRect(tt_rect.x()+25, cy, tt_w-35, 20), Qt.AlignRight, f"${val:,.0f}")
                    cy += 20


class DonutChartWidget(QWidget):
    def __init__(self, data=None, parent=None):
        super().__init__(parent)
        self.data = data or {}
        self.colors = ["#3B82F6", "#10B981", "#F59E0B", "#8B5CF6", "#EC4899", "#06B6D4", "#64748B"]
        self.setMinimumHeight(350)
        self.setAttribute(Qt.WA_Hover, True)
        self.hover_angle = -1
        self.setMouseTracking(True)
        
    def update_data(self, data):
        self.data = data
        self.update()

    def paintEvent(self, event):
        from PyQt6.QtGui import QPainter, QColor, QFont, QPen, QBrush
        from PyQt6.QtCore import QPoint, QRect, QRectF, Qt
        import math
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        # Leave enough horizontal space (80px on each side) for the text labels
        avail_w = self.width() - 180
        avail_h = self.height() - 40
        size = min(avail_w, avail_h)
        if size < 80: size = 80
        
        rect = QRect(int((self.width() - size) / 2), int((self.height() - size) / 2), size, size)
        
        total = sum(self.data.values())
        if total <= 0:
            painter.setBrush(QColor("#EEF2F8"))
            painter.drawEllipse(rect); return
            
        start_angle = 90 * 16
        for i, (cat, val) in enumerate(self.data.items()):
            if val <= 0: continue
            span_angle = int((val / total) * 360 * 16)
            color = QColor(self.colors[i % len(self.colors)])
            painter.setBrush(color)
            painter.setPen(QPen(QColor("#FFFFFF"), 2))
            painter.drawPie(rect, start_angle, -span_angle)
            
            # Draw outside line and label for top 5
            pct = (val / total) * 100
            if i < 5 and pct >= 2.0:
                mid_angle = (start_angle - span_angle / 2) / 16.0
                rad = math.radians(mid_angle)
                
                cx = rect.center().x()
                cy = rect.center().y()
                outer_radius = size / 2.0
                
                edge_x = cx + outer_radius * math.cos(rad)
                edge_y = cy - outer_radius * math.sin(rad)
                
                line_len = 15
                end_x = cx + (outer_radius + line_len) * math.cos(rad)
                end_y = cy - (outer_radius + line_len) * math.sin(rad)
                
                is_right = math.cos(rad) >= 0
                horiz_len = 10
                text_x = end_x + horiz_len if is_right else end_x - horiz_len
                
                painter.setPen(QPen(color, 2))
                painter.drawLine(int(edge_x), int(edge_y), int(end_x), int(end_y))
                painter.drawLine(int(end_x), int(end_y), int(text_x), int(end_y))
                
                painter.setPen(QColor("#1E293B"))
                painter.setFont(QFont("Segoe UI", 8, QFont.Bold))
                cat_str = str(cat)
                if len(cat_str) > 12: cat_str = cat_str[:10] + ".."
                label_text = f"{cat_str} {pct:.0f}%"
                
                if is_right:
                    text_rect = QRectF(text_x + 4, end_y - 10, 100, 20)
                    painter.drawText(text_rect, Qt.AlignLeft | Qt.AlignVCenter | Qt.TextDontClip, label_text)
                else:
                    text_rect = QRectF(text_x - 104, end_y - 10, 100, 20)
                    painter.drawText(text_rect, Qt.AlignRight | Qt.AlignVCenter | Qt.TextDontClip, label_text)
            
            start_angle -= span_angle
            
        # Center hole
        inner_size = int(size * 0.60)
        inner_rect = QRect(int((self.width() - inner_size) / 2), int((self.height() - inner_size) / 2), inner_size, inner_size)
        painter.setBrush(QColor("#F1F5F9"))
        painter.setPen(Qt.NoPen)
        painter.drawEllipse(inner_rect)
        
        painter.setPen(QColor("#1E293B"))
        
        # Format total to fit
        if total >= 1_000_000:
            t_str = f"${total/1_000_000:.1f}M"
        elif total >= 1_000:
            t_str = f"${total/1_000:.1f}K"
        else:
            t_str = f"${total:,.0f}"
            
        painter.setFont(QFont("Segoe UI", 13, QFont.Bold))
        painter.drawText(inner_rect, Qt.AlignCenter | Qt.TextDontClip, t_str)


class AIAssistantWidget(ModernCard):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("""
            #card {
                
                border: 2px solid #8B5CF6;
                border-radius: 20px;
            }
        """)
        
        lay = QVBoxLayout(self)
        lay.setContentsMargins(25, 20, 25, 20)
        lay.setSpacing(10)
        
        # Header
        h_lay = QHBoxLayout()
        lbl_icon = QLabel("🤖")
        lbl_icon.setStyleSheet("font-size: 24px;")
        lbl_title = QLabel("Antigravity AI - Análisis Estratégico")
        lbl_title.setStyleSheet("font-size: 18px; font-weight: 900; ")
        
        self.lbl_status = QLabel("Pensando...")
        self.lbl_status.setStyleSheet("font-size: 12px;  font-style: italic;")
        self.lbl_status.hide()
        
        h_lay.addWidget(lbl_icon)
        h_lay.addWidget(lbl_title)
        h_lay.addStretch()
        h_lay.addWidget(self.lbl_status)
        lay.addLayout(h_lay)
        
        # Content
        self.lbl_content = QLabel("Recopilando datos para generar insights...")
        self.lbl_content.setWordWrap(True)
        self.lbl_content.setStyleSheet("font-size: 14px;  line-height: 1.5;")
        lay.addWidget(self.lbl_content)
        
        # Timer for animation
        from PyQt6.QtCore import QTimer
        self.anim_timer = QTimer(self)
        self.anim_timer.timeout.connect(self._animate_typing)
        self.full_text = ""
        self.current_char = 0

    def update_insights(self, chart_data, pago_sum, donut_data):
        self.lbl_status.show()
        self.lbl_content.setText("")
        
        # Generar texto de insights
        total_ventas = sum([d.get('ventas', 0) for d in chart_data.values()])
        if total_ventas == 0:
            self.full_text = "No hay datos suficientes en este periodo para generar un análisis."
        else:
            # Insight 1: Mejor día
            mejor_dia = max(chart_data.items(), key=lambda x: x[1].get('ventas', 0))[0]
            val_mejor_dia = chart_data[mejor_dia].get('ventas', 0)
            
            # Insight 2: Depto estrella
            mejor_depto = "N/A"
            if donut_data:
                mejor_depto = max(donut_data.items(), key=lambda x: x[1])[0]
                
            # Insight 3: Forma de pago
            mejor_forma = "N/A"
            if pago_sum:
                mejor_forma = max(pago_sum.items(), key=lambda x: x[1])[0]
                
            self.full_text = f"""<ul>
                <li style='margin-bottom: 8px;'>📈 <b>Pico de Ventas:</b> El mejor desempeño fue el día <b>{mejor_dia}</b> con <b>${val_mejor_dia:,.2f}</b>. Asegúrate de replicar la estrategia de ese día.</li>
                <li style='margin-bottom: 8px;'>🏆 <b>Departamento Estrella:</b> <b>{mejor_depto}</b> está liderando en volumen. Considera ubicar promociones cruzadas cerca de esta sección.</li>
                <li style='margin-bottom: 8px;'>💳 <b>Preferencia de Pago:</b> La mayoría de tus clientes prefiere usar <b>{mejor_forma}</b>. Analiza si las comisiones de este método están optimizadas.</li>
            </ul>"""
            
        self.current_char = 0
        self.anim_timer.start(10) # 10ms por caracter HTML (approx)

    def _animate_typing(self):
        # We need to type HTML carefully, but for simplicity we just chunk it or show it directly if it's HTML.
        # To avoid breaking HTML tags during typing, we will just show it instantly after a small "thinking" delay.
        # Actually, let's just simulate 1 second of thinking, then show all.
        self.lbl_status.hide()
        self.lbl_content.setText(self.full_text)
        self.anim_timer.stop()

class DialogoVentasPorHora(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Ventas por Hora")
        self.resize(600, 400)
        from PyQt6.QtWidgets import QVBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView
        lay = QVBoxLayout(self)
        tabla = QTableWidget()
        tabla.setColumnCount(2)
        tabla.setHorizontalHeaderLabels(["Hora del Día", "Cantidad de Transacciones"])
        tabla.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tabla.setStyleSheet("QTableWidget { background: white; border: 1px solid #e2e8f0; border-radius: 8px; }")
        
        from src.utils.db import db_manager
        res = db_manager.execute_query("""
            SELECT substr(fecha, 12, 2) as hora, COUNT(id) as cant 
            FROM ventas
            WHERE estado IN ('COMPLETADA', 'COMPLETADO', 'CERRADA', 'CERRADO')
            GROUP BY hora
            ORDER BY hora ASC
        """)
        if res:
            tabla.setRowCount(len(res))
            for i, r in enumerate(res):
                hora = str(r['hora']) + ":00"
                cant = str(r['cant'])
                tabla.setItem(i, 0, QTableWidgetItem(hora))
                tabla.setItem(i, 1, QTableWidgetItem(cant))
        lay.addWidget(tabla)

class DialogoInventarioBajo(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Inventario Bajo (Reorden)")
        self.resize(800, 600)
        from PyQt6.QtWidgets import QVBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView
        lay = QVBoxLayout(self)
        tabla = QTableWidget()
        tabla.setColumnCount(3)
        tabla.setHorizontalHeaderLabels(["Código", "Producto", "Stock Actual"])
        tabla.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        from src.utils.db import db_manager
        res = db_manager.execute_query("SELECT codigo, nombre, stock FROM productos WHERE stock <= 5 ORDER BY stock ASC")
        if res:
            tabla.setRowCount(len(res))
            for i, r in enumerate(res):
                tabla.setItem(i, 0, QTableWidgetItem(str(r['codigo'] or '')))
                tabla.setItem(i, 1, QTableWidgetItem(str(r['nombre'] or '')))
                tabla.setItem(i, 2, QTableWidgetItem(str(r['stock'] or '0')))
        lay.addWidget(tabla)


from PyQt6.QtCore import QThread, pyqtSignal

class DataLoaderThread(QThread):
    data_loaded = pyqtSignal(dict)

    def __init__(self, periodo, start_str, end_str, period_type):
        super().__init__()
        self.periodo = periodo
        self.start_str = start_str
        self.end_str = end_str
        self.period_type = period_type

    def run(self):
        try:
            from src.utils.db import db_manager
            import datetime
            
            # KPIs
            res_kpi = db_manager.execute_query(
                "SELECT SUM(total) as v_bruta, SUM(total - descuento + recargo) as v_neta, COUNT(id) as cant "
                "FROM ventas WHERE (fecha BETWEEN ? AND ?) AND estado IN ('COMPLETADA', 'CERRADA')", 
                (self.start_str, self.end_str)
            )
            v_bruta = float(res_kpi[0]['v_bruta'] or 0.0) if res_kpi and res_kpi[0] else 0.0
            t_cant = int(res_kpi[0]['cant'] or 0) if res_kpi and res_kpi[0] else 0
            
            res_costo = db_manager.execute_query(
                "SELECT SUM(dv.cantidad * COALESCE(p.costo, 0)) as costo "
                "FROM detalles_ventas dv JOIN ventas v ON dv.id_venta = v.id "
                "LEFT JOIN productos p ON dv.id_producto = p.id "
                "WHERE (v.fecha BETWEEN ? AND ?) AND v.estado IN ('COMPLETADA', 'CERRADA')",
                (self.start_str, self.end_str)
            )
            costo = float(res_costo[0]['costo'] or 0.0) if res_costo and res_costo[0] else 0.0
            ganancia = v_bruta - costo
            
            # Chart Data
            s_dt_c = datetime.datetime.strptime(self.start_str, "%Y-%m-%d %H:%M:%S")
            e_dt_c = datetime.datetime.strptime(self.end_str, "%Y-%m-%d %H:%M:%S")
            days_diff = (e_dt_c - s_dt_c).days + 1
            
            display_chart_data = {}
            if self.period_type == "day":
                query_chart = "SELECT substr(fecha, 12, 2) as hora, SUM(total) as tot FROM ventas WHERE (fecha BETWEEN ? AND ?) AND estado IN ('COMPLETADA', 'CERRADA') GROUP BY hora"
                res = db_manager.execute_query(query_chart, (self.start_str, self.end_str))
                for r in (res or []): display_chart_data[f"{r['hora']}:00"] = float(r['tot'] or 0)
            elif self.period_type == "week" or days_diff <= 31:
                query_chart = "SELECT substr(fecha, 1, 10) as dia, SUM(total) as tot FROM ventas WHERE (fecha BETWEEN ? AND ?) AND estado IN ('COMPLETADA', 'CERRADA') GROUP BY dia"
                res = db_manager.execute_query(query_chart, (self.start_str, self.end_str))
                for r in (res or []):
                    dt_obj = datetime.datetime.strptime(r['dia'], "%Y-%m-%d")
                    display_chart_data[dt_obj.strftime("%d/%m")] = float(r['tot'] or 0)
            else:
                query_chart = "SELECT substr(fecha, 1, 7) as mes, SUM(total) as tot FROM ventas WHERE (fecha BETWEEN ? AND ?) AND estado IN ('COMPLETADA', 'CERRADA') GROUP BY mes"
                res = db_manager.execute_query(query_chart, (self.start_str, self.end_str))
                for r in (res or []):
                    try:
                        m_idx = int(r['mes'][-2:])
                        meses = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
                        display_chart_data[meses[m_idx]] = float(r['tot'] or 0)
                    except: display_chart_data[r['mes']] = float(r['tot'] or 0)
            
            # Tablas Varias
            res_diario = db_manager.execute_query(
                "SELECT substr(fecha, 1, 10) as dia, SUM(total) as tot FROM ventas WHERE (fecha BETWEEN ? AND ?) AND estado IN ('COMPLETADA', 'CERRADA') GROUP BY dia ORDER BY dia DESC", (self.start_str, self.end_str)
            )
            res_depto = db_manager.execute_query(
                "SELECT COALESCE(p.departamento, 'General') as depto, SUM(dv.subtotal) as tot, SUM(dv.cantidad * COALESCE(p.costo, 0)) as costo FROM detalles_ventas dv JOIN ventas v ON dv.id_venta = v.id LEFT JOIN productos p ON dv.id_producto = p.id WHERE (v.fecha BETWEEN ? AND ?) AND v.estado IN ('COMPLETADA', 'CERRADA') GROUP BY depto ORDER BY tot DESC", (self.start_str, self.end_str)
            )
            res_pago = db_manager.execute_query(
                "SELECT substr(fecha, 1, 10) as dia, COALESCE(metodo_pago, 'Efectivo') as m_pago, SUM(total) as tot FROM ventas WHERE (fecha BETWEEN ? AND ?) AND estado IN ('COMPLETADA', 'CERRADA') GROUP BY dia, m_pago", (self.start_str, self.end_str)
            )
            res_cajeros = db_manager.execute_query(
                "SELECT COALESCE(v.usuario, 'Desconocido') as cajero, COUNT(v.id) as cant, SUM(v.total) as tot FROM ventas v WHERE (v.fecha BETWEEN ? AND ?) AND v.estado IN ('COMPLETADA', 'CERRADA') GROUP BY cajero ORDER BY tot DESC LIMIT 5", (self.start_str, self.end_str)
            )
            res_productos = db_manager.execute_query(
                "SELECT dv.nombre_producto as nombre, SUM(dv.cantidad) as cant, SUM(dv.subtotal) as tot FROM detalles_ventas dv JOIN ventas v ON dv.id_venta = v.id WHERE (v.fecha BETWEEN ? AND ?) AND v.estado IN ('COMPLETADA', 'CERRADA') GROUP BY dv.nombre_producto ORDER BY tot DESC LIMIT 5", (self.start_str, self.end_str)
            )

            self.data_loaded.emit({
                "v_bruta": v_bruta,
                "ganancia": ganancia,
                "t_cant": t_cant,
                "display_chart_data": display_chart_data,
                "res_diario": res_diario,
                "res_depto": res_depto,
                "res_pago": res_pago,
                "res_cajeros": res_cajeros,
                "res_productos": res_productos
            })
        except Exception as e:
            print("Error in DataLoaderThread:", e)
            self.data_loaded.emit({})

class VistaFinanciero(QWidget):

    request_dashboard = pyqtSignal()
    _prod_map_cache = None
    _unique_prods_base = None
    def __init__(self):
        super().__init__()
        self.setup_ui()
        self.cargar_datos("Semana Actual")
        
        # Sincronización en Tiempo Real (Solo para Modo Espectador / Red)
        from src.config import config
        from PyQt6.QtCore import QTimer
        db_path = config.get("db_path", "")
        if db_path and db_path != "":
            self.sync_timer = QTimer(self)
            self.sync_timer.timeout.connect(self.sincronizacion_silenciosa)
            self.sync_timer.start(10000) # Cada 10 segundos

    def sincronizacion_silenciosa(self):
        if not self.isVisible(): return
        
        # Si estamos en la pestaña 1 (Gráficos), animamos también la actualización de gráficos
        if hasattr(self, 'stack_views') and self.stack_views.currentIndex() == 1:
            self._renderizar_graficos()
            if self.txt_audit_prod.hasFocus(): return
            self._buscar_auditoria()
        else:
            periodo = getattr(self, "current_period", "Semana Actual")
            self.cargar_datos(periodo)

    def _crear_tabla_estilizada(self, titulo, col_count=2):
        card = ModernCard()
        lay = QVBoxLayout(card)
        lay.setContentsMargins(25, 20, 25, 20)
        lay.setSpacing(15)
        
        lbl_tit = QLabel(titulo)
        lbl_tit.setObjectName("lbl_tit_tabla")
        lbl_tit.setStyleSheet(
            f"font-size: 17px; font-weight: 800; color: {_FIN['text']}; "
            "background: transparent; border: none; letter-spacing: -0.3px;"
        )
        lay.addWidget(lbl_tit)
        
        tabla = QTableWidget()
        tabla.setColumnCount(col_count)
        tabla.horizontalHeader().setVisible(False)
        tabla.verticalHeader().setVisible(False)
        tabla.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tabla.setSelectionMode(QAbstractItemView.NoSelection)
        tabla.setShowGrid(False)
        tabla.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        tabla.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        if col_count > 2:
            for i in range(2, col_count):
                tabla.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeToContents)
        _aplicar_paleta_tabla(tabla)
        lay.addWidget(tabla)
        
        lbl_total = QLabel("")
        lbl_total.setStyleSheet(
            f"font-size: 13px; font-weight: 700; color: {_FIN['text_soft']}; text-align: right; border: none;"
        )
        lbl_total.setAlignment(Qt.AlignRight)
        lay.addWidget(lbl_total)
        
        return card, tabla, lbl_total

    def setup_ui(self):
        self.setObjectName("Admin3Reportes")
        self.setStyleSheet(f"""
            QWidget#Admin3Reportes {{
                background: {_FIN['bg_page']};
                font-family: 'Segoe UI', sans-serif;
                color: {_FIN['text']};
            }}
            QScrollArea {{ border: none; background: transparent; }}
            QWidget#ScrollContainer {{ background: transparent; }}
            QScrollBar:vertical {{ background: transparent; width: 6px; }}
            QScrollBar::handle:vertical {{ background: rgba(75,85,99,0.35); border-radius: 3px; }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
        """)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        

        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content_widget = QWidget()
        content_widget.setObjectName("ScrollContainer")
        self.content_layout = QVBoxLayout(content_widget)
        self.content_layout.setContentsMargins(32, 24, 32, 32)
        self.content_layout.setSpacing(24)
        
        # TITLE AND FILTERS
        lbl_main = QLabel("Resumen de Ventas de Mayo")
        lbl_main.setObjectName("lbl_main_title_financial")
        lbl_main.setStyleSheet(
            f"font-size: 24px; font-weight: 800; color: {_FIN['text']}; margin-bottom: 5px; border: none;"
        )
        self.content_layout.addWidget(lbl_main)
        
        filters_layout = QHBoxLayout()
        filters_layout.setSpacing(12)
        self.period_buttons = {}
        self._period_btn_active = (
            f"QPushButton {{ color: white; font-size: 13px; font-weight: 800; "
            f"border: none; background: {_FIN['accent']}; border-radius: 10px; padding: 8px 16px; }}"
            f"QPushButton:hover {{ background: {_FIN['accent_hover']}; }}"
        )
        self._period_btn_idle = (
            f"QPushButton {{ color: {_FIN['text_soft']}; font-size: 13px; font-weight: 600; "
            f"background: white; padding: 8px 16px; border-radius: 10px; "
            f"border: 1px solid {_FIN['card_border']}; }}"
            f"QPushButton:hover {{ color: {_FIN['accent']}; background: {_FIN['accent_light']}; "
            f"border-color: #BFDBFE; }}"
        )
        for f_text in ["Semana Actual", "Mes Actual", "Mes Anterior", "Año actual", "Periodo..."]:
            f_btn = QPushButton(f_text)
            f_btn.setCursor(Qt.PointingHandCursor)
            f_btn.setStyleSheet(self._period_btn_idle)
            f_btn.clicked.connect(lambda checked, t=f_text: self.cargar_datos(t))
            self.period_buttons[f_text] = f_btn
            filters_layout.addWidget(f_btn)
        filters_layout.addStretch()
        
        from PyQt6.QtWidgets import QCheckBox
        self.chk_comparativa = QCheckBox("Comparar con periodo anterior")
        self.chk_comparativa.setStyleSheet(
            f"font-weight: 600; font-size: 13px; color: {_FIN['text_soft']}; border: none;"
        )
        self.chk_comparativa.stateChanged.connect(lambda: self.cargar_datos(getattr(self, "current_period", "Mes Actual")))
        filters_layout.addWidget(self.chk_comparativa)
        
        self.content_layout.addLayout(filters_layout)
        
        # AI ASSISTANT
        self.ai_assistant = AIAssistantWidget()
        self.content_layout.addWidget(self.ai_assistant)
        
        # MAIN CHART
        self.card_main_chart = ModernCard()
        self.card_main_chart.setMinimumHeight(380)
        chart_layout = QVBoxLayout(self.card_main_chart)
        chart_layout.setContentsMargins(15, 15, 15, 15)
        self.stock_chart = StockAreaChartWidget()
        chart_layout.addWidget(self.stock_chart)
        
        # Legend
        legend_layout = QHBoxLayout()
        legend_layout.addStretch()
        lbl_v = QLabel("■ Ventas")
        lbl_v.setStyleSheet(" font-weight: bold; font-size: 12px;")
        lbl_g = QLabel("■ Ganancia")
        lbl_g.setStyleSheet(" font-weight: bold; font-size: 12px;")
        legend_layout.addWidget(lbl_v)
        legend_layout.addWidget(lbl_g)
        chart_layout.addLayout(legend_layout)
        self.content_layout.addWidget(self.card_main_chart)
        
        # KPIS
        self.kpi_container = QFrame()
        self.kpi_container.setStyleSheet(
            f"background: transparent; border-top: 1px solid {_FIN['card_border']}; "
            f"border-bottom: 1px solid {_FIN['card_border']};"
        )
        self.kpi_layout = QGridLayout(self.kpi_container)
        self.kpi_layout.setContentsMargins(10, 20, 10, 20)
        self.kpi_layout.setSpacing(20)
        self.content_layout.addWidget(self.kpi_container)
        
        # TWO COLUMNS GRID
        self.tables_grid = QGridLayout()
        self.tables_grid.setSpacing(24)
        
        # Col 1
        self.card_vtas_tiempo, self.tabla_vtas_tiempo, _ = self._crear_tabla_estilizada("Ventas por semana")
        self.tables_grid.addWidget(self.card_vtas_tiempo, 0, 0)
        
        self.card_pago, self.tabla_pago, _ = self._crear_tabla_estilizada("Ventas por forma de pago")
        self.pago_chart = BarChartWidget()
        self.card_pago.layout().insertWidget(1, self.pago_chart)
        self.tables_grid.addWidget(self.card_pago, 1, 0)
        
        self.card_rec, self.tabla_rec, _ = self._crear_tabla_estilizada("Mayor Recaudación ($)", 3)
        self.tabla_rec.setMinimumHeight(350)
        self.tables_grid.addWidget(self.card_rec, 2, 0)
        
        self.card_impuestos, self.tabla_impuestos, _ = self._crear_tabla_estilizada("Impuestos")
        self.tables_grid.addWidget(self.card_impuestos, 3, 0, 1, 2)
        
        # Col 2
        self.card_vtas_depto, self.tabla_vtas_depto, _ = self._crear_tabla_estilizada("Ventas por Departamento")
        self.depto_donut = DonutChartWidget()
        self.depto_donut.setFixedSize(220, 220)
        d_layout = QHBoxLayout()
        d_layout.addWidget(self.depto_donut, alignment=Qt.AlignCenter)
        self.card_vtas_depto.layout().insertLayout(1, d_layout)
        self.tables_grid.addWidget(self.card_vtas_depto, 0, 1)
        
        self.card_gan_depto, self.tabla_gan_depto, _ = self._crear_tabla_estilizada("Ganancia por Departamento")
        self.tables_grid.addWidget(self.card_gan_depto, 1, 1)
        
        self.card_vol, self.tabla_vol, _ = self._crear_tabla_estilizada("Mayor Volumen (Cant.)", 3)
        self.tabla_vol.setMinimumHeight(350)
        self.tables_grid.addWidget(self.card_vol, 2, 1)
        
        self.content_layout.addLayout(self.tables_grid)
        self.content_layout.addStretch()
        
        scroll.setWidget(content_widget)
        main_layout.addWidget(scroll)
        self.cargar_datos("Semana Actual")


    def cargar_datos(self, periodo="Mes Actual"):
        self.current_period = periodo
        try:
            if periodo == "Periodo...":
                try:
                    dialog = DialogoSeleccionPeriodo(self)
                    if qt_exec(dialog) == QDialog.Accepted:
                        start_str, end_str = dialog.get_fechas()
                    else:
                        return
                except NameError:
                    return # Not implemented yet
            else:
                import datetime
                hoy = datetime.date.today()
                if periodo == "Semana Actual":
                    start_date = hoy - datetime.timedelta(days=hoy.weekday())
                    start_str = start_date.strftime("%Y-%m-%d 00:00:00")
                    end_str = hoy.strftime("%Y-%m-%d 23:59:59")
                elif periodo == "Mes Actual":
                    import calendar
                    start_str = hoy.replace(day=1).strftime("%Y-%m-%d 00:00:00")
                    last_day = calendar.monthrange(hoy.year, hoy.month)[1]
                    end_str = hoy.replace(day=last_day).strftime("%Y-%m-%d 23:59:59")
                elif periodo == "Mes Anterior":
                    first_day_this_month = hoy.replace(day=1)
                    last_day_prev = first_day_this_month - datetime.timedelta(days=1)
                    start_str = last_day_prev.replace(day=1).strftime("%Y-%m-%d 00:00:00")
                    end_str = last_day_prev.strftime("%Y-%m-%d 23:59:59")
                elif periodo == "Año actual":
                    start_str = hoy.replace(month=1, day=1).strftime("%Y-%m-%d 00:00:00")
                    end_str = hoy.strftime("%Y-%m-%d 23:59:59")
                else:
                    start_str, end_str = hoy.strftime("%Y-%m-%d 00:00:00"), hoy.strftime("%Y-%m-%d 23:59:59")
                    
            self.current_start_str = start_str
            self.current_end_str = end_str
            for text, btn in self.period_buttons.items():
                btn.setStyleSheet(self._period_btn_active if text == periodo else self._period_btn_idle)
            
            lbl_title = self.findChild(QLabel, "lbl_main_title_financial")
            if lbl_title:
                m_name = ""
                if "Mes" in periodo:
                    try:
                        m_idx = int(start_str[5:7])
                        meses = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
                        m_name = meses[m_idx]
                        lbl_title.setText(f"Resumen de Ventas de {m_name}")
                    except: lbl_title.setText(f"Resumen de Ventas ({periodo})")
                else:
                    lbl_title.setText(f"Resumen de Ventas ({periodo})")

            lbl_vtas_tiempo = self.card_vtas_tiempo.findChild(QLabel, "lbl_tit_tabla")
            if lbl_vtas_tiempo:
                if "Mes" in periodo:
                    lbl_vtas_tiempo.setText("Ventas por mes")
                elif "Año" in periodo:
                    lbl_vtas_tiempo.setText("Ventas por año")
                else:
                    lbl_vtas_tiempo.setText("Ventas por semana")

            # FETCH DATA
            # 1. Ventas Totales y KPIs
            res_kpi = db_manager.execute_query(
                "SELECT SUM(total) as v_bruta, SUM(total - descuento + recargo) as v_neta, COUNT(id) as cant "
                "FROM ventas WHERE (fecha BETWEEN ? AND ?) AND estado IN ('COMPLETADA', 'CERRADA')", 
                (start_str, end_str)
            )
            v_bruta = 0.0
            t_cant = 0
            t_promedio = 0.0
            if res_kpi and res_kpi[0]:
                v_bruta = float(res_kpi[0]['v_bruta'] or 0.0)
                t_cant = int(res_kpi[0]['cant'] or 0)
                t_promedio = v_bruta / t_cant if t_cant > 0 else 0.0

            # 2. Ganancia Total y Costos
            res_costo = db_manager.execute_query(
                "SELECT SUM(dv.cantidad * COALESCE(p.costo, 0)) as costo "
                "FROM detalles_ventas dv JOIN ventas v ON dv.id_venta = v.id "
                "LEFT JOIN productos p ON dv.id_producto = p.id "
                "WHERE (v.fecha BETWEEN ? AND ?) AND v.estado IN ('COMPLETADA', 'CERRADA')",
                (start_str, end_str)
            )
            costo = float(res_costo[0]['costo'] or 0.0) if res_costo and res_costo[0] else 0.0
            ganancia = v_bruta - costo
            margen = (ganancia / v_bruta * 100) if v_bruta > 0 else 0.0
            
            # Limpiar KPIs previos
            for i in reversed(range(self.kpi_layout.count())): 
                w = self.kpi_layout.itemAt(i).widget()
                if w: w.deleteLater()
            
            def create_kpi(title, value, color=None, align="left"):
                if color is None:
                    color = _FIN["text"]
                w = QFrame()
                w.setStyleSheet(
                    f"background: {_FIN['card']}; border-radius: 16px; "
                    f"border: 1px solid {_FIN['card_border']};"
                )
                l = QVBoxLayout(w)
                l.setContentsMargins(20, 18, 20, 18)
                l.setSpacing(6)
                t = QLabel(title.upper())
                t.setStyleSheet(
                    f"font-size: 11px; font-weight: 800; letter-spacing: 0.8px; "
                    f"color: {_FIN['muted_on_white']}; border: none;"
                )
                v = QLabel(value)
                v.setStyleSheet(f"color: {color}; font-size: 26px; font-weight: 800; border: none;")
                
                t.setAlignment(Qt.AlignCenter)
                v.setAlignment(Qt.AlignCenter)
                l.addWidget(t)
                l.addWidget(v)
                return w
                
            self.kpi_layout.addWidget(create_kpi("Ventas Totales", f"${v_bruta:,.2f}"), 0, 0)
            self.kpi_layout.addWidget(create_kpi("Ganancia Neta", f"${ganancia:,.2f}"), 0, 2)
            self.kpi_layout.addWidget(create_kpi("Número de Ventas", f"{t_cant}"), 1, 0)
            self.kpi_layout.addWidget(create_kpi("Ticket Promedio", f"${t_promedio:,.2f}"), 2, 0)
            self.kpi_layout.addWidget(create_kpi("Margen de utilidad promedio", f"{margen:,.0f}%", _FIN["green"]), 1, 2, 2, 1)
            
            # 3. Main Chart Data (Agrupacion inteligente)
            period_type = "day"
            if "Año" in self.current_period:
                period_type = "month"
            elif "Semana" in self.current_period:
                period_type = "week"
            else:
                period_type = "month_days"
                
            import datetime
            try:
                s_dt_c = datetime.datetime.strptime(start_str, "%Y-%m-%d %H:%M:%S")
                e_dt_c = datetime.datetime.strptime(end_str, "%Y-%m-%d %H:%M:%S")
            except:
                s_dt_c = datetime.datetime.now()
                e_dt_c = datetime.datetime.now()
                
            chart_data = {}
            if period_type == "month":
                meses_nombres = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
                for i, m in enumerate(meses_nombres):
                    chart_data[f"{i+1:02d}"] = {'ventas': 0.0, 'ganancia': 0.0, 'label': m}
            elif period_type == "week":
                dias_nombres = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
                for i in range(7):
                    curr_d = s_dt_c + datetime.timedelta(days=i)
                    chart_data[curr_d.strftime("%Y-%m-%d")] = {'ventas': 0.0, 'ganancia': 0.0, 'label': dias_nombres[i]}
            else:
                days_diff = (e_dt_c - s_dt_c).days + 1
                if days_diff > 60:
                    period_type = "month"
                    meses_nombres = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
                    for i, m in enumerate(meses_nombres):
                        chart_data[f"{i+1:02d}"] = {'ventas': 0.0, 'ganancia': 0.0, 'label': m}
                else:
                    for i in range(days_diff):
                        curr_d = s_dt_c + datetime.timedelta(days=i)
                        chart_data[curr_d.strftime("%Y-%m-%d")] = {'ventas': 0.0, 'ganancia': 0.0, 'label': curr_d.strftime("%d")}
                        
            if period_type == "month":
                res_tot = db_manager.execute_query(
                    "SELECT substr(fecha, 6, 2) as key_val, SUM(total) as tot "
                    "FROM ventas WHERE (fecha BETWEEN ? AND ?) AND estado IN ('COMPLETADA', 'CERRADA') "
                    "GROUP BY key_val", (start_str, end_str)
                )
                res_cost = db_manager.execute_query(
                    "SELECT substr(v.fecha, 6, 2) as key_val, SUM(dv.cantidad * COALESCE(p.costo, 0)) as costo "
                    "FROM ventas v JOIN detalles_ventas dv ON v.id = dv.id_venta "
                    "LEFT JOIN productos p ON dv.id_producto = p.id "
                    "WHERE (v.fecha BETWEEN ? AND ?) AND v.estado IN ('COMPLETADA', 'CERRADA') "
                    "GROUP BY key_val", (start_str, end_str)
                )
            else:
                res_tot = db_manager.execute_query(
                    "SELECT substr(fecha, 1, 10) as key_val, SUM(total) as tot "
                    "FROM ventas WHERE (fecha BETWEEN ? AND ?) AND estado IN ('COMPLETADA', 'CERRADA') "
                    "GROUP BY key_val", (start_str, end_str)
                )
                res_cost = db_manager.execute_query(
                    "SELECT substr(v.fecha, 1, 10) as key_val, SUM(dv.cantidad * COALESCE(p.costo, 0)) as costo "
                    "FROM ventas v JOIN detalles_ventas dv ON v.id = dv.id_venta "
                    "LEFT JOIN productos p ON dv.id_producto = p.id "
                    "WHERE (v.fecha BETWEEN ? AND ?) AND v.estado IN ('COMPLETADA', 'CERRADA') "
                    "GROUP BY key_val", (start_str, end_str)
                )
                
            if res_tot:
                for r in res_tot:
                    k = str(r['key_val'])
                    if k in chart_data:
                        chart_data[k]['ventas'] = float(r['tot'] or 0.0)
                        chart_data[k]['ganancia'] = float(r['tot'] or 0.0)
            if res_cost:
                for r in res_cost:
                    k = str(r['key_val'])
                    if k in chart_data:
                        c = float(r['costo'] or 0.0)
                        chart_data[k]['ganancia'] = chart_data[k]['ventas'] - c

            display_chart_data = {v['label']: {'ventas': v['ventas'], 'ganancia': v['ganancia']} for k, v in chart_data.items()}
            
            res_diario = []
            for k, v in chart_data.items():
                if period_type == "month":
                    res_diario.append({'dia': v['label'], 'tot': v['ventas'], 'label': v['label']})
                elif period_type == "week":
                    res_diario.append({'dia': v['label'], 'tot': v['ventas'], 'label': v['label']})
                else:
                    res_diario.append({'dia': f"Día {v['label']}", 'tot': v['ventas'], 'label': v['label']})

            chart_data_prev = None
            display_chart_data_prev = None
            if hasattr(self, 'chk_comparativa') and self.chk_comparativa.isChecked():
                try:
                    if "Año" in self.current_period:
                        try:
                            from dateutil.relativedelta import relativedelta
                            prev_start_dt = s_dt_c - relativedelta(years=1)
                            prev_end_dt = s_dt_c - datetime.timedelta(days=1)
                        except:
                            prev_start_dt = s_dt_c - datetime.timedelta(days=365)
                            prev_end_dt = s_dt_c - datetime.timedelta(days=1)
                    elif "Mes" in self.current_period:
                        try:
                            from dateutil.relativedelta import relativedelta
                            prev_start_dt = s_dt_c - relativedelta(months=1)
                            prev_end_dt = s_dt_c - datetime.timedelta(days=1)
                        except:
                            diff = e_dt_c - s_dt_c
                            prev_end_dt = s_dt_c - datetime.timedelta(days=1)
                            prev_start_dt = prev_end_dt - diff
                    elif "Semana" in self.current_period:
                        prev_start_dt = s_dt_c - datetime.timedelta(days=7)
                        prev_end_dt = s_dt_c - datetime.timedelta(days=1)
                    else:
                        try:
                            from dateutil.relativedelta import relativedelta
                            prev_start_dt = s_dt_c - relativedelta(years=1)
                            prev_end_dt = e_dt_c - relativedelta(years=1)
                        except:
                            prev_start_dt = s_dt_c - datetime.timedelta(days=365)
                            prev_end_dt = e_dt_c - datetime.timedelta(days=365)
                        
                    prev_start = prev_start_dt.strftime("%Y-%m-%d 00:00:00")
                    prev_end = prev_end_dt.strftime("%Y-%m-%d 23:59:59")
                    
                    if period_type == "month":
                        res_prev = db_manager.execute_query(
                            "SELECT substr(v.fecha, 6, 2) as key_val, SUM(v.total) as tot "
                            "FROM ventas v WHERE (v.fecha BETWEEN ? AND ?) AND v.estado IN ('COMPLETADA', 'CERRADA') "
                            "GROUP BY key_val", (prev_start, prev_end)
                        )
                    else:
                        res_prev = db_manager.execute_query(
                            "SELECT substr(v.fecha, 1, 10) as key_val, SUM(v.total) as tot "
                            "FROM ventas v WHERE (v.fecha BETWEEN ? AND ?) AND v.estado IN ('COMPLETADA', 'CERRADA') "
                            "GROUP BY key_val", (prev_start, prev_end)
                        )
                        
                    display_chart_data_prev = {v['label']: {'ventas': 0.0} for k, v in chart_data.items()}
                    
                    if res_prev:
                        prev_keys_sorted = []
                        if period_type == "month":
                            prev_keys_sorted = [f"{i+1:02d}" for i in range(12)]
                        else:
                            days_diff_prev = (prev_end_dt - prev_start_dt).days + 1
                            for i in range(days_diff_prev):
                                d = prev_start_dt + datetime.timedelta(days=i)
                                prev_keys_sorted.append(d.strftime("%Y-%m-%d"))
                        
                        curr_keys = list(chart_data.keys())
                        for r in res_prev:
                            pk = str(r['key_val'])
                            v = float(r['tot'] or 0.0)
                            if pk in prev_keys_sorted:
                                idx = prev_keys_sorted.index(pk)
                                if idx < len(curr_keys):
                                    ck = curr_keys[idx]
                                    label = chart_data[ck]['label']
                                    display_chart_data_prev[label]['ventas'] += v
                                    
                    chart_data_prev = display_chart_data_prev
                except Exception as e:
                    print("Error parsing prev period logic:", e)

            self.stock_chart.update_data(display_chart_data, chart_data_prev)

            
            is_comp = chart_data_prev is not None
            # Helper to calculate and return formatted QTableWidgetItem
            def create_diff_item(v_act, v_prev):
                diff = v_act - v_prev
                if v_prev > 0:
                    pct = (diff / v_prev) * 100
                else:
                    pct = 100.0 if diff > 0 else 0.0
                
                if diff > 0:
                    it = QTableWidgetItem(f"▲ +{pct:.1f}%")
                    it.setForeground(QBrush(QColor("#10B981"))) # Green
                elif diff < 0:
                    it = QTableWidgetItem(f"▼ {pct:.1f}%")
                    it.setForeground(QBrush(QColor("#EF4444"))) # Red
                else:
                    it = QTableWidgetItem("=")
                    it.setForeground(QBrush(QColor("#94A3B8")))
                it.setTextAlignment(Qt.AlignCenter)
                font = QFont("Segoe UI", 10, QFont.Bold)
                it.setFont(font)
                return it

            table_style = _estilo_tabla_financiera()

            # 4. Ventas por semana (Día a Día)
            self.tabla_vtas_tiempo.setRowCount(0)
            self.tabla_vtas_tiempo.setColumnCount(4 if is_comp else 2)
            self.tabla_vtas_tiempo.setSelectionBehavior(QTableWidget.SelectRows)
            if is_comp:
                self.tabla_vtas_tiempo.horizontalHeader().setVisible(True)
                self.tabla_vtas_tiempo.setHorizontalHeaderLabels(["Día", "Actual", "Anterior", "Rendimiento"])
            else:
                self.tabla_vtas_tiempo.horizontalHeader().setVisible(False)
                
            for tbl in (self.tabla_vtas_tiempo, self.tabla_vtas_depto, self.tabla_gan_depto, self.tabla_pago):
                _aplicar_paleta_tabla(tbl)
            
            self.tabla_vtas_depto.setSelectionBehavior(QTableWidget.SelectRows)
            self.tabla_gan_depto.setSelectionBehavior(QTableWidget.SelectRows)
            self.tabla_pago.setSelectionBehavior(QTableWidget.SelectRows)

            if res_diario:
                for i, r in enumerate(res_diario):
                    d = str(r['dia'])
                    short_d = d
                    short_d_key = r.get('label', d)
                    v = float(r['tot'] or 0.0)
                    self.tabla_vtas_tiempo.insertRow(i)
                    self.tabla_vtas_tiempo.setItem(i, 0, QTableWidgetItem(short_d))
                    it_v = QTableWidgetItem(f"${v:,.2f}")
                    it_v.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    it_v.setFont(QFont("Segoe UI", 10, QFont.Bold))
                    
                    if is_comp:
                        v_prev = chart_data_prev.get(short_d_key, {}).get('ventas', 0) if chart_data_prev else 0
                        it_p = QTableWidgetItem(f"${v_prev:,.2f}")
                        it_p.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        it_p.setForeground(QBrush(QColor("#94A3B8")))
                        self.tabla_vtas_tiempo.setItem(i, 1, it_v)
                        self.tabla_vtas_tiempo.setItem(i, 2, it_p)
                        self.tabla_vtas_tiempo.setItem(i, 3, create_diff_item(v, v_prev))
                    else:
                        self.tabla_vtas_tiempo.setItem(i, 1, it_v)
            
            # PREV QUERIES FOR DEPTOS AND PAGOS
            depto_prev_map = {}
            pago_prev_map = {}
            if is_comp:
                try:
                    import datetime
                    s_dt = datetime.datetime.strptime(start_str, "%Y-%m-%d %H:%M:%S")
                    e_dt = datetime.datetime.strptime(end_str, "%Y-%m-%d %H:%M:%S")
                    
                    # Fix date math to align months perfectly
                    periodo_str = getattr(self, "current_period", "")
                    if "Año" in periodo_str:
                        try:
                            from dateutil.relativedelta import relativedelta
                            prev_start = s_dt - relativedelta(years=1)
                            prev_end = s_dt - datetime.timedelta(days=1)
                        except:
                            prev_start = s_dt - datetime.timedelta(days=365)
                            prev_end = s_dt - datetime.timedelta(days=1)
                    elif "Mes" in periodo_str:
                        try:
                            from dateutil.relativedelta import relativedelta
                            prev_start = s_dt - relativedelta(months=1)
                            prev_end = s_dt - datetime.timedelta(days=1)
                        except:
                            diff = e_dt - s_dt
                            prev_end = s_dt - datetime.timedelta(days=1)
                            prev_start = prev_end - diff
                    elif "Semana" in periodo_str:
                        prev_start = s_dt - datetime.timedelta(days=7)
                        prev_end = s_dt - datetime.timedelta(days=1)
                    else:
                        try:
                            from dateutil.relativedelta import relativedelta
                            prev_start = s_dt - relativedelta(years=1)
                            prev_end = e_dt - relativedelta(years=1)
                        except:
                            prev_start = s_dt - datetime.timedelta(days=365)
                            prev_end = e_dt - datetime.timedelta(days=365)
                        
                    res_depto_prev = db_manager.execute_query(
                        "SELECT COALESCE(p.departamento, 'S/D') as depto, SUM(dv.subtotal) as tot, SUM(dv.cantidad * COALESCE(p.costo, 0)) as costo "
                        "FROM detalles_ventas dv JOIN ventas v ON dv.id_venta = v.id "
                        "LEFT JOIN productos p ON dv.id_producto = p.id "
                        "WHERE (v.fecha BETWEEN ? AND ?) AND v.estado IN ('COMPLETADA', 'CERRADA') "
                        "GROUP BY depto",
                        (prev_start.strftime("%Y-%m-%d 00:00:00"), prev_end.strftime("%Y-%m-%d 23:59:59"))
                    )
                    if res_depto_prev:
                        for r in res_depto_prev:
                            nom = str(r['depto'] or 'S/D').upper()
                            depto_prev_map[nom] = {'tot': float(r['tot'] or 0.0), 'costo': float(r['costo'] or 0.0)}
                            
                    res_pago_prev = db_manager.execute_query(
                        "SELECT COALESCE(v.metodo_pago, 'Efectivo') as pago, SUM(v.total) as tot "
                        "FROM ventas v "
                        "WHERE (v.fecha BETWEEN ? AND ?) AND v.estado IN ('COMPLETADA', 'CERRADA') "
                        "GROUP BY pago",
                        (prev_start.strftime("%Y-%m-%d 00:00:00"), prev_end.strftime("%Y-%m-%d 23:59:59"))
                    )
                    if res_pago_prev:
                        for r in res_pago_prev:
                            pago_prev_map[str(r['pago'])] = float(r['tot'] or 0.0)
                except Exception as e:
                    print("Error getting prev depto/pago data:", e)

            # 5. Ventas por Departamento
            self.tabla_vtas_depto.setRowCount(0)
            self.tabla_vtas_depto.setColumnCount(4 if is_comp else 2)
            if is_comp:
                self.tabla_vtas_depto.horizontalHeader().setVisible(True)
                self.tabla_vtas_depto.setHorizontalHeaderLabels(["Departamento", "Actual", "Anterior", "Crecimiento"])
            else:
                self.tabla_vtas_depto.horizontalHeader().setVisible(False)
                
            res_depto = db_manager.execute_query(
                "SELECT COALESCE(p.departamento, 'S/D') as depto, SUM(dv.subtotal) as tot, SUM(dv.cantidad * COALESCE(p.costo, 0)) as costo "
                "FROM detalles_ventas dv JOIN ventas v ON dv.id_venta = v.id "
                "LEFT JOIN productos p ON dv.id_producto = p.id "
                "WHERE (v.fecha BETWEEN ? AND ?) AND v.estado IN ('COMPLETADA', 'CERRADA') "
                "GROUP BY depto ORDER BY tot DESC", (start_str, end_str)
            )
            donut_data = {}
            if res_depto:
                for i, r in enumerate(res_depto):
                    nom = str(r['depto'] or 'S/D').upper()
                    v = float(r['tot'] or 0.0)
                    donut_data[nom] = v
                    self.tabla_vtas_depto.insertRow(i)
                    
                    colors = ["#3B82F6", "#10B981", "#F59E0B", "#8B5CF6", "#EC4899", "#06B6D4", "#64748B"]
                    c_hex = colors[i % len(colors)]
                    
                    it_nom = QTableWidgetItem(f"● {nom}")
                    it_nom.setForeground(QBrush(QColor(c_hex)))
                    font_nom = QFont("Segoe UI", 10, QFont.Bold)
                    it_nom.setFont(font_nom)
                    
                    self.tabla_vtas_depto.setItem(i, 0, it_nom)
                    it_v = QTableWidgetItem(f"${v:,.2f}")
                    it_v.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    it_v.setFont(QFont("Segoe UI", 10, QFont.Bold))
                    
                    if is_comp:
                        v_prev = depto_prev_map.get(nom, {}).get('tot', 0)
                        it_p = QTableWidgetItem(f"${v_prev:,.2f}")
                        it_p.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        it_p.setForeground(QBrush(QColor("#94A3B8")))
                        self.tabla_vtas_depto.setItem(i, 1, it_v)
                        self.tabla_vtas_depto.setItem(i, 2, it_p)
                        self.tabla_vtas_depto.setItem(i, 3, create_diff_item(v, v_prev))
                    else:
                        self.tabla_vtas_depto.setItem(i, 1, it_v)
            self.depto_donut.data = donut_data
            self.depto_donut.update()
            
            # 6. Ganancia por Departamento
            self.tabla_gan_depto.setRowCount(0)
            self.tabla_gan_depto.setColumnCount(4 if is_comp else 2)
            if is_comp:
                self.tabla_gan_depto.horizontalHeader().setVisible(True)
                self.tabla_gan_depto.setHorizontalHeaderLabels(["Departamento", "Actual", "Anterior", "Crecimiento"])
            else:
                self.tabla_gan_depto.horizontalHeader().setVisible(False)
                
            if res_depto:
                for i, r in enumerate(res_depto):
                    nom = str(r['depto'] or 'S/D').upper()
                    v = float(r['tot'] or 0.0)
                    c = float(r['costo'] or 0.0)
                    g = v - c
                    self.tabla_gan_depto.insertRow(i)
                    self.tabla_gan_depto.setItem(i, 0, QTableWidgetItem(nom))
                    it_v = QTableWidgetItem(f"${g:,.2f}")
                    it_v.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    it_v.setFont(QFont("Segoe UI", 10, QFont.Bold))
                    
                    if is_comp:
                        p_t = depto_prev_map.get(nom, {}).get('tot', 0)
                        p_c = depto_prev_map.get(nom, {}).get('costo', 0)
                        g_prev = p_t - p_c
                        it_p = QTableWidgetItem(f"${g_prev:,.2f}")
                        it_p.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        it_p.setForeground(QBrush(QColor("#94A3B8")))
                        self.tabla_gan_depto.setItem(i, 1, it_v)
                        self.tabla_gan_depto.setItem(i, 2, it_p)
                        self.tabla_gan_depto.setItem(i, 3, create_diff_item(g, g_prev))
                    else:
                        self.tabla_gan_depto.setItem(i, 1, it_v)
                    
            # 7. Formas de Pago
            self.tabla_pago.setRowCount(0)
            self.tabla_pago.setColumnCount(4 if is_comp else 2)
            if is_comp:
                self.tabla_pago.horizontalHeader().setVisible(True)
                self.tabla_pago.setHorizontalHeaderLabels(["Forma de Pago", "Actual", "Anterior", "Crecimiento"])
            else:
                self.tabla_pago.horizontalHeader().setVisible(False)
                
            res_pago = db_manager.execute_query(
                "SELECT COALESCE(v.metodo_pago, 'Efectivo') as pago, substr(v.fecha, 1, 10) as dia, SUM(v.total) as tot "
                "FROM ventas v "
                "WHERE (v.fecha BETWEEN ? AND ?) AND v.estado IN ('COMPLETADA', 'CERRADA') "
                "GROUP BY pago, dia ORDER BY dia ASC", (start_str, end_str)
            )
            pago_sum = {}
            bar_data = {}
            if res_pago:
                for r in res_pago:
                    p = str(r['pago'])
                    d = str(r['dia'])
                    v = float(r['tot'] or 0.0)
                    short_d = d[-2:] if "-" in d else d
                    if short_d not in bar_data: bar_data[short_d] = {}
                    bar_data[short_d][p] = v
                    pago_sum[p] = pago_sum.get(p, 0.0) + v
                    
            self.pago_chart.data = bar_data
            self.pago_chart.update()
            
            for i, (p, v) in enumerate(sorted(pago_sum.items(), key=lambda x: x[1], reverse=True)):
                self.tabla_pago.insertRow(i)
                self.tabla_pago.setItem(i, 0, QTableWidgetItem(p))
                it_v = QTableWidgetItem(f"${v:,.2f}")
                it_v.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                it_v.setFont(QFont("Segoe UI", 10, QFont.Bold))
                
                if is_comp:
                    v_prev = pago_prev_map.get(p, 0)
                    it_p = QTableWidgetItem(f"${v_prev:,.2f}")
                    it_p.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    it_p.setForeground(QBrush(QColor("#94A3B8")))
                    self.tabla_pago.setItem(i, 1, it_v)
                    self.tabla_pago.setItem(i, 2, it_p)
                    self.tabla_pago.setItem(i, 3, create_diff_item(v, v_prev))
                else:
                    self.tabla_pago.setItem(i, 1, it_v)

            # 8. Impuestos Placeholder
            self.tabla_impuestos.setRowCount(1)
            self.tabla_impuestos.setColumnCount(1)
            self.tabla_impuestos.horizontalHeader().setVisible(False)
            it = QTableWidgetItem("- No se registró ninguna venta con impuestos -")
            it.setTextAlignment(Qt.AlignCenter)
            it.setForeground(QBrush(QColor("#94A3B8")))
            self.tabla_impuestos.setItem(0, 0, it)
                
            # Rendimiento de Productos
            if VistaFinanciero._prod_map_cache is None:
                res_prod = db_manager.execute_query("SELECT id, codigo, nombre, departamento, categoria, unidad FROM productos")
                prod_map = {}
                if res_prod:
                    for p in res_prod:
                        pid = str(p['id'])
                        pcod = str(p['codigo'] or '')
                        data = {
                            'nombre': p['nombre'], 
                            'codigo': pcod,
                            'departamento': p['departamento'] or 'ALMACEN',
                            'categoria': p['categoria'] or 'GENERAL',
                            'unidad': p['unidad'] or 'UN'
                        }
                        prod_map[pid] = data
                        if pcod: prod_map[pcod] = data
                VistaFinanciero._prod_map_cache = prod_map
                
                unique_prods = []
                vistos = set()
                for k, data in prod_map.items():
                    dict_id = id(data)
                    if dict_id not in vistos:
                        vistos.add(dict_id)
                        unique_prods.append(data)
                VistaFinanciero._unique_prods_base = unique_prods
                        
            res_vtas_rend = db_manager.execute_query(
                "SELECT dv.id_producto, SUM(dv.subtotal) as rec, SUM(dv.cantidad) as vol "
                "FROM detalles_ventas dv JOIN ventas v ON dv.id_venta = v.id "
                "WHERE v.estado IN ('COMPLETADA', 'CERRADA') AND (v.fecha BETWEEN ? AND ?) "
                "GROUP BY dv.id_producto", (start_str, end_str)
            )
            
            prev_prod_stats = {}
            if is_comp:
                res_prev_prod = db_manager.execute_query(
                    "SELECT dv.id_producto, SUM(dv.subtotal) as rec, SUM(dv.cantidad) as vol "
                    "FROM detalles_ventas dv JOIN ventas v ON dv.id_venta = v.id "
                    "WHERE v.estado IN ('COMPLETADA', 'CERRADA') AND (v.fecha BETWEEN ? AND ?) "
                    "GROUP BY dv.id_producto", (prev_start.strftime("%Y-%m-%d 00:00:00"), prev_end.strftime("%Y-%m-%d 23:59:59"))
                )
                if res_prev_prod:
                    for v in res_prev_prod:
                        prev_prod_stats[str(v['id_producto'])] = {'rec': float(v['rec'] or 0.0), 'vol': float(v['vol'] or 0.0)}
            
            top_sold = []
            sold_ids = set()
            if res_vtas_rend:
                for v in res_vtas_rend:
                    pid = str(v['id_producto'])
                    sold_ids.add(pid)
                    base = VistaFinanciero._prod_map_cache.get(pid)
                    if base:
                        top_sold.append({'pid': pid, 'nombre': base['nombre'], 'codigo': base['codigo'], 'rec': float(v['rec'] or 0.0), 'vol': float(v['vol'] or 0.0)})
            
            unsold_base = []
            for base in VistaFinanciero._unique_prods_base:
                if base['codigo'] not in sold_ids:
                    unsold_base.append(base)
                    if len(unsold_base) >= 100: break
                    
            bottom_sold = [{'pid': b['codigo'], 'nombre': b['nombre'], 'codigo': b['codigo'], 'rec': 0.0, 'vol': 0.0} for b in unsold_base]
            
            # Top Recaudación
            top_sold.sort(key=lambda x: x['rec'], reverse=True)
            display_rec = top_sold[:100] + bottom_sold if len(top_sold) > 0 else bottom_sold
            self.tabla_rec.setRowCount(len(display_rec))
            self.tabla_rec.setColumnCount(5 if is_comp else 3)
            self.tabla_rec.horizontalHeader().setVisible(True)
            if is_comp:
                self.tabla_rec.setHorizontalHeaderLabels(["Producto", "Cant. act.", "Monto act. ($)", "Cant. ant.", "Monto ant. ($)"])
            else:
                self.tabla_rec.setHorizontalHeaderLabels(["Producto", "Cant. act.", "Monto act. ($)"])
            _aplicar_paleta_tabla(self.tabla_rec)
                
            for row, p in enumerate(display_rec):
                self.tabla_rec.setItem(row, 0, QTableWidgetItem(p['nombre']))
                it_vol = QTableWidgetItem(f"{p['vol']:,.0f}")
                it_vol.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tabla_rec.setItem(row, 1, it_vol)
                it_rec = QTableWidgetItem(f"${p['rec']:,.0f}")
                it_rec.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if is_comp:
                    prev = prev_prod_stats.get(p['pid'], {'vol': 0.0, 'rec': 0.0})
                    
                    it_p_vol = QTableWidgetItem(f"{prev['vol']:,.0f}")
                    it_p_vol.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    it_p_vol.setForeground(QBrush(QColor("#94A3B8")))
                    
                    it_p_rec = QTableWidgetItem(f"${prev['rec']:,.0f}")
                    it_p_rec.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    it_p_rec.setForeground(QBrush(QColor("#94A3B8")))
                    
                    self.tabla_rec.setItem(row, 2, it_rec)
                    self.tabla_rec.setItem(row, 3, it_p_vol)
                    self.tabla_rec.setItem(row, 4, it_p_rec)
                else:
                    self.tabla_rec.setItem(row, 2, it_rec)
                
            # Top Volumen
            top_sold.sort(key=lambda x: x['vol'], reverse=True)
            display_vol = top_sold[:100] + bottom_sold if len(top_sold) > 0 else bottom_sold
            self.tabla_vol.setRowCount(len(display_vol))
            self.tabla_vol.setColumnCount(5 if is_comp else 3)
            self.tabla_vol.horizontalHeader().setVisible(True)
            if is_comp:
                self.tabla_vol.setHorizontalHeaderLabels(["Producto", "Cant. act.", "Monto act. ($)", "Cant. ant.", "Monto ant. ($)"])
            else:
                self.tabla_vol.setHorizontalHeaderLabels(["Producto", "Cant. act.", "Monto act. ($)"])
            _aplicar_paleta_tabla(self.tabla_vol)
                
            for row, p in enumerate(display_vol):
                self.tabla_vol.setItem(row, 0, QTableWidgetItem(p['nombre']))
                it_vol = QTableWidgetItem(f"{p['vol']:,.0f}")
                it_vol.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tabla_vol.setItem(row, 1, it_vol)
                it_rec = QTableWidgetItem(f"${p['rec']:,.0f}")
                it_rec.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if is_comp:
                    prev = prev_prod_stats.get(p['pid'], {'vol': 0.0, 'rec': 0.0})
                    
                    it_p_vol = QTableWidgetItem(f"{prev['vol']:,.0f}")
                    it_p_vol.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    it_p_vol.setForeground(QBrush(QColor("#94A3B8")))
                    
                    it_p_rec = QTableWidgetItem(f"${prev['rec']:,.0f}")
                    it_p_rec.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    it_p_rec.setForeground(QBrush(QColor("#94A3B8")))
                    
                    self.tabla_vol.setItem(row, 2, it_rec)
                    self.tabla_vol.setItem(row, 3, it_p_vol)
                    self.tabla_vol.setItem(row, 4, it_p_rec)
                else:
                    self.tabla_vol.setItem(row, 2, it_rec)

            # Resize columns to contents dynamically for all tables
            try:
                for tbl in [self.tabla_vtas_tiempo, self.tabla_vtas_depto, self.tabla_gan_depto, self.tabla_pago]:
                    header = tbl.horizontalHeader()
                    header.setSectionResizeMode(0, QHeaderView.Stretch)
                    for c in range(1, tbl.columnCount()):
                        header.setSectionResizeMode(c, QHeaderView.ResizeToContents)
                
                for tbl in [self.tabla_rec, self.tabla_vol]:
                    header = tbl.horizontalHeader()
                    header.setSectionResizeMode(0, QHeaderView.Stretch)
                    for c in range(1, tbl.columnCount()):
                        header.setSectionResizeMode(c, QHeaderView.ResizeToContents)
            except Exception as e:
                print("Error resizing columns:", e)

            for tbl in (
                self.tabla_vtas_tiempo, self.tabla_vtas_depto, self.tabla_gan_depto,
                self.tabla_pago, self.tabla_rec, self.tabla_vol,
            ):
                _pintar_filas_tabla(tbl)

            # Actualizar IA
            if hasattr(self, 'ai_assistant'):
                from PyQt6.QtCore import QTimer
                self.ai_assistant.lbl_status.show()
                self.ai_assistant.lbl_content.setText("")
                QTimer.singleShot(800, lambda: self.ai_assistant.update_insights(chart_data, pago_sum, donut_data))
                
        except Exception as e:
            import traceback
            print("Error cargando datos de reporte financiero:", traceback.format_exc())
            QMessageBox.critical(self, "Error", f"Ocurrió un error al cargar los datos: {e}")

        # Historial view
        self.historial_view = QWidget()



    def _show_historial_tab(self):
        self.stack_views.setCurrentIndex(2)
        self._update_tab_buttons()
        self._cargar_historial_tickets()
        self._buscar_auditoria()






    def _calcular_comparativa(self):
        idx_mes = self.cmb_audit_mes.currentIndex()
        anio_sel = self.cmb_audit_anio.currentText()
        
        now = datetime.datetime.now()
        
        if idx_mes > 0:
            mes_eval = idx_mes
        else:
            mes_eval = now.month
            
        if anio_sel.isdigit():
            anio_eval = int(anio_sel)
        else:
            anio_eval = now.year
            
        if mes_eval == 1:
            mes_prev = 12
            anio_prev = anio_eval - 1
        else:
            mes_prev = mes_eval - 1
            anio_prev = anio_eval
            
        start_eval = f"{anio_eval:04d}-{mes_eval:02d}-01 00:00:00"
        if mes_eval == 12:
            end_eval = f"{anio_eval+1:04d}-01-01 00:00:00"
        else:
            end_eval = f"{anio_eval:04d}-{mes_eval+1:02d}-01 00:00:00"
            
        start_prev = f"{anio_prev:04d}-{mes_prev:02d}-01 00:00:00"
        if mes_prev == 12:
            end_prev = f"{anio_prev+1:04d}-01-01 00:00:00"
        else:
            end_prev = f"{anio_prev:04d}-{mes_prev+1:02d}-01 00:00:00"
            
        monto_eval = db_manager.execute_scalar(
            "SELECT SUM(total) FROM ventas WHERE fecha >= ? AND fecha < ? AND estado IN ('COMPLETADA','COMPLETADO','CERRADA','CERRADO')",
            (start_eval, end_eval)
        ) or 0.0
        
        monto_prev = db_manager.execute_scalar(
            "SELECT SUM(total) FROM ventas WHERE fecha >= ? AND fecha < ? AND estado IN ('COMPLETADA','COMPLETADO','CERRADA','CERRADO')",
            (start_prev, end_prev)
        ) or 0.0
        
        if monto_prev > 0:
            diff = ((monto_eval - monto_prev) / monto_prev) * 100
        else:
            diff = 100.0 if monto_eval > 0 else 0.0
            
        return monto_eval, diff

    def _actualizar_audit_kpis(self, tot_monto, tot_unidades, tot_kilos):
        while self.audit_kpi_layout.count():
            item = self.audit_kpi_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()
            
        monto_eval, diff = self._calcular_comparativa()
        
        if diff > 0:
            comp_text = f"▲ +{diff:,.1f}%"
            palette_key = "green"
        elif diff < 0:
            comp_text = f"▼ {diff:,.1f}%"
            palette_key = "red"
        else:
            comp_text = "● 0.0%"
            palette_key = "slate"
            
        _PALETTE = {
            "blue":    ("#EFF6FF", "#3B82F6"),
            "green":   ("#ECFDF5", "#10B981"),
            "amber":   ("#FFFBEB", "#F59E0B"),
            "red":     ("#FEF2F2", "#EF4444"),
            "slate":   ("#F8FAFC", "#64748B"),
        }
        
        def build_kpi_card(titulo, valor, p_key, extra_text=None):
            bg, accent = _PALETTE.get(p_key, _PALETTE["slate"])
            f = QFrame()
            f.setObjectName("audit_kpi")
            f.setStyleSheet(f"""
                #audit_kpi {{
                    background: {bg};
                    border-radius: 18px;
                    border: none;
                }}
            """)
            
            h_color = accent.lstrip('#')
            r, g, b = tuple(int(h_color[i:i+2], 16) for i in (0, 2, 4))
            
            
            l = QVBoxLayout(f)
            l.setContentsMargins(18, 14, 18, 14)
            l.setSpacing(4)
            
            h_title = QHBoxLayout()
            h_title.setSpacing(6)
            dot = QFrame()
            dot.setFixedSize(6, 6)
            dot.setStyleSheet(f"background: {accent}; border-radius: 3px; border: none;")
            h_title.addWidget(dot)
            
            lbl_t = QLabel(titulo.upper())
            lbl_t.setStyleSheet("font-size: 10px; font-weight: 800;  border: none; background: none;")
            h_title.addWidget(lbl_t)
            h_title.addStretch()
            l.addLayout(h_title)
            
            lbl_v = QLabel(str(valor))
            lbl_v.setStyleSheet("font-size: 18px; font-weight: 900;  border: none; background: none;")
            l.addWidget(lbl_v)
            
            if extra_text:
                lbl_e = QLabel(extra_text)
                lbl_e.setStyleSheet("font-size: 11px; font-weight: bold;  border: none; background: none;")
                l.addWidget(lbl_e)
                
            return f
            
        self.audit_kpi_layout.addWidget(build_kpi_card("Facturado Filtrado", f"${tot_monto:,.2f}", "green"))
        
        idx_mes = self.cmb_audit_mes.currentIndex()
        mes_nombre = self.cmb_audit_mes.currentText() if idx_mes > 0 else "Este Mes"
        self.audit_kpi_layout.addWidget(build_kpi_card(
            "Comparativa vs Mes Anterior", 
            comp_text, 
            palette_key, 
            extra_text=f"Eval: {mes_nombre}"
        ))
        
        self.audit_kpi_layout.addWidget(build_kpi_card("Artículos (Unidades)", f"{tot_unidades:,.2f} ud", "blue"))
        self.audit_kpi_layout.addWidget(build_kpi_card("Volumen Físico (Peso)", f"{tot_kilos:,.3f} kg", "amber"))

    def _exportar_auditoria(self):
        from datetime import datetime
        nombre_def = f"auditoria_ventas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Exportar auditoría a Excel", nombre_def,
            "Excel (*.xlsx);;Todos los archivos (*)")
        if not filepath: return
        
        row_count = self.table_audit.rowCount()
        if row_count == 0:
            QMessageBox.information(self, "Exportar", "No hay datos en la tabla para exportar.")
            return
            
        class WorkerExportAudit(QThread):
            finished = pyqtSignal(bool, str)
            def __init__(self, path, table_widget):
                super().__init__()
                self.path = path
                self.headers = [table_widget.horizontalHeaderItem(col).text() for col in range(table_widget.columnCount())]
                self.data = []
                for row in range(table_widget.rowCount()):
                    row_data = []
                    for col in range(table_widget.columnCount()):
                        item = table_widget.item(row, col)
                        row_data.append(item.text() if item else "")
                    self.data.append(row_data)
                    
            def run(self):
                try:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Auditoría de Ventas"
                    
                    header_fill = PatternFill("solid", fgColor="059669")
                    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                    border_thin = Border(
                        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
                    )
                    cell_font = Font(name="Segoe UI", size=10)
                    
                    for col_idx, h in enumerate(self.headers, 1):
                        cell = ws.cell(row=1, column=col_idx, value=h)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border_thin
                        
                    for row_idx, row_vals in enumerate(self.data, 2):
                        for col_idx, val in enumerate(row_vals, 1):
                            cell_val = val
                            if col_idx in (6, 8, 9):
                                clean_val = val.replace("$", "").replace(",", "").strip()
                                try:
                                    cell_val = float(clean_val)
                                except:
                                    pass
                                    
                            cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
                            cell.font = cell_font
                            cell.border = border_thin
                            cell.alignment = Alignment(vertical="center")
                            
                            if col_idx in (8, 9):
                                cell.number_format = '"$"#,##0.00'
                                cell.alignment = Alignment(horizontal="right", vertical="center")
                            elif col_idx == 6:
                                u_med = row_vals[6] if len(row_vals) > 6 else "UN"
                                if "KG" in str(u_med).upper():
                                    cell.number_format = '#,##0.000'
                                else:
                                    cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal="right", vertical="center")
                            elif col_idx in (1, 7, 10, 11):
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                
                    for col in range(1, len(self.headers) + 1):
                        ws.column_dimensions[get_column_letter(col)].width = 16
                    ws.column_dimensions['E'].width = 28
                    
                    ws.freeze_panes = "A2"
                    wb.save(self.path)
                    self.finished.emit(True, f"Se exportaron {len(self.data)} filas exitosamente a:\n{self.path}")
                except Exception as e:
                    self.finished.emit(False, str(e))

        self.btn_audit_exportar.setText("⏳ EXPORTANDO...")
        self.btn_audit_exportar.setEnabled(False)
        
        self._worker_exp_audit = WorkerExportAudit(filepath, self.table_audit)
        
        def on_finished(ok, msg):
            self.btn_audit_exportar.setText("📤 EXPORTAR")
            self.btn_audit_exportar.setEnabled(True)
            if ok:
                QMessageBox.information(self, "Exportación Completada", msg)
            else:
                QMessageBox.warning(self, "Error al Exportar", f"No se pudo exportar: {msg}")
                
        self._worker_exp_audit.finished.connect(on_finished)
        self._worker_exp_audit.start()


    def _cargar_historial_tickets(self):
        query = """
            SELECT v.id, v.fecha, v.usuario, p.departamento, dv.nombre_producto,
                   dv.cantidad, p.unidad AS unidad_medida, dv.precio_unitario, dv.subtotal,
                   v.metodo_pago, v.estado
            FROM ventas v
            JOIN detalles_ventas dv ON dv.id_venta = v.id
            LEFT JOIN productos p ON dv.id_producto = p.id
            ORDER BY v.id DESC LIMIT 3000
        """
        rows = db_manager.execute_query(query) or []
        
        filtro = self.txt_hist_buscar.text().strip().lower()
        if filtro:
            rows = [r for r in rows if filtro in str(r.get('id','')) or 
                                      filtro in str(r.get('nombre_producto','')).lower() or
                                      filtro in str(r.get('usuario','')).lower() or
                                      filtro in str(r.get('departamento','')).lower()]
        
        self.tabla_historial_crudo.setRowCount(0)
        from PyQt6.QtGui import QColor, QFont
        
        font_mono = QFont("Consolas", 10, QFont.Bold)
        col_green = QColor("#039855") # Verde oscuro legible
        col_red = QColor("#D92D20")   # Rojo oscuro legible
        col_text = QColor("#1E293B")  # Gris pizarra oscuro
        col_gray = QColor("#64748B")  # Gris pizarra medio
        col_prod = QColor("#4F46E5")  # Indigo oscuro
        
        for i, r in enumerate(rows):
            self.tabla_historial_crudo.insertRow(i)
            
            estado_val = str(r.get('estado', '')).upper()
            subtotal_val = float(r.get('subtotal', 0))
            is_negative = estado_val in ['CANCELADA', 'ANULADA', 'REEMBOLSADA'] or subtotal_val < 0
            
            trade_color = col_red if is_negative else col_green
            sign = "-" if is_negative else "+"
            
            it_id = QTableWidgetItem(str(r.get('id', '')))
            it_id.setTextAlignment(Qt.AlignCenter)
            it_id.setForeground(col_gray)
            
            it_fecha = QTableWidgetItem(str(r.get('fecha', '')))
            it_fecha.setTextAlignment(Qt.AlignCenter)
            it_fecha.setForeground(col_gray)
            
            it_cajero = QTableWidgetItem(str(r.get('usuario', '')).upper())
            it_cajero.setForeground(col_text)
            
            it_depto = QTableWidgetItem(str(r.get('departamento', '')).upper())
            it_depto.setForeground(col_text)
            
            it_prod = QTableWidgetItem(str(r.get('nombre_producto', '')).upper())
            it_prod.setForeground(col_prod) # Destacar el activo
            it_prod.setFont(QFont("Segoe UI", 9, QFont.Bold))
            
            it_cant = QTableWidgetItem(f"{r.get('cantidad', 0):.2f}")
            it_cant.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            it_cant.setFont(font_mono)
            it_cant.setForeground(col_text)
            
            it_um = QTableWidgetItem(str(r.get('unidad_medida', 'UN')))
            it_um.setTextAlignment(Qt.AlignCenter)
            it_um.setForeground(col_gray)
            
            it_punit = QTableWidgetItem(f"{r.get('precio_unitario', 0):.2f}")
            it_punit.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            it_punit.setFont(font_mono)
            it_punit.setForeground(col_gray)
            
            it_subt = QTableWidgetItem(f"{sign} {abs(subtotal_val):.2f}")
            it_subt.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            it_subt.setFont(font_mono)
            it_subt.setForeground(trade_color)
            
            it_pago = QTableWidgetItem(str(r.get('metodo_pago', '')).upper())
            it_pago.setForeground(col_gray)
            it_pago.setTextAlignment(Qt.AlignCenter)
            
            it_estado = QTableWidgetItem(estado_val)
            it_estado.setTextAlignment(Qt.AlignCenter)
            it_estado.setFont(QFont("Segoe UI", 9, QFont.Bold))
            it_estado.setForeground(trade_color)
            
            # Aplicar fondo alternado para máxima estética y legibilidad
            bg = QColor("#ffffff") if i % 2 == 0 else QColor("#F8FAFC")
            for it in [it_id, it_fecha, it_cajero, it_depto, it_prod, it_cant, it_um, it_punit, it_subt, it_pago, it_estado]:
                it.setBackground(bg)
                it.setFont(QFont("Segoe UI", 9))
            
            self.tabla_historial_crudo.setItem(i, 0, it_id)
            self.tabla_historial_crudo.setItem(i, 1, it_fecha)
            self.tabla_historial_crudo.setItem(i, 2, it_cajero)
            self.tabla_historial_crudo.setItem(i, 3, it_depto)
            self.tabla_historial_crudo.setItem(i, 4, it_prod)
            self.tabla_historial_crudo.setItem(i, 5, it_cant)
            self.tabla_historial_crudo.setItem(i, 6, it_um)
            self.tabla_historial_crudo.setItem(i, 7, it_punit)
            self.tabla_historial_crudo.setItem(i, 8, it_subt)
            self.tabla_historial_crudo.setItem(i, 9, it_pago)
            self.tabla_historial_crudo.setItem(i, 10, it_estado)