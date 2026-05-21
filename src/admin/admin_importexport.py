"""
Módulo de Importación / Exportación de Productos
Formato Excel Flexible: Soportando precios, ofertas y lectura robusta de columnas sin importar su posición.
Optimizado para cargas masivas instantáneas (17,000+ registros en lote/WAL).
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

try:
    from src.database import db_manager
except ImportError:
    from database import db_manager

# ── Columnas por Defecto para Exportación ───────
COLS = [
    ("Código",        "id"),
    ("Producto",      "nombre"),
    ("P. Costo",      "costo"),
    ("P. Venta",      "precio"),
    ("P. Mayoreo",    "precio_mayoreo"),
    ("Cant. Oferta",  "cant_oferta"),
    ("P. Oferta",     "precio_oferta"),
    ("Departamento",  "departamento"),
    ("Existencia",    "stock"),
    ("Inv. Mínimo",   "stock_minimo"),
    ("Inv. Máximo",   "stock_maximo"),
    ("Tipo de Venta", "unidad"),      # KG→GRANEL, UN→UNIDAD
]

HEADER_FILL  = PatternFill("solid", fgColor="FFD966")   # amarillo dorado
HEADER_FONT  = Font(bold=True, size=10)
BORDER_THIN  = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

# ─────────────────────────────────────────────────────────
def exportar_excel(filepath: str) -> tuple[bool, str]:
    """Exporta todos los productos al archivo Excel indicado con las columnas completas de ofertas."""
    try:
        rows = db_manager.execute_query(
            "SELECT id, nombre, costo, precio, precio_mayoreo, cant_oferta, precio_oferta, departamento, stock, stock_minimo, stock_maximo, unidad "
            "FROM productos ORDER BY departamento, nombre"
        ) or []

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"

        # ── Encabezado ────────────────────────────────────
        headers = [c[0] for c in COLS]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = BORDER_THIN

        # ── Datos ─────────────────────────────────────────
        for row_idx, r in enumerate(rows, 2):
            tipo = "GRANEL" if (r['unidad'] or '').upper() == 'KG' else "UNIDAD"
            valores = [
                r['id'],
                r['nombre'] or '',
                r['costo'] or 0.0,
                r['precio'] or 0.0,
                r['precio_mayoreo'] or 0.0,
                r['cant_oferta'] or 0.0,
                r['precio_oferta'] or 0.0,
                r['departamento'] or '',
                r['stock'] or 0.0,
                r['stock_minimo'] or 0.0,
                r['stock_maximo'] or 0.0,
                tipo,
            ]
            for col_idx, val in enumerate(valores, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border    = BORDER_THIN
                cell.alignment = Alignment(vertical="center")
                # Formato moneda para precios
                if col_idx in (3, 4, 5, 7):
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx in (6, 9, 10, 11):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        # ── Anchos de columna ────────────────────────────
        anchos = [12, 28, 12, 12, 12, 12, 12, 16, 12, 12, 12, 14]
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        ws.freeze_panes = "A2"
        wb.save(filepath)
        return True, f"Se exportaron {len(rows)} productos a:\n{filepath}"

    except Exception as e:
        return False, f"Error al exportar: {e}"


# ─────────────────────────────────────────────────────────
def importar_excel(filepath: str) -> tuple[bool, str]:
    """
    Importa productos desde un Excel con lectura flexible de columnas por su nombre de encabezado,
    sin importar el orden o posición de la tabla. Optimizado para 17,000+ productos en lote/WAL.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        # Detectar fila de encabezado buscando cualquiera de las columnas principales en las primeras 5 filas
        header_row = 1
        for row in ws.iter_rows(min_row=1, max_row=5):
            row_vals = [str(c.value or '').strip().lower() for c in row]
            if any(h in row_vals for h in ('código', 'codigo', 'cod', 'producto', 'nombre', 'p. venta', 'precio')):
                header_row = row[0].row
                break

        # Leer encabezados para mapear columnas de forma totalmente agnóstica de la posición
        headers = {}
        for cell in ws[header_row]:
            if cell.value is not None:
                headers[str(cell.value).strip().lower()] = cell.column - 1  # índice 0-based

        insertados = 0; actualizados = 0; errores = 0

        # MODO ALTO RENDIMIENTO: Usar 1 sola conexión y 1 sola transacción en memoria
        conn = db_manager.get_connection()
        cursor = conn.cursor()

        # Pre-cargar identificadores existentes para búsqueda O(1) en memoria
        cursor.execute("SELECT id, codigo FROM productos")
        existentes_id = {}
        existentes_cod = {}
        for r in cursor.fetchall():
            existentes_id[str(r['id'])] = r['id']
            if r['codigo']:
                existentes_cod[str(r['codigo'])] = r['id']

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(row): continue  # fila vacía

            def get(key_variants, default=''):
                for k in key_variants:
                    if k in headers:
                        v = row[headers[k]]
                        return v if v is not None else default
                return default

            def parse_float(val):
                if val is None or val == '': return 0.0
                if isinstance(val, (int, float)): return float(val)
                s = str(val).replace('$', '').strip()
                if ',' in s and '.' not in s: s = s.replace(',', '.')
                elif ',' in s and '.' in s: s = s.replace('.', '').replace(',', '.')
                try: return float(s)
                except: return 0.0

            codigo_val  = str(get(['código','codigo','cod'], '')).strip()
            if codigo_val.endswith('.0'):
                codigo_val = codigo_val[:-2]

            nombre      = str(get(['producto','nombre','descripcion','descripción'], '')).strip()
            if not nombre: continue  # saltar filas sin nombre

            costo       = parse_float(get(['p. costo','costo','p.costo'], 0))
            precio      = parse_float(get(['p. venta','precio','p.venta'], 0))
            mayoreo     = parse_float(get(['p. mayoreo','mayoreo','p.mayoreo'], 0))
            cant_of     = parse_float(get(['cant. oferta','cant oferta','cantidad oferta','oferta_cant'], 0))
            precio_of   = parse_float(get(['p. oferta','precio oferta','precio_oferta'], 0))
            stock       = parse_float(get(['existencia','stock','cantidad'], 0))
            minimo      = parse_float(get(['inv. mínimo','inv. minimo','mínimo','minimo'], 0))
            maximo      = parse_float(get(['inv. máximo','inv. maximo','máximo','maximo'], 0))

            depto  = str(get(['departamento','depto'], '')).strip() or None
            tipo   = str(get(['tipo de venta','tipo'], 'UNIDAD')).strip().upper()
            unidad = 'KG' if 'GRANEL' in tipo else 'UN'
            es_pes = 1 if unidad == 'KG' else 0
            cat    = depto.upper() if depto else 'GENERAL'

            # Búsqueda instantánea O(1) en los diccionarios pre-cargados
            existe_id = None
            if codigo_val and codigo_val.isdigit():
                existe_id = existentes_id.get(codigo_val)
            if not existe_id and codigo_val:
                existe_id = existentes_cod.get(codigo_val)

            try:
                if existe_id:
                    cursor.execute(
                        "UPDATE productos SET nombre=?,precio=?,costo=?,stock=?,precio_mayoreo=?,stock_minimo=?,stock_maximo=?,"
                        "unidad=?,es_pesable=?,departamento=?,categoria=?,cant_oferta=?,precio_oferta=? WHERE id=?",
                        (nombre, precio, costo, stock, mayoreo, minimo, maximo, unidad, es_pes, depto, cat, cant_of, precio_of, existe_id))
                    actualizados += 1
                else:
                    params = (codigo_val if not codigo_val.isdigit() else None,
                              nombre, precio, costo, stock, mayoreo, minimo, maximo, unidad, es_pes, depto, cat, cant_of, precio_of)
                    if codigo_val and codigo_val.isdigit():
                        cursor.execute(
                            "INSERT OR IGNORE INTO productos "
                            "(id,codigo,nombre,precio,costo,stock,precio_mayoreo,stock_minimo,stock_maximo,unidad,es_pesable,departamento,categoria,cant_oferta,precio_oferta) "
                            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                            (int(codigo_val), None, nombre, precio, costo, stock, mayoreo, minimo, maximo, unidad, es_pes, depto, cat, cant_of, precio_of))
                        existentes_id[codigo_val] = int(codigo_val)
                    else:
                        cursor.execute(
                            "INSERT INTO productos "
                            "(codigo,nombre,precio,costo,stock,precio_mayoreo,stock_minimo,stock_maximo,unidad,es_pesable,departamento,categoria,cant_oferta,precio_oferta) "
                            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                            params)
                        if codigo_val:
                            existentes_cod[codigo_val] = cursor.lastrowid
                    insertados += 1
            except Exception as ex:
                errores += 1

        # Confirmar todo el lote masivo en una sola ráfaga atómica
        conn.commit()
        conn.close()

        msg = (f"Importación masiva completada en modo Ultra-Rápido:\n"
               f"  ✔ Insertados:   {insertados}\n"
               f"  ✔ Actualizados: {actualizados}\n"
               f"  ✖ Errores:      {errores}")
        return True, msg

    except Exception as e:
        return False, f"Error al importar: {e}"
