from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame, QPushButton, 
    QScrollArea, QGridLayout, QGraphicsDropShadowEffect, QStackedWidget,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QComboBox, QLineEdit, QFileDialog, QMessageBox
)
from PyQt5.QtCore import Qt, pyqtSignal, QRect, QPoint, QSize, QThread
from PyQt5.QtGui import QColor, QFont, QPainter, QBrush, QPen, QLinearGradient, QPolygon, QPainterPath
import datetime

try:
    from src.database import db_manager
except ImportError:
    from database import db_manager

def get_depto_icon(depto_name):
    if not depto_name:
        return "📦"
    name = depto_name.strip().upper()
    if "CARNE" in name or "VACUNO" in name or "RES" in name or "CERDO" in name or "VACUN" in name or "ASADO" in name or "TERNER" in name:
        return "🥩"
    if "AVE" in name or "POLLO" in name or "GRANJA" in name or "POLLER" in name:
        return "🍗"
    if "ACHURA" in name or "CHINCHU" in name or "MENUDE" in name or "RIÑON" in name or "MOLLEJ" in name or "INTESTI" in name:
        return "🍢"
    if "PREPARADO" in name or "ELABORADO" in name or "HAMBUR" in name or "MILANE" in name or "ROTIS" in name:
        return "🍳"
    if "EMBUTIDO" in name or "FIAMBRE" in name or "SALCHI" in name or "CHORI" in name or "JAMON" in name or "SALA" in name or "CHARCU" in name:
        return "🌭"
    if "ALMACEN" in name or "ALMACÉN" in name or "ABARRO" in name or "DESPEN" in name:
        return "🥫"
    if "BEBIDA" in name or "REFRES" in name or "GASEO" in name or "CERVE" in name or "VINO" in name or "TRAGO" in name:
        return "🥤"
    if "VERDU" in name or "FRUTA" in name or "VEGETA" in name or "HORTE" in name:
        return "🥦"
    if "PANAD" in name or "PAN" in name or "FACTU" in name or "FACTUR" in name or "BIZCO" in name:
        return "🍞"
    if "LACTEO" in name or "LÁCTEO" in name or "QUESO" in name or "LECHE" in name or "MANTE" in name or "YOGU" in name:
        return "🧀"
    if "LIMPIE" in name or "HIGIEN" in name or "JABON" in name or "DETER" in name or "PERFU" in name:
        return "🧼"
    if "CONGEL" in name or "HELA" in name:
        return "❄️"
    if "KIOS" in name or "GOLO" in name or "CARAME" in name or "CHOCO" in name:
        return "🍬"
    return "📦"

class ModernCard(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("card")
        self.setStyleSheet("""
            #card {
                background-color: #ffffff;
                border: none;
                border-radius: 20px;
            }
        """)
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(24)
        shadow.setXOffset(0)
        shadow.setYOffset(6)
        shadow.setColor(QColor(0, 0, 0, 15))
        self.setGraphicsEffect(shadow)

class BarChartWidget(QWidget):
    def __init__(self, data=None, parent=None):
        super().__init__(parent)
        self.data = data or {}
        self.colors = {
            "Efectivo": QColor("#3B82F6"),       # Elegant Blue
            "Vales": QColor("#60A5FA"),          # Light Blue
            "Tarjeta": QColor("#EF4444"),        # Red
            "Crédito": QColor("#F59E0B"),        # Amber
            "Transferencia": QColor("#10B981"),  # Emerald
            "Cheque": QColor("#8B5CF6")          # Purple
        }
        self.setMinimumHeight(350)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        w = self.width()
        h = self.height()
        padding_l, padding_r, padding_t, padding_b = 80, 40, 40, 60
        chart_w, chart_h = w - padding_l - padding_r, h - padding_t - padding_b
        
        max_total = 0
        for day, methods in self.data.items():
            max_total = max(max_total, sum(methods.values()))
        if max_total == 0: max_total = 1000
        magnitude = 10**(len(str(int(max_total))) - 1)
        max_total = ((int(max_total) // magnitude) + 1) * magnitude
        
        # Grid lines (very subtle)
        painter.setPen(QPen(QColor("#EEF2F8"), 1))
        painter.setFont(QFont("Segoe UI", 9))
        for i in range(6):
            y = h - padding_b - (i * chart_h / 5)
            painter.drawLine(padding_l, int(y), w - padding_r, int(y))
            val = (max_total / 5) * i
            painter.setPen(QColor("#94A3B8"))
            painter.drawText(QRect(0, int(y - 10), padding_l - 10, 20), Qt.AlignRight | Qt.AlignVCenter, f"${val:,.0f}")
            painter.setPen(QPen(QColor("#EEF2F8"), 1))

        # Bars
        days = list(self.data.keys())
        if not days: return
        bar_w = min(40, (chart_w / len(days)) * 0.6)
        spacing = chart_w / len(days)
        
        for i, day in enumerate(days):
            x = padding_l + (i * spacing) + (spacing - bar_w) / 2
            current_y = h - padding_b
            for method, value in self.data[day].items():
                if value <= 0: continue
                bar_h = (value / max_total) * chart_h
                painter.setBrush(self.colors.get(method, QColor("#94a3b8")))
                painter.setPen(Qt.NoPen)
                painter.drawRoundedRect(int(x), int(current_y - bar_h), int(bar_w), int(bar_h), 3, 3)
                current_y -= bar_h
            # X Label
            painter.setPen(QColor("#64748B"))
            painter.setFont(QFont("Segoe UI", 9, QFont.Bold))
            painter.drawText(QRect(int(x - 20), int(h - padding_b + 10), int(bar_w + 40), 30), Qt.AlignCenter, day)

class DonutChartWidget(QWidget):
    def __init__(self, data=None, parent=None):
        super().__init__(parent)
        self.data = data or {}
        self.colors = ["#3B82F6", "#10B981", "#F59E0B", "#8B5CF6", "#EC4899", "#EF4444", "#06B6D4"]
        self.setMinimumHeight(350)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        size = min(self.width(), self.height()) - 80
        rect = QRect(int((self.width() - size) / 2), int((self.height() - size) / 2), size, size)
        total = sum(self.data.values())
        if total <= 0:
            painter.setBrush(QColor("#EEF2F8"))
            painter.drawEllipse(rect); return
        start_angle = 90 * 16
        for i, (cat, val) in enumerate(self.data.items()):
            if val <= 0: continue
            span_angle = int((val / total) * 360 * 16)
            painter.setBrush(QColor(self.colors[i % len(self.colors)]))
            painter.setPen(Qt.NoPen)
            painter.drawPie(rect, start_angle, -span_angle)
            start_angle -= span_angle
        # Center hole
        inner_size = int(size * 0.72)
        inner_rect = QRect(int((self.width() - inner_size) / 2), int((self.height() - inner_size) / 2), inner_size, inner_size)
        painter.setBrush(QColor("#ffffff"))
        painter.drawEllipse(inner_rect)

class Admin3Reportes(QWidget):
    request_dashboard = pyqtSignal()
    def __init__(self):
        super().__init__()
        self.setup_ui()
        self.cargar_datos("Semana Actual")
        
        # Sincronización en Tiempo Real (Solo para Modo Espectador / Red)
        from src.config import config
        from PyQt5.QtCore import QTimer
        db_path = config.get("db_path", "")
        if db_path and db_path != "":
            self.sync_timer = QTimer(self)
            self.sync_timer.timeout.connect(self.sincronizacion_silenciosa)
            self.sync_timer.start(10000) # Cada 10 segundos

    def sincronizacion_silenciosa(self):
        if not self.isVisible(): return
        
        if self.stack_views.currentIndex() == 1:
            if self.txt_audit_prod.hasFocus(): return
            self._buscar_auditoria()
        else:
            periodo = getattr(self, "current_period", "Semana Actual")
            self.cargar_datos(periodo)

    def setup_ui(self):
        self.setObjectName("Admin3Reportes")
        self.setStyleSheet("""
            QWidget#Admin3Reportes {
                background-color: #F4F6FB;
                font-family: 'Segoe UI', sans-serif;
            }
            QScrollArea { border: none; background: transparent; }
            QWidget#ScrollContainer { background: transparent; }
            QScrollBar:vertical {
                background: transparent; width: 5px;
            }
            QScrollBar::handle:vertical {
                background: rgba(99,102,241,0.25); border-radius: 3px;
            }
            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical { height: 0; }
        """)
        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # ── NAVBAR ────────────────────────────────────────────────────────────
        header = QFrame()
        header.setObjectName("header")
        header.setFixedHeight(64)
        header.setStyleSheet("""
            QFrame#header {
                background: #FFFFFF;
                border-bottom: 1px solid #EEF2F8;
            }
        """)
        h_layout = QHBoxLayout(header)
        h_layout.setContentsMargins(32, 0, 32, 0)
        
        btn_back = QPushButton("← VOLVER")
        btn_back.setCursor(Qt.PointingHandCursor)
        btn_back.setStyleSheet("""
            QPushButton {
                background: #F1F5F9;
                color: #475569;
                font-weight: 700;
                border-radius: 12px;
                padding: 8px 18px;
                font-size: 11px;
                border: none;
            }
            QPushButton:hover {
                background: #E2E8F0;
                color: #0F172A;
            }
        """)
        btn_back.clicked.connect(self.request_dashboard.emit)
        h_layout.addWidget(btn_back)
        h_layout.addSpacing(25)
        
        self.btn_ventas = QPushButton("📈 REPORTE DE VENTAS")
        self.btn_clientes = QPushButton("🔍 AUDITORÍA DE VENTAS")
        self.btn_ventas.setCursor(Qt.PointingHandCursor)
        self.btn_clientes.setCursor(Qt.PointingHandCursor)
        self.btn_ventas.clicked.connect(self._show_ventas_tab)
        self.btn_clientes.clicked.connect(self._show_auditoria_tab)
        
        h_layout.addWidget(self.btn_ventas)
        h_layout.addWidget(self.btn_clientes)
        h_layout.addStretch()
        main_layout.addWidget(header)
  
        # ── CONTENIDO: VENTAS (SCROLL) ────────────────────────────────────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content_widget = QWidget()
        content_widget.setObjectName("ScrollContainer")
        self.content_layout = QVBoxLayout(content_widget)
        self.content_layout.setContentsMargins(32, 24, 32, 32)
        self.content_layout.setSpacing(24)

        top_section = QVBoxLayout()
        top_section.setSpacing(10)
        
        title_section = QVBoxLayout()
        title_section.setSpacing(4)
        lbl_main = QLabel("Reportes & Analíticas")
        lbl_main.setStyleSheet("""
            font-size: 22px; font-weight: 900; color: #0F172A;
            background: transparent; border: none;
            font-family: 'Segoe UI', sans-serif;
        """)
        title_section.addWidget(lbl_main)
        
        self.lbl_main_title = QLabel("SEMANA ACTUAL")
        self.lbl_main_title.setStyleSheet("""
            font-size: 11px; font-weight: 700; color: #6366F1;
            letter-spacing: 0.5px; background: transparent; border: none;
            font-family: 'Segoe UI', sans-serif;
        """)
        title_section.addWidget(self.lbl_main_title)
        top_section.addLayout(title_section)
        
        filters_layout = QHBoxLayout()
        filters_layout.setSpacing(8)
        filters_layout.setContentsMargins(0, 8, 0, 8)
        
        self.period_buttons = {}
        for f_text in ["Semana Actual", "Mes Actual", "Mes Anterior", "Año actual", "Periodo..."]:
            f_btn = QPushButton(f_text)
            f_btn.setCursor(Qt.PointingHandCursor)
            f_btn.clicked.connect(lambda checked, t=f_text: self.cargar_datos(t))
            filters_layout.addWidget(f_btn)
            self.period_buttons[f_text] = f_btn
            
        filters_layout.addStretch()
        top_section.addLayout(filters_layout)
        
        self.kpi_layout = QHBoxLayout()
        self.kpi_layout.setSpacing(15)
        top_section.addSpacing(15)
        top_section.addLayout(self.kpi_layout)
        
        self.content_layout.addLayout(top_section)

        # Gráficos e Información Detallada
        charts_grid = QGridLayout()
        charts_grid.setSpacing(30)
        charts_grid.setContentsMargins(0, 10, 0, 0)
        
        self.card_payments = ModernCard()
        pay_layout = QVBoxLayout(self.card_payments)
        pay_layout.setContentsMargins(25, 25, 25, 25)
        pay_layout.setSpacing(15)
        
        pay_title = QLabel("VENTAS POR PAGO")
        pay_title.setAlignment(Qt.AlignCenter)
        pay_title.setStyleSheet("""
            font-size: 13px; 
            color: #0F172A; 
            font-weight: 800; 
            letter-spacing: 0.5px;
            background: transparent;
            border: none;
            font-family: 'Segoe UI', sans-serif;
        """)
        pay_layout.addWidget(pay_title)
        
        self.bar_chart = BarChartWidget()
        pay_layout.addWidget(self.bar_chart)
        
        self.pay_table_layout = QVBoxLayout()
        self.pay_table_layout.setSpacing(10)
        pay_layout.addLayout(self.pay_table_layout)
        
        charts_grid.addWidget(self.card_payments, 0, 0)
        
        self.card_depts = ModernCard()
        dept_layout = QVBoxLayout(self.card_depts)
        dept_layout.setContentsMargins(25, 25, 25, 25)
        dept_layout.setSpacing(15)
        
        dept_title = QLabel("GANANCIA POR DEPTO")
        dept_title.setAlignment(Qt.AlignCenter)
        dept_title.setStyleSheet("""
            font-size: 13px; 
            color: #0F172A; 
            font-weight: 800; 
            letter-spacing: 0.5px;
            background: transparent;
            border: none;
            font-family: 'Segoe UI', sans-serif;
        """)
        dept_layout.addWidget(dept_title)
        
        donut_h_layout = QHBoxLayout()
        donut_h_layout.setSpacing(20)
        self.donut_chart = DonutChartWidget()
        donut_h_layout.addWidget(self.donut_chart, 3)
        
        self.dept_legend_layout = QVBoxLayout()
        self.dept_legend_layout.setSpacing(8)
        self.dept_legend_layout.setAlignment(Qt.AlignTop)
        donut_h_layout.addLayout(self.dept_legend_layout, 2)
        dept_layout.addLayout(donut_h_layout)
        
        self.dept_table_layout = QVBoxLayout()
        self.dept_table_layout.setSpacing(10)
        dept_layout.addLayout(self.dept_table_layout)
        
        charts_grid.addWidget(self.card_depts, 0, 1)
        self.content_layout.addLayout(charts_grid)
        
        scroll.setWidget(content_widget)
        
        # ── SWITCH VIEWS ──────────────────────────────────────────────────────
        self.stack_views = QStackedWidget()
        self.stack_views.addWidget(scroll)
        
        self.audit_view = QWidget()
        self.audit_view.setStyleSheet("background-color: #F4F6FB;")
        self.setup_audit_ui()
        self.stack_views.addWidget(self.audit_view)
        
        main_layout.addWidget(self.stack_views)
        self._update_tab_buttons()

    def _update_tab_buttons(self):
        active_style = """
            QPushButton {
                background: #EEF2FF;
                color: #6366F1;
                font-weight: 800;
                border-radius: 12px;
                padding: 8px 18px;
                font-size: 12px;
                border: none;
            }
        """
        inactive_style = """
            QPushButton {
                background: transparent;
                color: #64748B;
                font-weight: 700;
                border-radius: 12px;
                padding: 8px 18px;
                font-size: 12px;
                border: none;
            }
            QPushButton:hover {
                background: #F8FAFC;
                color: #334155;
            }
        """
        if self.stack_views.currentIndex() == 0:
            self.btn_ventas.setStyleSheet(active_style)
            self.btn_clientes.setStyleSheet(inactive_style)
        else:
            self.btn_ventas.setStyleSheet(inactive_style)
            self.btn_clientes.setStyleSheet(active_style)

    def cargar_datos(self, periodo="Semana Actual"):
        self.current_period = periodo
        self.lbl_main_title.setText(f"{periodo.upper()}")
        
        # Highlight active period button
        for name, btn in self.period_buttons.items():
            if name == periodo:
                btn.setStyleSheet("""
                    QPushButton {
                        background: #6366F1;
                        color: white;
                        font-weight: 700;
                        font-size: 11px;
                        padding: 6px 14px;
                        border-radius: 12px;
                        border: none;
                    }
                """)
            else:
                btn.setStyleSheet("""
                    QPushButton {
                        background: white;
                        color: #64748B;
                        font-weight: 600;
                        font-size: 11px;
                        padding: 6px 14px;
                        border-radius: 12px;
                        border: none;
                    }
                    QPushButton:hover {
                        background: #E2E8F0;
                        color: #0F172A;
                    }
                """)
                
        today = datetime.date.today()
        if periodo == "Semana Actual":
            start_date = today - datetime.timedelta(days=today.weekday())
            end_date = start_date + datetime.timedelta(days=6)
        elif periodo == "Mes Actual":
            start_date = today.replace(day=1)
            next_month = today.replace(day=28) + datetime.timedelta(days=4)
            end_date = next_month - datetime.timedelta(days=next_month.day)
        elif periodo == "Mes Anterior":
            last_month = today.replace(day=1) - datetime.timedelta(days=1)
            start_date = last_month.replace(day=1)
            end_date = last_month
        elif periodo == "Año actual":
            start_date = today.replace(month=1, day=1)
            end_date = today.replace(month=12, day=31)
        else:
            start_date = today - datetime.timedelta(days=30)
            end_date = today
            
        start_str, end_str = start_date.strftime("%Y-%m-%d 00:00:00"), end_date.strftime("%Y-%m-%d 23:59:59")
        dias_nombres = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]
        metodos = ["Efectivo", "Vales", "Tarjeta", "Crédito", "Transferencia", "Cheque"]
        data_bar = {}
        
        for i in range(7):
            current_day_date = start_date + datetime.timedelta(days=i)
            day_name = dias_nombres[current_day_date.weekday()]
            day_str = current_day_date.strftime("%Y-%m-%d")
            day_data = {}
            for m in metodos:
                val = db_manager.execute_scalar("SELECT SUM(total) FROM ventas WHERE metodo_pago = ? AND (fecha LIKE ?) AND estado IN ('COMPLETADA', 'COMPLETADO', 'CERRADA', 'CERRADO')", (m, f"{day_str}%")) or 0
                day_data[m] = val
            data_bar[day_name] = day_data
            
        self.bar_chart.data = data_bar
        self.bar_chart.update()
        
        while self.pay_table_layout.count():
            item = self.pay_table_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()
            
        for m in ["Efectivo", "Tarjeta", "Transferencia"]:
            total_m = db_manager.execute_scalar("SELECT SUM(total) FROM ventas WHERE metodo_pago = ? AND (fecha BETWEEN ? AND ?) AND estado IN ('COMPLETADA', 'COMPLETADO', 'CERRADA', 'CERRADO')", (m, start_str, end_str)) or 0
            row = QFrame()
            row.setStyleSheet("""
                background: #F8FAFC; 
                border-radius: 12px; 
                border: none;
            """)
            rl = QHBoxLayout(row)
            rl.setContentsMargins(16, 12, 16, 12)
            
            dot_color = "#3B82F6"
            if m.lower() == "tarjeta": dot_color = "#EF4444"
            elif m.lower() == "transferencia": dot_color = "#10B981"
            
            dot = QFrame()
            dot.setFixedSize(8, 8)
            dot.setStyleSheet(f"background: {dot_color}; border-radius: 4px; border: none;")
            rl.addWidget(dot)
            
            lbl_m = QLabel(m.upper())
            lbl_m.setStyleSheet("font-size: 11px; color: #475569; font-weight: 700; background: none; border: none;")
            rl.addWidget(lbl_m)
            
            rl.addStretch()
            lbl_v = QLabel(f"${total_m:,.2f}")
            lbl_v.setStyleSheet("font-size: 13px; font-weight: 800; color: #0F172A; background: none; border: none;")
            rl.addWidget(lbl_v)
            self.pay_table_layout.addWidget(row)
            
        res_dept = db_manager.execute_query(
            "SELECT COALESCE(p.departamento, 'ALMACEN') as depto, SUM(dv.subtotal) as total "
            "FROM detalles_ventas dv "
            "JOIN ventas v ON dv.id_venta = v.id "
            "LEFT JOIN (SELECT id, codigo, departamento FROM productos GROUP BY codigo) p ON (dv.id_producto = CAST(p.id AS TEXT) OR dv.id_producto = p.codigo) "
            "WHERE (v.fecha BETWEEN ? AND ?) AND v.estado IN ('COMPLETADA', 'COMPLETADO', 'CERRADA', 'CERRADO') "
            "GROUP BY depto ORDER BY total DESC", (start_str, end_str)
        )
        
        data_donut = {}
        if res_dept:
            for r in res_dept[:6]:
                name_with_icon = f"{get_depto_icon(r['depto'])} {r['depto'].upper()}"
                data_donut[name_with_icon] = r['total']
        else:
            data_donut = {"🥫 ALMACEN": 1}
            
        self.donut_chart.data = data_donut
        self.donut_chart.update()
        
        while self.dept_legend_layout.count():
            item = self.dept_legend_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
            elif item.layout():
                while item.layout().count():
                    subitem = item.layout().takeAt(0)
                    if subitem.widget(): subitem.widget().deleteLater()
                    
        while self.dept_table_layout.count():
            item = self.dept_table_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()
            
        if res_dept:
            for i, r in enumerate(res_dept[:6]):
                row = QFrame()
                row.setStyleSheet("""
                    background: #F8FAFC; 
                    border-radius: 12px; 
                    border: none;
                """)
                rl = QHBoxLayout(row)
                rl.setContentsMargins(16, 10, 16, 10)
                depto_name = r['depto'].upper()
                icon = get_depto_icon(depto_name)
                
                color = self.donut_chart.colors[i % len(self.donut_chart.colors)]
                dot = QFrame()
                dot.setFixedSize(8, 8)
                dot.setStyleSheet(f"background: {color}; border-radius: 4px; border: none;")
                rl.addWidget(dot)
                
                lbl_c = QLabel(f"{icon} {depto_name}")
                lbl_c.setStyleSheet("font-size: 11px; color: #475569; font-weight: 700; background: none; border: none;")
                rl.addWidget(lbl_c)
                
                rl.addStretch()
                lbl_v = QLabel(f"${r['total']:,.2f}")
                lbl_v.setStyleSheet("font-size: 13px; font-weight: 800; color: #0F172A; background: none; border: none;")
                rl.addWidget(lbl_v)
                self.dept_table_layout.addWidget(row)
                
            for i, r in enumerate(res_dept[:6]):
                row = QHBoxLayout()
                row.setSpacing(6)
                dot = QFrame()
                dot.setFixedSize(8, 8)
                color = self.donut_chart.colors[i % len(self.donut_chart.colors)]
                dot.setStyleSheet(f"background: {color}; border-radius: 4px; border: none;")
                row.addWidget(dot)
                depto_name = r['depto'].upper()
                icon = get_depto_icon(depto_name)
                lbl_text = QLabel(f"{icon} {depto_name} (${r['total']:,.2f})")
                lbl_text.setStyleSheet("font-size: 11px; color: #475569; font-weight: 700; background: none; border: none;")
                row.addWidget(lbl_text)
                self.dept_legend_layout.addLayout(row)

        # ── KPI GENERALES (LIGHT PASTEL) ──────────────────────────────────────
        while self.kpi_layout.count():
            item = self.kpi_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()
            
        res_tot = db_manager.execute_query("SELECT SUM(total) as neto, SUM(total + descuento - recargo) as bruto, COUNT(id) as cant FROM ventas WHERE (fecha BETWEEN ? AND ?) AND estado IN ('COMPLETADA', 'COMPLETADO', 'CERRADA', 'CERRADO')", (start_str, end_str))
        v_neta = float((res_tot[0]['neto'] if res_tot else 0) or 0)
        v_bruta = float((res_tot[0]['bruto'] if res_tot else 0) or 0)
        t_cant = int((res_tot[0]['cant'] if res_tot else 0) or 0)
        t_promedio = (v_neta / t_cant) if t_cant > 0 else 0.0
        
        res_estrella = db_manager.execute_query("SELECT nombre_producto, SUM(subtotal) as tot FROM detalles_ventas dv JOIN ventas v ON dv.id_venta=v.id WHERE (v.fecha BETWEEN ? AND ?) AND v.estado IN ('COMPLETADA', 'COMPLETADO', 'CERRADA', 'CERRADO') GROUP BY id_producto ORDER BY tot DESC LIMIT 1", (start_str, end_str))
        p_estrella = res_estrella[0]['nombre_producto'] if res_estrella else "Ninguno"
        
        _PALETTE = {
            "blue":    ("#EFF6FF", "#3B82F6"),
            "green":   ("#ECFDF5", "#10B981"),
            "amber":   ("#FFFBEB", "#F59E0B"),
            "violet":  ("#F5F3FF", "#8B5CF6"),
            "pink":    ("#FDF2F8", "#EC4899"),
        }
        
        def build_kpi_card(titulo, valor, palette_key):
            bg1, accent = _PALETTE.get(palette_key, _PALETTE["blue"])
            f = QFrame()
            f.setObjectName("kpi_card")
            f.setStyleSheet(f"""
                #kpi_card {{
                    background: {bg1};
                    border-radius: 20px;
                    border: none;
                }}
            """)
            
            h_color = accent.lstrip('#')
            r, g, b = tuple(int(h_color[i:i+2], 16) for i in (0, 2, 4))
            
            sh = QGraphicsDropShadowEffect(f)
            sh.setBlurRadius(16)
            sh.setColor(QColor(r, g, b, 25))
            sh.setOffset(0, 4)
            f.setGraphicsEffect(sh)
            
            l = QVBoxLayout(f)
            l.setContentsMargins(18, 14, 18, 14)
            l.setSpacing(6)
            
            h_title = QHBoxLayout()
            h_title.setSpacing(8)
            dot = QFrame()
            dot.setFixedSize(8, 8)
            dot.setStyleSheet(f"background: {accent}; border-radius: 4px; border: none;")
            h_title.addWidget(dot)
            
            lbl_t = QLabel(titulo.upper())
            lbl_t.setStyleSheet("font-size: 11px; font-weight: 800; color: #475569; border: none; font-family: 'Segoe UI'; letter-spacing: 0.5px; background: none;")
            h_title.addWidget(lbl_t)
            h_title.addStretch()
            l.addLayout(h_title)
            
            lbl_v = QLabel(str(valor))
            lbl_v.setStyleSheet("font-size: 20px; font-weight: 900; color: #0F172A; border: none; font-family: 'Segoe UI'; background: none;")
            l.addWidget(lbl_v)
            return f
            
        self.kpi_layout.addWidget(build_kpi_card("Venta Bruta", f"${v_bruta:,.2f}", "blue"))
        self.kpi_layout.addWidget(build_kpi_card("Venta Neta", f"${v_neta:,.2f}", "green"))
        self.kpi_layout.addWidget(build_kpi_card("Transacciones", f"{t_cant}", "amber"))
        self.kpi_layout.addWidget(build_kpi_card("Ticket Promedio", f"${t_promedio:,.2f}", "violet"))
        self.kpi_layout.addWidget(build_kpi_card("Producto Estrella", p_estrella.upper(), "pink"))

    def _show_ventas_tab(self):
        self.stack_views.setCurrentIndex(0)
        self._update_tab_buttons()

    def _show_auditoria_tab(self):
        self.stack_views.setCurrentIndex(1)
        self._update_tab_buttons()
        self._buscar_auditoria()

    def setup_audit_ui(self):
        lay = QVBoxLayout(self.audit_view)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(20)
        
        filter_card = ModernCard()
        fl = QHBoxLayout(filter_card)
        fl.setContentsMargins(25, 18, 25, 18)
        fl.setSpacing(15)
        
        self.txt_audit_prod = QLineEdit()
        self.txt_audit_prod.setPlaceholderText("🔎 Buscar producto o código...")
        
        self.cmb_audit_mes = QComboBox()
        self.cmb_audit_mes.addItems(["Todos los Meses", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"])
        
        self.cmb_audit_anio = QComboBox()
        self.cmb_audit_anio.addItem("Todos los Años")
        for y in range(2025, 2031):
            self.cmb_audit_anio.addItem(str(y))
            
        input_style = """
            QLineEdit, QComboBox {
                background: #F8FAFC;
                border: 1px solid #E2E8F0;
                border-radius: 10px;
                padding: 8px 12px;
                font-size: 13px;
                color: #1E293B;
                font-family: 'Segoe UI';
            }
            QLineEdit:focus, QComboBox:focus {
                border-color: #6366F1;
                background: #FFFFFF;
            }
        """
        self.txt_audit_prod.setStyleSheet(input_style)
        self.cmb_audit_mes.setStyleSheet(input_style)
        self.cmb_audit_anio.setStyleSheet(input_style)
        
        self.btn_audit_buscar = QPushButton("🔎 FILTRAR")
        self.btn_audit_buscar.setCursor(Qt.PointingHandCursor)
        self.btn_audit_buscar.setStyleSheet("""
            QPushButton {
                background: #6366F1;
                color: white;
                font-weight: 700;
                border-radius: 10px;
                padding: 10px 20px;
                font-size: 13px;
                border: none;
            }
            QPushButton:hover {
                background: #4F46E5;
            }
        """)
        self.btn_audit_buscar.clicked.connect(self._buscar_auditoria)
        
        self.btn_audit_limpiar = QPushButton("🧹 REINICIAR")
        self.btn_audit_limpiar.setCursor(Qt.PointingHandCursor)
        self.btn_audit_limpiar.setStyleSheet("""
            QPushButton {
                background: #F1F5F9;
                color: #64748B;
                font-weight: 700;
                border-radius: 10px;
                padding: 10px 16px;
                font-size: 13px;
                border: none;
            }
            QPushButton:hover {
                background: #E2E8F0;
                color: #0F172A;
            }
        """)
        self.btn_audit_limpiar.clicked.connect(self._limpiar_filtros_audit)
        
        self.btn_audit_exportar = QPushButton("📤 EXPORTAR")
        self.btn_audit_exportar.setCursor(Qt.PointingHandCursor)
        self.btn_audit_exportar.setStyleSheet("""
            QPushButton {
                background: #10B981;
                color: white;
                font-weight: 700;
                border-radius: 10px;
                padding: 10px 20px;
                font-size: 13px;
                border: none;
            }
            QPushButton:hover {
                background: #059669;
            }
        """)
        self.btn_audit_exportar.clicked.connect(self._exportar_auditoria)
        
        lbl_prod = QLabel("Producto:")
        lbl_prod.setStyleSheet("font-weight: 700; color: #475569; font-size: 12px; border: none; background: transparent;")
        lbl_mes = QLabel("Mes:")
        lbl_mes.setStyleSheet("font-weight: 700; color: #475569; font-size: 12px; border: none; background: transparent;")
        lbl_anio = QLabel("Año:")
        lbl_anio.setStyleSheet("font-weight: 700; color: #475569; font-size: 12px; border: none; background: transparent;")
        
        fl.addWidget(lbl_prod, 0)
        fl.addWidget(self.txt_audit_prod, 3)
        fl.addWidget(lbl_mes, 0)
        fl.addWidget(self.cmb_audit_mes, 2)
        fl.addWidget(lbl_anio, 0)
        fl.addWidget(self.cmb_audit_anio, 2)
        fl.addWidget(self.btn_audit_buscar, 0)
        fl.addWidget(self.btn_audit_limpiar, 0)
        fl.addWidget(self.btn_audit_exportar, 0)
        
        lay.addWidget(filter_card)
        
        self.audit_kpi_layout = QHBoxLayout()
        self.audit_kpi_layout.setSpacing(20)
        lay.addLayout(self.audit_kpi_layout)
        
        self.table_audit = QTableWidget()
        self.table_audit.setColumnCount(11)
        self.table_audit.setHorizontalHeaderLabels([
            "ID Venta", "Fecha y Hora", "Cajero", "Departamento", "Producto", "Cant / Peso", "U. Medida", "P. Unitario", "Subtotal", "Pago", "Estado"
        ])
        self.table_audit.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_audit.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_audit.verticalHeader().setVisible(False)
        self.table_audit.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: none;
                gridline-color: #F1F5F9;
                font-size: 12px;
                border-radius: 16px;
            }
            QTableWidget::item {
                padding: 10px;
                border-bottom: 1px solid #F1F5F9;
            }
            QHeaderView::section {
                background-color: #F8FAFC;
                color: #475569;
                font-weight: 800;
                border: none;
                border-bottom: 2px solid #E2E8F0;
                padding: 10px;
                font-size: 12px;
            }
        """)
        self.table_audit.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table_audit.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table_audit.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table_audit.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.table_audit.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        self.table_audit.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeToContents)
        self.table_audit.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeToContents)
        lay.addWidget(self.table_audit)
        
        self.audit_footer = QFrame()
        self.audit_footer.setStyleSheet("""
            QFrame {
                background: white;
                border-radius: 16px;
                border: none;
            }
        """)
        sh = QGraphicsDropShadowEffect(self.audit_footer)
        sh.setBlurRadius(20)
        sh.setColor(QColor(0, 0, 0, 12))
        sh.setOffset(0, 4)
        self.audit_footer.setGraphicsEffect(sh)
        
        foot_main_lay = QVBoxLayout(self.audit_footer)
        foot_main_lay.setContentsMargins(20, 15, 20, 15)
        foot_main_lay.setSpacing(10)
        
        row1 = QHBoxLayout()
        self.lbl_foot_regs = QLabel("Total Transacciones: 0")
        self.lbl_foot_regs.setStyleSheet("font-weight: 700; color: #1E293B; font-size: 12px; background: none; border: none;")
        self.lbl_foot_unidades = QLabel("Unidades Vendidas: 0.00 ud")
        self.lbl_foot_unidades.setStyleSheet("font-weight: 700; color: #475569; font-size: 12px; background: none; border: none;")
        self.lbl_foot_kilos = QLabel("Peso Carne/Aves (Kilos): 0.000 kg")
        self.lbl_foot_kilos.setStyleSheet("font-weight: 700; color: #D97706; font-size: 12px; background: none; border: none;")
        
        row1.addWidget(self.lbl_foot_regs)
        row1.addStretch()
        row1.addWidget(self.lbl_foot_unidades)
        row1.addStretch()
        row1.addWidget(self.lbl_foot_kilos)
        foot_main_lay.addLayout(row1)
        
        row2 = QHBoxLayout()
        self.lbl_foot_carnes = QLabel("🥩 Carnes: $0.00")
        self.lbl_foot_carnes.setStyleSheet("font-weight: 700; color: #EF4444; font-size: 12px; background: none; border: none;")
        self.lbl_foot_aves = QLabel("🍗 Aves: $0.00")
        self.lbl_foot_aves.setStyleSheet("font-weight: 700; color: #F97316; font-size: 12px; background: none; border: none;")
        self.lbl_foot_almacen = QLabel("🥫 Almacén: $0.00")
        self.lbl_foot_almacen.setStyleSheet("font-weight: 700; color: #3B82F6; font-size: 12px; background: none; border: none;")
        self.lbl_foot_monto = QLabel("Facturado Total: $0.00")
        self.lbl_foot_monto.setStyleSheet("font-weight: 900; color: #10B981; font-size: 15px; background: none; border: none;")
        
        row2.addWidget(self.lbl_foot_carnes)
        row2.addStretch()
        row2.addWidget(self.lbl_foot_aves)
        row2.addStretch()
        row2.addWidget(self.lbl_foot_almacen)
        row2.addStretch()
        row2.addWidget(self.lbl_foot_monto)
        foot_main_lay.addLayout(row2)
        
        lay.addWidget(self.audit_footer)

    def _limpiar_filtros_audit(self):
        self.txt_audit_prod.clear()
        self.cmb_audit_mes.setCurrentIndex(0)
        self.cmb_audit_anio.setCurrentIndex(0)
        self._buscar_auditoria()

    def _buscar_auditoria(self):
        prod = self.txt_audit_prod.text().strip()
        idx_mes = self.cmb_audit_mes.currentIndex()
        anio_sel = self.cmb_audit_anio.currentText()
        
        query = """
            SELECT 
                v.id AS id_venta,
                v.fecha,
                v.usuario,
                dv.nombre_producto,
                COALESCE(p.departamento, 'ALMACEN') as depto,
                COALESCE(p.categoria, 'GENERAL') as categoria,
                dv.cantidad,
                dv.precio_unitario,
                dv.subtotal,
                v.metodo_pago,
                v.estado,
                COALESCE(p.unidad, 'UN') as unidad
            FROM detalles_ventas dv
            JOIN ventas v ON dv.id_venta = v.id
            LEFT JOIN (SELECT id, codigo, departamento, categoria, unidad FROM productos GROUP BY codigo) p ON (dv.id_producto = CAST(p.id AS TEXT) OR dv.id_producto = p.codigo)
            WHERE 1=1
        """
        params = []
        
        if prod:
            query += " AND (dv.nombre_producto LIKE ? OR dv.id_producto LIKE ?)"
            params.extend([f"%{prod}%", f"%{prod}%"])
            
        if idx_mes > 0:
            query += " AND strftime('%m', v.fecha) = ?"
            params.append(f"{idx_mes:02d}")
            
        if anio_sel.isdigit():
            query += " AND strftime('%Y', v.fecha) = ?"
            params.append(anio_sel)
            
        query += " ORDER BY v.fecha DESC"
        
        rows = db_manager.execute_query(query, tuple(params)) or []
        
        self.table_audit.setRowCount(0)
        self.table_audit.setRowCount(len(rows))
        
        tot_monto = 0.0
        tot_unidades = 0.0
        tot_kilos = 0.0
        monto_carnes = 0.0
        monto_aves = 0.0
        monto_almacen = 0.0
        
        for i, r in enumerate(rows):
            id_v = str(r['id_venta'])
            fecha = r['fecha'] or ''
            cajero = r['usuario'] or ''
            prod_name = r['nombre_producto'] or ''
            depto = (r['depto'] or 'ALMACEN').strip().upper()
            cant = r['cantidad'] if r['cantidad'] is not None else 0.0
            precio = r['precio_unitario'] if r['precio_unitario'] is not None else 0.0
            subt = r['subtotal'] if r['subtotal'] is not None else 0.0
            pago = r['metodo_pago'] or 'Efectivo'
            estado = r['estado'] or 'COMPLETADA'
            unidad = (r['unidad'] or 'UN').strip().upper()
            
            tot_monto += subt
            
            if unidad == 'KG':
                tot_kilos += cant
                uni_str = "KG"
            else:
                tot_unidades += cant
                uni_str = "UN"
                
            if "CARNE" in depto or "RES" in depto or "CERDO" in depto or "VACUNO" in depto:
                monto_carnes += subt
            elif "AVE" in depto or "POLLO" in depto or "GRANJA" in depto:
                monto_aves += subt
            else:
                monto_almacen += subt
            
            item_id = QTableWidgetItem(id_v); item_id.setTextAlignment(Qt.AlignCenter)
            item_fec = QTableWidgetItem(fecha)
            item_caj = QTableWidgetItem(cajero)
            item_dep = QTableWidgetItem(depto)
            item_prod = QTableWidgetItem(prod_name)
            
            if unidad == 'KG':
                item_cant = QTableWidgetItem(f"{cant:,.3f}")
            else:
                item_cant = QTableWidgetItem(f"{cant:,.2f}")
            item_cant.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            
            item_uni = QTableWidgetItem(uni_str); item_uni.setTextAlignment(Qt.AlignCenter)
            item_prec = QTableWidgetItem(f"${precio:,.2f}"); item_prec.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            item_subt = QTableWidgetItem(f"${subt:,.2f}"); item_subt.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            item_pago = QTableWidgetItem(pago)
            item_est = QTableWidgetItem(estado); item_est.setTextAlignment(Qt.AlignCenter)
            
            bg = QColor("#ffffff") if i % 2 == 0 else QColor("#F8FAFC")
            for col, item in enumerate([item_id, item_fec, item_caj, item_dep, item_prod, item_cant, item_uni, item_prec, item_subt, item_pago, item_est]):
                item.setBackground(bg)
                item.setForeground(QColor("#1E293B"))
                item.setFont(QFont("Segoe UI", 9))
                self.table_audit.setItem(i, col, item)
                
        self.lbl_foot_regs.setText(f"Total Transacciones: {len(rows)}")
        self.lbl_foot_unidades.setText(f"Unidades Vendidas: {tot_unidades:,.2f} ud")
        self.lbl_foot_kilos.setText(f"Peso Carne/Aves (Kilos): {tot_kilos:,.3f} kg")
        self.lbl_foot_carnes.setText(f"🥩 Carnes: ${monto_carnes:,.2f}")
        self.lbl_foot_aves.setText(f"🍗 Aves: ${monto_aves:,.2f}")
        self.lbl_foot_almacen.setText(f"🥫 Almacén: ${monto_almacen:,.2f}")
        self.lbl_foot_monto.setText(f"Facturado Total: ${tot_monto:,.2f}")
        
        self._actualizar_audit_kpis(tot_monto, tot_unidades, tot_kilos)

    def _calcular_comparativa(self):
        idx_mes = self.cmb_audit_mes.currentIndex()
        anio_sel = self.cmb_audit_anio.currentText()
        
        now = datetime.datetime.now()
        
        if idx_mes > 0:
            mes_eval = idx_mes
        else:
            mes_eval = now.month
            
        if anio_sel.isdigit():
            anio_eval = int(anio_sel)
        else:
            anio_eval = now.year
            
        if mes_eval == 1:
            mes_prev = 12
            anio_prev = anio_eval - 1
        else:
            mes_prev = mes_eval - 1
            anio_prev = anio_eval
            
        start_eval = f"{anio_eval:04d}-{mes_eval:02d}-01 00:00:00"
        if mes_eval == 12:
            end_eval = f"{anio_eval+1:04d}-01-01 00:00:00"
        else:
            end_eval = f"{anio_eval:04d}-{mes_eval+1:02d}-01 00:00:00"
            
        start_prev = f"{anio_prev:04d}-{mes_prev:02d}-01 00:00:00"
        if mes_prev == 12:
            end_prev = f"{anio_prev+1:04d}-01-01 00:00:00"
        else:
            end_prev = f"{anio_prev:04d}-{mes_prev+1:02d}-01 00:00:00"
            
        monto_eval = db_manager.execute_scalar(
            "SELECT SUM(total) FROM ventas WHERE fecha >= ? AND fecha < ? AND estado IN ('COMPLETADA','COMPLETADO','CERRADA','CERRADO')",
            (start_eval, end_eval)
        ) or 0.0
        
        monto_prev = db_manager.execute_scalar(
            "SELECT SUM(total) FROM ventas WHERE fecha >= ? AND fecha < ? AND estado IN ('COMPLETADA','COMPLETADO','CERRADA','CERRADO')",
            (start_prev, end_prev)
        ) or 0.0
        
        if monto_prev > 0:
            diff = ((monto_eval - monto_prev) / monto_prev) * 100
        else:
            diff = 100.0 if monto_eval > 0 else 0.0
            
        return monto_eval, diff

    def _actualizar_audit_kpis(self, tot_monto, tot_unidades, tot_kilos):
        while self.audit_kpi_layout.count():
            item = self.audit_kpi_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()
            
        monto_eval, diff = self._calcular_comparativa()
        
        if diff > 0:
            comp_text = f"▲ +{diff:,.1f}%"
            palette_key = "green"
        elif diff < 0:
            comp_text = f"▼ {diff:,.1f}%"
            palette_key = "red"
        else:
            comp_text = "● 0.0%"
            palette_key = "slate"
            
        _PALETTE = {
            "blue":    ("#EFF6FF", "#3B82F6"),
            "green":   ("#ECFDF5", "#10B981"),
            "amber":   ("#FFFBEB", "#F59E0B"),
            "red":     ("#FEF2F2", "#EF4444"),
            "slate":   ("#F8FAFC", "#64748B"),
        }
        
        def build_kpi_card(titulo, valor, p_key, extra_text=None):
            bg, accent = _PALETTE.get(p_key, _PALETTE["slate"])
            f = QFrame()
            f.setObjectName("audit_kpi")
            f.setStyleSheet(f"""
                #audit_kpi {{
                    background: {bg};
                    border-radius: 18px;
                    border: none;
                }}
            """)
            
            h_color = accent.lstrip('#')
            r, g, b = tuple(int(h_color[i:i+2], 16) for i in (0, 2, 4))
            
            sh = QGraphicsDropShadowEffect(f)
            sh.setBlurRadius(16)
            sh.setColor(QColor(r, g, b, 20))
            sh.setOffset(0, 4)
            f.setGraphicsEffect(sh)
            
            l = QVBoxLayout(f)
            l.setContentsMargins(18, 14, 18, 14)
            l.setSpacing(4)
            
            h_title = QHBoxLayout()
            h_title.setSpacing(6)
            dot = QFrame()
            dot.setFixedSize(6, 6)
            dot.setStyleSheet(f"background: {accent}; border-radius: 3px; border: none;")
            h_title.addWidget(dot)
            
            lbl_t = QLabel(titulo.upper())
            lbl_t.setStyleSheet("font-size: 10px; font-weight: 800; color: #475569; border: none; background: none;")
            h_title.addWidget(lbl_t)
            h_title.addStretch()
            l.addLayout(h_title)
            
            lbl_v = QLabel(str(valor))
            lbl_v.setStyleSheet("font-size: 18px; font-weight: 900; color: #0F172A; border: none; background: none;")
            l.addWidget(lbl_v)
            
            if extra_text:
                lbl_e = QLabel(extra_text)
                lbl_e.setStyleSheet("font-size: 11px; font-weight: bold; color: #64748B; border: none; background: none;")
                l.addWidget(lbl_e)
                
            return f
            
        self.audit_kpi_layout.addWidget(build_kpi_card("Facturado Filtrado", f"${tot_monto:,.2f}", "green"))
        
        idx_mes = self.cmb_audit_mes.currentIndex()
        mes_nombre = self.cmb_audit_mes.currentText() if idx_mes > 0 else "Este Mes"
        self.audit_kpi_layout.addWidget(build_kpi_card(
            "Comparativa vs Mes Anterior", 
            comp_text, 
            palette_key, 
            extra_text=f"Eval: {mes_nombre}"
        ))
        
        self.audit_kpi_layout.addWidget(build_kpi_card("Artículos (Unidades)", f"{tot_unidades:,.2f} ud", "blue"))
        self.audit_kpi_layout.addWidget(build_kpi_card("Volumen Físico (Peso)", f"{tot_kilos:,.3f} kg", "amber"))

    def _exportar_auditoria(self):
        from datetime import datetime
        nombre_def = f"auditoria_ventas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Exportar auditoría a Excel", nombre_def,
            "Excel (*.xlsx);;Todos los archivos (*)")
        if not filepath: return
        
        row_count = self.table_audit.rowCount()
        if row_count == 0:
            QMessageBox.information(self, "Exportar", "No hay datos en la tabla para exportar.")
            return
            
        class WorkerExportAudit(QThread):
            finished = pyqtSignal(bool, str)
            def __init__(self, path, table_widget):
                super().__init__()
                self.path = path
                self.headers = [table_widget.horizontalHeaderItem(col).text() for col in range(table_widget.columnCount())]
                self.data = []
                for row in range(table_widget.rowCount()):
                    row_data = []
                    for col in range(table_widget.columnCount()):
                        item = table_widget.item(row, col)
                        row_data.append(item.text() if item else "")
                    self.data.append(row_data)
                    
            def run(self):
                try:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Auditoría de Ventas"
                    
                    header_fill = PatternFill("solid", fgColor="059669")
                    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                    border_thin = Border(
                        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
                    )
                    cell_font = Font(name="Segoe UI", size=10)
                    
                    for col_idx, h in enumerate(self.headers, 1):
                        cell = ws.cell(row=1, column=col_idx, value=h)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border_thin
                        
                    for row_idx, row_vals in enumerate(self.data, 2):
                        for col_idx, val in enumerate(row_vals, 1):
                            cell_val = val
                            if col_idx in (6, 8, 9):
                                clean_val = val.replace("$", "").replace(",", "").strip()
                                try:
                                    cell_val = float(clean_val)
                                except:
                                    pass
                                    
                            cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
                            cell.font = cell_font
                            cell.border = border_thin
                            cell.alignment = Alignment(vertical="center")
                            
                            if col_idx in (8, 9):
                                cell.number_format = '"$"#,##0.00'
                                cell.alignment = Alignment(horizontal="right", vertical="center")
                            elif col_idx == 6:
                                u_med = row_vals[6] if len(row_vals) > 6 else "UN"
                                if "KG" in str(u_med).upper():
                                    cell.number_format = '#,##0.000'
                                else:
                                    cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal="right", vertical="center")
                            elif col_idx in (1, 7, 10, 11):
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                
                    for col in range(1, len(self.headers) + 1):
                        ws.column_dimensions[get_column_letter(col)].width = 16
                    ws.column_dimensions['E'].width = 28
                    
                    ws.freeze_panes = "A2"
                    wb.save(self.path)
                    self.finished.emit(True, f"Se exportaron {len(self.data)} filas exitosamente a:\n{self.path}")
                except Exception as e:
                    self.finished.emit(False, str(e))

        self.btn_audit_exportar.setText("⏳ EXPORTANDO...")
        self.btn_audit_exportar.setEnabled(False)
        
        self._worker_exp_audit = WorkerExportAudit(filepath, self.table_audit)
        
        def on_finished(ok, msg):
            self.btn_audit_exportar.setText("📤 EXPORTAR")
            self.btn_audit_exportar.setEnabled(True)
            if ok:
                QMessageBox.information(self, "Exportación Completada", msg)
            else:
                QMessageBox.warning(self, "Error al Exportar", f"No se pudo exportar: {msg}")
                
        self._worker_exp_audit.finished.connect(on_finished)
        self._worker_exp_audit.start()
