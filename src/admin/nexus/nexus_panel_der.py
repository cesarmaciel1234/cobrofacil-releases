import os
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QComboBox, 
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView, 
    QAbstractItemView, QFrame, QMessageBox, QFileDialog, QStackedWidget
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QFont
from datetime import datetime
from src.base_de_datos.database import db_manager

class NexusPanelDer(QFrame):
    def __init__(self):
        super().__init__()
        self.setStyleSheet("background-color: transparent;")
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # --- HEADER PERSONALIZADO (5 Pestañas) ---
        header_layout = QHBoxLayout()
        header_layout.setContentsMargins(0, 0, 0, 0)
        header_layout.setSpacing(2)
        
        # Pestañas (Las 5 juntas)
        self.btn_tab1 = QPushButton("👁️ EN VIVO")
        self.btn_tab2 = QPushButton("🏁 CIERRES")
        self.btn_tab3 = QPushButton("🚨 ALERTAS")
        self.btn_tab4 = QPushButton("👤 CONTROL")
        self.btn_tab5 = QPushButton("💰 CAJONES")
        
        self.tabs = [self.btn_tab1, self.btn_tab2, self.btn_tab3, self.btn_tab4, self.btn_tab5]
        for idx, btn in enumerate(self.tabs):
            btn.setCursor(Qt.PointingHandCursor)
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #1E293B;
                    color: #94A3B8;
                    border: none;
                    border-top-left-radius: 8px;
                    border-top-right-radius: 8px;
                    padding: 10px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #334155;
                    color: white;
                }
            """)
            btn.clicked.connect(lambda checked, i=idx: self.switch_tab(i))
            header_layout.addWidget(btn)
            
        header_layout.addStretch()
        main_layout.addLayout(header_layout)
        
        # --- CONTENIDO (QStackedWidget) ---
        self.stack = QStackedWidget()
        self.stack.setStyleSheet("QStackedWidget { background: #0F172A; border-radius: 8px; border-top-left-radius: 0px; }")
        main_layout.addWidget(self.stack)
        
        # Tab 1: Auditoría
        tab_auditoria = QWidget()
        tab_auditoria.setStyleSheet("background-color: #0F172A; border-radius: 8px;")
        layout_inf = QVBoxLayout(tab_auditoria)
        layout_inf.setContentsMargins(15, 15, 15, 15)
        layout_inf.setSpacing(12)
        

        filt_bar = QHBoxLayout()
        filt_bar.setSpacing(10)
        
        self.txt_buscar = QLineEdit()
        self.txt_buscar.setPlaceholderText("🔍 Buscar por usuario, ID, tipo o descripción...")
        self.txt_buscar.setMinimumWidth(220)
        self.txt_buscar.setStyleSheet("""
            QLineEdit {
                background: #1E293B;
                border: 1px solid #334155;
                border-radius: 10px;
                padding: 8px 12px;
                font-size: 13px;
                color: #F8FAFC;
            }
            QLineEdit:focus {
                border-color: #38BDF8;
                background: #0F172A;
            }
        """)
        self.txt_buscar.textChanged.connect(self.filtrar_auditoria)
        filt_bar.addWidget(self.txt_buscar)
        
        self.cmb_tipo_evento = QComboBox()
        self.cmb_tipo_evento.addItems([
            "Todos los Eventos",
            "Brechas de Seguridad",
            "Intervenciones Supervisor",
            "Turnos y Aperturas",
            "Cancelaciones de Tickets",
            "Ingresos y Retiros de Efectivo"
        ])
        self.cmb_tipo_evento.setStyleSheet("""
            QComboBox {
                background: #1E293B;
                border: 1px solid #334155;
                border-radius: 10px;
                padding: 8px 12px;
                font-size: 13px;
                color: #F8FAFC;
            }
            QComboBox:focus {
                border-color: #38BDF8;
                background: #0F172A;
            }
            QComboBox QAbstractItemView {
                background: #1E293B;
                color: #F8FAFC;
                selection-background-color: #38BDF8;
            }
        """)
        self.cmb_tipo_evento.currentIndexChanged.connect(self.filtrar_auditoria)
        filt_bar.addWidget(self.cmb_tipo_evento)
        
        btn_exportar = QPushButton("📥 EXPORTAR BITÁCORA")
        btn_exportar.setCursor(Qt.PointingHandCursor)
        btn_exportar.setStyleSheet("""
            QPushButton {
                background: #10B981;
                color: white;
                font-weight: 700;
                border-radius: 10px;
                padding: 8px 20px;
                font-size: 13px;
                border: none;
            }
            QPushButton:hover {
                background: #059669;
            }
        """)
        btn_exportar.clicked.connect(self._exportar_auditoria_excel)
        filt_bar.addWidget(btn_exportar)
        
        layout_inf.addLayout(filt_bar)
        
        self.tabla_eventos = QTableWidget()
        self.tabla_eventos.setColumnCount(5)
        self.tabla_eventos.setHorizontalHeaderLabels(["💻 PC", "📅 FECHA / HORA", "🚨 EVENTO DE AUDITORÍA", "👤 USUARIO", "📝 DETALLE / OBSERVACIONES"])
        self.tabla_eventos.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabla_eventos.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabla_eventos.verticalHeader().setVisible(False)
        
        self.tabla_eventos.setStyleSheet("""
            QTableWidget {
                background-color: #0F172A;
                border: none;
                gridline-color: #1E293B;
                font-size: 12px;
                border-radius: 16px;
                color: #F8FAFC;
            }
            QTableWidget::item {
                padding: 10px;
                border-bottom: 1px solid #1E293B;
            }
            QHeaderView::section {
                background-color: #1E293B;
                color: #38BDF8;
                font-weight: bold;
                padding: 12px;
                border: none;
                border-bottom: 2px solid #334155;
            }
            QScrollBar:vertical {
                border: none;
                background: #0F172A;
                width: 10px;
                margin: 0px 0px 0px 0px;
            }
            QScrollBar::handle:vertical {
                background: #334155;
                min-height: 20px;
                border-radius: 5px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
        """)
        
        hh_e = self.tabla_eventos.horizontalHeader()
        hh_e.setSectionResizeMode(0, QHeaderView.Interactive)
        self.tabla_eventos.setColumnWidth(0, 90)
        hh_e.setSectionResizeMode(1, QHeaderView.Interactive)
        self.tabla_eventos.setColumnWidth(1, 170)
        hh_e.setSectionResizeMode(2, QHeaderView.Interactive)
        self.tabla_eventos.setColumnWidth(2, 220)
        hh_e.setSectionResizeMode(3, QHeaderView.Interactive)
        self.tabla_eventos.setColumnWidth(3, 110)
        hh_e.setSectionResizeMode(4, QHeaderView.Stretch)
        
        self.tabla_eventos.verticalScrollBar().valueChanged.connect(self._al_hacer_scroll)
        layout_inf.addWidget(self.tabla_eventos)

        self.stack.addWidget(tab_auditoria)

        # Tab 2: Cierres Z
        self._setup_tab_cierres()

        self.active_query = ""
        self.active_params = []
        self.offset = 0
        self.total_logs_count = 0
        self.caja_filter = 0  # 0 = TODAS LAS CAJAS

        # Cargar inicial
        self.filtrar_auditoria()
        self.cargar_cierres()
        
        # Iniciar en la primera pestaña
        self.switch_tab(0)

    def _setup_tab_cierres(self):
        tab_cierres = QWidget()
        tab_cierres.setStyleSheet("background-color: white; border-radius: 8px;")
        layout_c = QVBoxLayout(tab_cierres)
        layout_c.setContentsMargins(15, 15, 15, 15)
        
        hdr_cierres = QHBoxLayout()
        lbl_c_title = QLabel("🏁 HISTORIAL DE CIERRES Z DE TERMINALES")
        lbl_c_title.setStyleSheet("font-weight: 900; color: #0F172A; font-size: 13px; border: none; background: transparent;")
        hdr_cierres.addWidget(lbl_c_title)
        hdr_cierres.addStretch()
        layout_c.addLayout(hdr_cierres)
        
        self.tabla_cierres = QTableWidget()
        self.tabla_cierres.setColumnCount(6)
        self.tabla_cierres.setHorizontalHeaderLabels([
            "💻 PC / CAJA", 
            "⏰ HORARIO INICIO", 
            "🏁 HORARIO DE CIERRE", 
            "👤 CAJERO", 
            "⚖️ DIFERENCIA", 
            "💵 FÍSICO / ESPERADO"
        ])
        self.tabla_cierres.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabla_cierres.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabla_cierres.verticalHeader().setVisible(False)
        
        self.tabla_cierres.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: none;
                gridline-color: #F1F5F9;
                font-size: 12px;
                border-radius: 16px;
                color: #1E293B;
            }
            QTableWidget::item {
                padding: 10px;
                border-bottom: 1px solid #F1F5F9;
            }
            QHeaderView::section {
                background-color: #F8FAFC;
                color: #475569;
                font-weight: 800;
                border: none;
                border-bottom: 2px solid #E2E8F0;
                padding: 10px;
                font-size: 11px;
            }
        """)
        
        hh_c = self.tabla_cierres.horizontalHeader()
        hh_c.setSectionResizeMode(0, QHeaderView.Interactive)
        self.tabla_cierres.setColumnWidth(0, 90)
        hh_c.setSectionResizeMode(1, QHeaderView.Stretch)
        hh_c.setSectionResizeMode(2, QHeaderView.Stretch)
        hh_c.setSectionResizeMode(3, QHeaderView.Interactive)
        self.tabla_cierres.setColumnWidth(3, 100)
        hh_c.setSectionResizeMode(4, QHeaderView.Interactive)
        hh_c.setSectionResizeMode(4, QHeaderView.Stretch)
        hh_c.setSectionResizeMode(5, QHeaderView.Stretch)
        
        layout_c.addWidget(self.tabla_cierres)
        self.stack.addWidget(tab_cierres)

    def cargar_cierres(self):
        def fmt_moneda(val):
            return f"${val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            
    def switch_tab(self, index):
        # Actualizar estilos de las pestañas
        for i, btn in enumerate(self.tabs):
            if i == index:
                btn.setStyleSheet("""
                    QPushButton {
                        background-color: white; color: #6366F1; font-weight: 900;
                        border-top-left-radius: 12px; border-top-right-radius: 12px;
                        border: none; padding: 12px 20px; font-size: 13px;
                    }
                """)
            else:
                btn.setStyleSheet("""
                    QPushButton {
                        background-color: #E2E8F0; color: #64748B; font-weight: 800;
                        border-top-left-radius: 8px; border-top-right-radius: 8px;
                        border: none; padding: 12px 20px; font-size: 12px;
                    }
                    QPushButton:hover { background-color: #CBD5E1; color: #475569; }
                """)
        
        # Lógica de cambio de vista y filtros
        if index == 0:
            self.stack.setCurrentIndex(0)
            self.cmb_tipo_evento.setCurrentIndex(0) # Eventos en Vivo -> Todos
        elif index == 1:
            self.stack.setCurrentIndex(1) # Historial Cierres
        elif index == 2:
            self.stack.setCurrentIndex(0)
            self.cmb_tipo_evento.setCurrentIndex(1) # Alertas
        elif index == 3:
            self.stack.setCurrentIndex(0)
            self.cmb_tipo_evento.setCurrentIndex(2) # Intervención
        elif index == 4:
            self.stack.setCurrentIndex(0)
            self.cmb_tipo_evento.setCurrentIndex(3) # Apertura Cajón

    def set_master_reference(self, admin_window):
        # Permite al panel invocar funciones del contenedor principal
        self.admin_master = admin_window

    def update_historical_z_combo(self):
        # Recargar fechas únicas de cortes Z
        fechas = db_manager.execute_query("SELECT DISTINCT DATE(fecha) as d FROM cortes_z ORDER BY d DESC")
        self.combo_fecha.clear()
        if fechas:
            for f in fechas:
                self.combo_fecha.addItem(str(f['d']))
        else:
            self.combo_fecha.addItem(datetime.now().strftime("%Y-%m-%d"))
            
        try:
            cierres = db_manager.execute_query("""
                SELECT 
                    c.id, 
                    c.fecha as fecha_cierre, 
                    c.usuario, 
                    c.monto, 
                    c.observaciones, 
                    c.caja_id,
                    (
                        SELECT a.fecha 
                        FROM movimientos_caja a 
                        WHERE a.tipo='APERTURA' 
                          AND a.caja_id = c.caja_id 
                          AND a.fecha < c.fecha 
                        ORDER BY a.fecha DESC 
                        LIMIT 1
                    ) as fecha_inicio
                FROM movimientos_caja c
                WHERE c.tipo='CIERRE_Z'
                ORDER BY c.id DESC
            """) or []
            self.tabla_cierres.setRowCount(len(cierres))
            for i, r in enumerate(cierres):
                # Col 0: PC ID (Caja ID)
                try:
                    c_id_val = int(r['caja_id'])
                    item_pc = QTableWidgetItem(f"0{c_id_val}" if c_id_val < 10 else str(c_id_val))
                except:
                    item_pc = QTableWidgetItem(str(r['caja_id'] or '01'))
                item_pc.setTextAlignment(Qt.AlignCenter)
                item_pc.setFont(QFont("Consolas", 10, QFont.Bold))
                self.tabla_cierres.setItem(i, 0, item_pc)
                
                # Col 1: Fecha/Hora Inicio
                f_ini_val = str(r['fecha_inicio']) if r['fecha_inicio'] else "-"
                item_ini = QTableWidgetItem(f_ini_val)
                item_ini.setTextAlignment(Qt.AlignCenter)
                self.tabla_cierres.setItem(i, 1, item_ini)
                
                # Col 2: Fecha/Hora Cierre
                item_fin = QTableWidgetItem(str(r['fecha_cierre']))
                item_fin.setTextAlignment(Qt.AlignCenter)
                self.tabla_cierres.setItem(i, 2, item_fin)
                
                # Col 3: Cajero
                item_usr = QTableWidgetItem(str(r['usuario']).upper())
                item_usr.setTextAlignment(Qt.AlignCenter)
                self.tabla_cierres.setItem(i, 3, item_usr)
                
                # Col 4: Diferencia (monto)
                dif = float(r['monto'] or 0.0)
                item_dif = QTableWidgetItem(fmt_moneda(dif))
                item_dif.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if dif < 0:
                    item_dif.setForeground(QColor("#ef4444"))
                    item_dif.setBackground(QColor("#fef2f2"))
                elif dif > 0:
                    item_dif.setForeground(QColor("#10b981"))
                    item_dif.setBackground(QColor("#f0fdf4"))
                else:
                    item_dif.setForeground(QColor("#1e293b"))
                self.tabla_cierres.setItem(i, 4, item_dif)

                # Col 5: Detalles (Físico / Esperado)
                obs = str(r['observaciones'] or '')
                detalles = "-"
                if "Fis:" in obs and "Esp:" in obs:
                    try:
                        fis_val = obs.split("Fis:")[1].split("|")[0].strip()
                        esp_val = obs.split("Esp:")[1].strip()
                        detalles = f"{fis_val} / {esp_val}"
                    except:
                        detalles = obs
                else:
                    detalles = obs
                
                item_det = QTableWidgetItem(detalles)
                item_det.setTextAlignment(Qt.AlignCenter)
                self.tabla_cierres.setItem(i, 5, item_det)
        except Exception as e:
            print(f"Error cargando logs BD: {e}")

    def aplicar_tema(self, is_dark):
        if is_dark:
            self._tema_oscuro()
        else:
            self._tema_claro()

    def _tema_oscuro(self):
        self.stack.setStyleSheet("QStackedWidget { background: #0F172A; border-radius: 8px; border-top-left-radius: 0px; }")
        self.stack.widget(0).setStyleSheet("background-color: #0F172A; border-radius: 8px;")
        
        # Tabs
        for btn in self.tabs:
            btn.setStyleSheet("""
                QPushButton { background-color: #1E293B; color: #94A3B8; border: none; border-top-left-radius: 8px; border-top-right-radius: 8px; padding: 10px; font-weight: bold; }
                QPushButton:hover { background-color: #334155; color: white; }
            """)
        
        # Inputs
        self.txt_buscar.setStyleSheet("QLineEdit { background: #1E293B; border: 1px solid #334155; border-radius: 10px; padding: 8px 12px; font-size: 13px; color: #F8FAFC; } QLineEdit:focus { border-color: #38BDF8; background: #0F172A; }")
        self.cmb_tipo_evento.setStyleSheet("QComboBox { background: #1E293B; border: 1px solid #334155; border-radius: 10px; padding: 8px 12px; font-size: 13px; color: #F8FAFC; } QComboBox:focus { border-color: #38BDF8; background: #0F172A; } QComboBox QAbstractItemView { background: #1E293B; color: #F8FAFC; selection-background-color: #38BDF8; }")
        
        # Tabla
        self.tabla_eventos.setStyleSheet("""
            QTableWidget { background-color: #0F172A; border: none; gridline-color: #1E293B; font-size: 12px; border-radius: 16px; color: #F8FAFC; }
            QTableWidget::item { padding: 10px; border-bottom: 1px solid #1E293B; }
            QHeaderView::section { background-color: #1E293B; color: #38BDF8; font-weight: bold; padding: 12px; border: none; border-bottom: 2px solid #334155; }
            QScrollBar:vertical { border: none; background: #0F172A; width: 10px; margin: 0px 0px 0px 0px; }
            QScrollBar::handle:vertical { background: #334155; min-height: 20px; border-radius: 5px; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0px; }
        """)

    def _tema_claro(self):
        self.stack.setStyleSheet("QStackedWidget { background: white; border-radius: 8px; border-top-left-radius: 0px; }")
        self.stack.widget(0).setStyleSheet("background-color: white; border-radius: 8px;")
        
        # Tabs
        for btn in self.tabs:
            btn.setStyleSheet("""
                QPushButton { background-color: transparent; color: #64748B; border: none; border-top-left-radius: 8px; border-top-right-radius: 8px; padding: 10px; font-weight: bold; }
                QPushButton:hover { background-color: #F8FAFC; color: #1E293B; }
            """)
        
        # Inputs
        self.txt_buscar.setStyleSheet("QLineEdit { background: #F8FAFC; border: 1px solid #E2E8F0; border-radius: 10px; padding: 8px 12px; font-size: 13px; color: #1E293B; } QLineEdit:focus { border-color: #6366F1; background: #FFFFFF; }")
        self.cmb_tipo_evento.setStyleSheet("QComboBox { background: #F8FAFC; border: 1px solid #E2E8F0; border-radius: 10px; padding: 8px 12px; font-size: 13px; color: #1E293B; } QComboBox:focus { border-color: #6366F1; background: #FFFFFF; }")
        
        # Tabla
        self.tabla_eventos.setStyleSheet("""
            QTableWidget { background-color: white; border: none; gridline-color: #F1F5F9; font-size: 12px; border-radius: 16px; color: #1E293B; }
            QTableWidget::item { padding: 10px; border-bottom: 1px solid #F1F5F9; }
            QHeaderView::section { background-color: #F8FAFC; color: #64748B; font-weight: bold; padding: 12px; border: none; border-bottom: 2px solid #E2E8F0; }
        """)

    def agregar_log(self, src, payload, fg_color):
        # Esta funcion es llamada desde admin7_nexus por el motor en vivo
        # Como es una tabla estructurada de 6 columnas, vamos a insertar el log
        # en la fila 0 simulando la vista
        row = 0
        self.tabla_eventos.insertRow(row)
        ts = datetime.now().strftime("%m-%d %H:%M")
        
        # Parseamos el payload básico: ej. "[VENTA] Cobro Efectivo - $100.00"
        tipo_evento = "📝 EVENTO"
        icon = "📝"
        if "[VENTA]" in payload:
            tipo_evento = "💰 VENTA"
            icon = "💰"
        elif "ALERTA" in payload or "CRITICAL" in payload:
            tipo_evento = "🚨 SEGURIDAD"
            icon = "🚨"

        # Col 0: PC
        pc_clean = src
        import re
        match = re.search(r'PC-(\d+)', src)
        if match:
            pc_clean = f"PC-{match.group(1)}"
        else:
            match_digit = re.search(r'\d+', src)
            if match_digit:
                c_num = int(match_digit.group())
                pc_clean = f"PC-0{c_num}" if c_num < 10 else f"PC-{c_num}"
            else:
                pc_clean = "PC-01"

        it_pc = QTableWidgetItem(pc_clean)
        it_pc.setTextAlignment(Qt.AlignCenter)
        it_pc.setFont(QFont("Segoe UI", 9, QFont.Bold))
        self.tabla_eventos.setItem(row, 0, it_pc)

        # Col 1: Fecha/Hora
        it_f = QTableWidgetItem(ts)
        it_f.setTextAlignment(Qt.AlignCenter)
        self.tabla_eventos.setItem(row, 1, it_f)

        # Col 2: Evento
        it_ev = QTableWidgetItem(f"{icon} {tipo_evento}")
        it_ev.setTextAlignment(Qt.AlignCenter)
        it_ev.setForeground(QColor(fg_color))
        it_ev.setFont(QFont("Segoe UI", 9, QFont.Bold))
        self.tabla_eventos.setItem(row, 2, it_ev)

        # Col 3: Usuario
        it_u = QTableWidgetItem("SISTEMA")
        it_u.setTextAlignment(Qt.AlignCenter)
        self.tabla_eventos.setItem(row, 3, it_u)

        # Col 4: Payload
        self.tabla_eventos.setItem(row, 4, QTableWidgetItem(payload))

        if self.tabla_eventos.rowCount() > 100:
            self.tabla_eventos.removeRow(100)

    def filtrar_auditoria(self):
        buscar = self.txt_buscar.text().strip()
        idx_tipo = self.cmb_tipo_evento.currentIndex()

        q = "SELECT id, fecha, tipo, usuario, observaciones, monto, caja_id FROM movimientos_caja WHERE 1=1"
        p = []

        if idx_tipo == 1:
            q += " AND tipo='ALERTA_SEGURIDAD'"
        elif idx_tipo == 2:
            q += " AND tipo='INTERVENCION'"
        elif idx_tipo == 3:
            q += " AND tipo IN ('APERTURA', 'CIERRE_Z', 'CIERRE_AUTO')"
        elif idx_tipo == 4:
            q += " AND tipo='CANCELACION'"
        elif idx_tipo == 5:
            q += " AND tipo IN ('INGRESO', 'RETIRO')"
        else:
            # Excluir tickets de ventas si vemos "Todos los Eventos" para limpiar el ruido
            q += " AND observaciones NOT LIKE '[TICKET]%' AND tipo NOT LIKE '[TICKET]%' AND tipo != 'VENTA'"
            
        if self.caja_filter > 0:
            q += " AND caja_id=?"
            p.append(self.caja_filter)

        if buscar:
            q += " AND (usuario LIKE ? OR observaciones LIKE ? OR tipo LIKE ? OR CAST(id AS CHAR) LIKE ?)"
            p += [f"%{buscar}%"] * 4

        q_count = "SELECT COUNT(*) " + q[q.find("FROM movimientos_caja"):]
        try:
            self.total_logs_count = db_manager.execute_scalar(q_count, tuple(p)) or 0
        except Exception as e:
            print(f"Error contando registros: {e}")
            self.total_logs_count = 0

        q += " ORDER BY id DESC"
        
        self.active_query = q
        self.active_params = p
        self.offset = 0
        self.tabla_eventos.setRowCount(0)
        self._cargar_siguiente_pagina()
        
    def set_caja_filter(self, caja_id):
        if caja_id == "todas":
            self.caja_filter = 0
        else:
            try:
                import re
                num_match = re.search(r'\d+', str(caja_id))
                self.caja_filter = int(num_match.group()) if num_match else 0
            except:
                self.caja_filter = 0
        # Automáticamente refrescar la tabla y el historial de cierres si es necesario
        self.filtrar_auditoria()
        # Opcional: También filtrar Cierres Z por caja si se desea (por ahora solo auditoria)

    def _cargar_siguiente_pagina(self):
        current_rows = self.tabla_eventos.rowCount()
        if current_rows >= self.total_logs_count:
            return
            
        q_paginated = f"{self.active_query} LIMIT 50 OFFSET {self.offset}"
        try:
            page_logs = db_manager.execute_query(q_paginated, tuple(self.active_params)) or []
        except Exception as e:
            print(f"Error cargando página de auditoría: {e}")
            return
            
        if not page_logs:
            return

        self.tabla_eventos.blockSignals(True)
        inicio = current_rows
        fin = inicio + len(page_logs)
        self.tabla_eventos.setRowCount(fin)
        
        for i_idx, r in enumerate(page_logs):
            i = inicio + i_idx
            tipo = str(r['tipo']).upper()
            obs = str(r['observaciones'] or '')
            usuario = str(r['usuario'] or '').upper()
            try:
                c_id = int(r['caja_id']) if r['caja_id'] is not None else 1
            except:
                c_id = 1
            pc_name = f"0{c_id}" if c_id < 10 else str(c_id)

            icon = "📝"
            badge = tipo
            fg_color = "#1e293b"
            bg_color = "white"

            if tipo == 'ALERTA_SEGURIDAD':
                icon = "🚨"
                badge = "SEGURIDAD"
                fg_color = "#dc2626"
                bg_color = "#fef2f2"
            elif tipo == 'INTERVENCION':
                icon = "🔑"
                badge = "INTERVENCIÓN"
                fg_color = "#d97706"
                bg_color = "#fffbeb"
            elif tipo == 'CANCELACION':
                icon = "❌"
                badge = "CANCELACIÓN"
                fg_color = "#7c3aed"
                bg_color = "#f5f3ff"
            elif tipo == 'APERTURA':
                icon = "👤"
                badge = "APERTURA CAJA"
                fg_color = "#2563eb"
                bg_color = "#eff6ff"
            elif tipo == 'CIERRE_Z' or tipo == 'CIERRE_AUTO':
                icon = "🏁"
                badge = "CIERRE DE CAJA"
                fg_color = "#059669"
                bg_color = "#f0fdf4"
            elif tipo == 'RETIRO':
                icon = "💸"
                badge = "RETIRO EFECTIVO"
                fg_color = "#ea580c"
                bg_color = "#fff7ed"
            elif tipo == 'INGRESO':
                icon = "💸"
                badge = "INGRESO CAPITAL"
                fg_color = "#0d9488"
                bg_color = "#f0fdfa"

            item_pc = QTableWidgetItem(f"PC-{pc_name}")
            item_pc.setTextAlignment(Qt.AlignCenter)
            item_pc.setFont(QFont("Segoe UI", 9, QFont.Bold))
            self.tabla_eventos.setItem(i, 0, item_pc)

            fecha_str = str(r['fecha'])
            try:
                if len(fecha_str) >= 16:
                    fecha_str = fecha_str[5:16]
            except:
                pass
            item_f = QTableWidgetItem(fecha_str)
            item_f.setTextAlignment(Qt.AlignCenter)
            self.tabla_eventos.setItem(i, 1, item_f)

            it_ev = QTableWidgetItem(f"{icon} {badge}")
            it_ev.setTextAlignment(Qt.AlignCenter)
            it_ev.setForeground(QColor(fg_color))
            it_ev.setBackground(QColor(bg_color))
            it_ev.setFont(QFont("Segoe UI", 9, QFont.Bold))
            self.tabla_eventos.setItem(i, 2, it_ev)

            it_u = QTableWidgetItem(usuario)
            it_u.setFont(QFont("Segoe UI", 9, QFont.Bold))
            it_u.setTextAlignment(Qt.AlignCenter)
            self.tabla_eventos.setItem(i, 3, it_u)

            self.tabla_eventos.setItem(i, 4, QTableWidgetItem(obs))

            if tipo not in ('ALERTA_SEGURIDAD', 'INTERVENCION', 'CANCELACION') and i % 2 == 1:
                for col in (0, 1, 3, 4):
                    self.tabla_eventos.item(i, col).setBackground(QColor("#f8fafc"))

        self.offset += 50
        self.tabla_eventos.blockSignals(False)

    def _al_hacer_scroll(self, value):
        bar = self.tabla_eventos.verticalScrollBar()
        if bar.maximum() > 0 and value >= bar.maximum() - 15:
            self._cargar_siguiente_pagina()

    def _exportar_auditoria_excel(self):
        try:
            all_matching_logs = db_manager.execute_query(self.active_query, tuple(self.active_params)) or []
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo consultar el log completo para exportar: {e}")
            return

        nombre_def = f"auditoria_seguridad_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Exportar Bitácora de Seguridad", nombre_def,
            "Excel (*.xlsx);;Todos los archivos (*)")
        if not filepath:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Bitacora Seguridad"

            HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
            HEADER_FONT = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
            BORDER_THIN = Border(
                left=Side(style='thin', color="E2E8F0"), right=Side(style='thin', color="E2E8F0"),
                top=Side(style='thin', color="E2E8F0"), bottom=Side(style='thin', color="E2E8F0")
            )

            headers = ["ID", "Fecha / Hora", "Tipo Evento", "Usuario", "Observaciones", "Monto"]
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BORDER_THIN

            for row_idx, r in enumerate(all_matching_logs, 2):
                tipo = str(r['tipo']).upper()
                monto = float(r['monto'] or 0.0)
                
                if tipo == 'ALERTA_SEGURIDAD': tipo = "🚨 SECURITY BREACH"
                elif tipo == 'INTERVENCION': tipo = "🔑 SUPERVISOR INTERVENTION"
                elif tipo == 'CANCELACION': tipo = "❌ TICKET VOID / CANCEL"
                elif tipo == 'APERTURA': tipo = "👤 TURN OPEN"
                elif tipo == 'CIERRE_Z': tipo = "🏁 FISCAL Z-CLOSE"
                elif tipo == 'CIERRE_AUTO': tipo = "🏁 AUTO Z-CLOSE"
                elif tipo == 'RETIRO': tipo = "💸 CASH WITHDRAWAL"
                elif tipo == 'INGRESO': tipo = "💸 CASH INJECTION"

                valores = [
                    r['id'],
                    str(r['fecha']),
                    tipo,
                    str(r['usuario']).upper(),
                    str(r['observaciones'] or ''),
                    monto
                ]

                for col_idx, val in enumerate(valores, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = BORDER_THIN
                    cell.alignment = Alignment(vertical="center")
                    cell.font = Font(name="Segoe UI", size=10)

                    if col_idx == 3 and "SECURITY" in tipo:
                        cell.fill = PatternFill("solid", fgColor="FEE2E2")
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
                    elif col_idx == 3 and "SUPERVISOR" in tipo:
                        cell.fill = PatternFill("solid", fgColor="FEF3C7")
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="92400E")
                    elif col_idx == 3 and "CANCEL" in tipo:
                        cell.fill = PatternFill("solid", fgColor="F5F3FF")
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="5B21B6")

                    if col_idx == 6:
                        cell.number_format = '"$"#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")

            anchos = [8, 22, 28, 14, 45, 15]
            for i, ancho in enumerate(anchos, 1):
                ws.column_dimensions[get_column_letter(i)].width = ancho

            ws.freeze_panes = "A2"
            wb.save(filepath)
            QMessageBox.information(
                self, "Exportación Exitosa", f"Se exportaron {len(all_matching_logs)} registros de auditoría a:\n{filepath}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo exportar a Excel: {e}")
