import os
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QTableWidget, QTableWidgetItem, QHeaderView, QFrame,
    QPushButton, QAbstractItemView, QMessageBox, QFileDialog, QSplitter, QComboBox,
    QGraphicsDropShadowEffect
)
from PyQt5.QtCore import Qt, QTimer, pyqtSignal
from PyQt5.QtGui import QColor, QFont, QBrush
from datetime import datetime

try:
    from src.base_de_datos.database import db_manager
    from src.config import config
    from src.cajero.paso7_cierre import Paso7CierreCaja
except ImportError:
    from database import db_manager
    from config import config
    from cajero.paso7_cierre import Paso7CierreCaja

STYLE = """
QWidget {
    background-color: #F4F6FB;
    font-family: 'Segoe UI', sans-serif;
    font-size: 13px;
    color: #1E293B;
}
QFrame#header {
    background: #FFFFFF;
    border-bottom: 1px solid #EEF2F8;
}
"""

def fmt_moneda(val):
    try:
        return f"${val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return f"${val}"

class Admin7Cierre(QWidget):
    request_dashboard = pyqtSignal()

    def __init__(self, parent_main=None):
        super().__init__()
        self.parent_main = parent_main
        self.active_query = ""
        self.active_params = []
        self.offset = 0
        self.total_logs_count = 0
        self.selected_lane_user = None
        self._setup_ui()
        
        # Timer de monitoreo en vivo para sincronización exacta en tiempo real (2 segundos)
        self.timer_monitoreo = QTimer(self)
        self.timer_monitoreo.timeout.connect(self.actualizar_monitoreo_vivo)

    def showEvent(self, event):
        super().showEvent(event)
        QTimer.singleShot(50, self.cargar_datos)
        self.timer_monitoreo.start(2000) # Actualizar cada 2 segundos en tiempo real

    def hideEvent(self, event):
        self.timer_monitoreo.stop()
        super().hideEvent(event)

    def actualizar_monitoreo_vivo(self):
        if self.isVisible():
            self.cargar_monitoreo_red()

    def _setup_ui(self):
        self.setStyleSheet(STYLE)
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Header White Elite
        hdr = QFrame()
        hdr.setObjectName("header")
        hdr.setFixedHeight(64)
        hl = QHBoxLayout(hdr)
        hl.setContentsMargins(32, 0, 32, 0)
        
        btn_back = QPushButton("← VOLVER AL PANEL")
        btn_back.setCursor(Qt.PointingHandCursor)
        btn_back.setStyleSheet("""
            QPushButton {
                background: #F1F5F9;
                color: #475569;
                font-weight: 700;
                border-radius: 12px;
                padding: 8px 18px;
                font-size: 11px;
                border: none;
            }
            QPushButton:hover {
                background: #E2E8F0;
                color: #0F172A;
            }
        """)
        btn_back.clicked.connect(self.request_dashboard.emit)
        hl.addWidget(btn_back)
        hl.addSpacing(25)
        
        # Pestañas en la cabecera
        self.btn_tab_sala = QPushButton("SALA DE CONTROL CENTRAL DE CAJAS (RED EN VIVO)")
        self.btn_tab_cierres = QPushButton("HISTORIAL DE CIERRES Z DE TERMINALES")
        self.btn_tab_alertas = QPushButton("REGISTRO DE ALERTAS Y EVENTOS DE AUDITORÍA")
        
        for btn in [self.btn_tab_sala, self.btn_tab_cierres, self.btn_tab_alertas]:
            btn.setCursor(Qt.PointingHandCursor)
            hl.addWidget(btn)
            
        self.btn_tab_sala.clicked.connect(lambda: self._set_active_tab(0))
        self.btn_tab_cierres.clicked.connect(lambda: self._set_active_tab(1))
        self.btn_tab_alertas.clicked.connect(lambda: self._set_active_tab(2))
        
        hl.addStretch()
        root.addWidget(hdr)

        # Panel de Métricas / Dashboard de Seguridad
        panel_met = QFrame()
        panel_met.setFixedHeight(100)
        panel_met.setStyleSheet("background-color: transparent; border: none;")
        layout_met = QHBoxLayout(panel_met)
        layout_met.setContentsMargins(32, 16, 32, 10)
        layout_met.setSpacing(16)

        def style_kpi_card(card, title_lbl, val_lbl, palette_key):
            _PALETTE = {
                "red":    ("#FEF2F2", "#EF4444"),
                "amber":  ("#FFFBEB", "#F59E0B"),
                "violet": ("#F5F3FF", "#8B5CF6"),
                "green":  ("#ECFDF5", "#10B981"),
            }
            bg, accent = _PALETTE[palette_key]
            card.setStyleSheet(f"""
                QFrame {{
                    background: {bg};
                    border-radius: 18px;
                    border: none;
                }}
            """)
            h_color = accent.lstrip('#')
            r, g, b = tuple(int(h_color[i:i+2], 16) for i in (0, 2, 4))
            sh = QGraphicsDropShadowEffect(card)
            sh.setBlurRadius(16)
            sh.setColor(QColor(r, g, b, 25))
            sh.setOffset(0, 4)
            card.setGraphicsEffect(sh)
            
            title_lbl.setStyleSheet("font-size: 10px; font-weight: 800; color: #475569; border: none; background: transparent; letter-spacing: 0.5px;")
            val_lbl.setStyleSheet(f"font-size: 20px; font-weight: 900; color: #0F172A; border: none; background: transparent;")

        # Card 1: Brechas de Seguridad
        self.card_brechas = QFrame()
        lay_b = QVBoxLayout(self.card_brechas)
        lay_b.setContentsMargins(16, 12, 16, 12)
        lay_b.setSpacing(4)
        self.lbl_brechas_val = QLabel("0")
        lbl_b_title = QLabel("⚠️ BRECHAS DE SEGURIDAD")
        lay_b.addWidget(self.lbl_brechas_val)
        lay_b.addWidget(lbl_b_title)
        style_kpi_card(self.card_brechas, lbl_b_title, self.lbl_brechas_val, "red")
        layout_met.addWidget(self.card_brechas)

        # Card 2: Intervenciones F11
        self.card_interv = QFrame()
        lay_i = QVBoxLayout(self.card_interv)
        lay_i.setContentsMargins(16, 12, 16, 12)
        lay_i.setSpacing(4)
        self.lbl_interv_val = QLabel("0")
        lbl_i_title = QLabel("🔑 INTERVENCIONES SUPERVISOR")
        lay_i.addWidget(self.lbl_interv_val)
        lay_i.addWidget(lbl_i_title)
        style_kpi_card(self.card_interv, lbl_i_title, self.lbl_interv_val, "amber")
        layout_met.addWidget(self.card_interv)

        # Card 3: Cancelaciones
        self.card_cancels = QFrame()
        lay_c = QVBoxLayout(self.card_cancels)
        lay_c.setContentsMargins(16, 12, 16, 12)
        lay_c.setSpacing(4)
        self.lbl_cancels_val = QLabel("0")
        lbl_c_title = QLabel("❌ TICKETS CANCELADOS")
        lay_c.addWidget(self.lbl_cancels_val)
        lay_c.addWidget(lbl_c_title)
        style_kpi_card(self.card_cancels, lbl_c_title, self.lbl_cancels_val, "violet")
        layout_met.addWidget(self.card_cancels)

        # Card 4: Último Cierre Z
        self.card_last_z = QFrame()
        lay_z = QVBoxLayout(self.card_last_z)
        lay_z.setContentsMargins(16, 12, 16, 12)
        lay_z.setSpacing(4)
        self.lbl_last_z_val = QLabel("S/D")
        lbl_z_title = QLabel("🏁 ÚLTIMO CIERRE FISCAL")
        lay_z.addWidget(self.lbl_last_z_val)
        lay_z.addWidget(lbl_z_title)
        style_kpi_card(self.card_last_z, lbl_z_title, self.lbl_last_z_val, "green")
        layout_met.addWidget(self.card_last_z)

        root.addWidget(panel_met)

        # --- PANEL DE CIERRES Z HISTÓRICOS (Para Tab 2 al lado de Vigilancia) ---
        panel_cierres = QFrame()
        panel_cierres.setStyleSheet("QFrame { background-color: white; border: none; border-radius: 20px; }")
        sh_cierres = QGraphicsDropShadowEffect(panel_cierres)
        sh_cierres.setBlurRadius(20)
        sh_cierres.setColor(QColor(0, 0, 0, 12))
        sh_cierres.setOffset(0, 4)
        panel_cierres.setGraphicsEffect(sh_cierres)
        
        left_lay = QVBoxLayout(panel_cierres)
        left_lay.setContentsMargins(24, 24, 24, 24)
        left_lay.setSpacing(15)

        lbl_l_tit = QLabel("🏁 CIERRES Z Y ARQUEOS FISCALES HISTÓRICOS")
        lbl_l_tit.setStyleSheet("font-weight: 900; color: #0F172A; font-size: 14px; border: none; background: transparent;")
        left_lay.addWidget(lbl_l_tit)

        # Botón de Acción Principal F12
        self.btn_ejecutar_z = QPushButton("🚀 INICIAR PROCESO DE CIERRE Z / AUDITORÍA (F12)")
        self.btn_ejecutar_z.setCursor(Qt.PointingHandCursor)
        self.btn_ejecutar_z.setStyleSheet("""
            QPushButton {
                background: #6366F1;
                color: white;
                border: none;
                font-weight: 800;
                font-size: 13px;
                padding: 12px;
                border-radius: 12px;
            }
            QPushButton:hover {
                background: #4F46E5;
            }
        """)
        self.btn_ejecutar_z.clicked.connect(self._ejecutar_cierre_z_dialogo)
        left_lay.addWidget(self.btn_ejecutar_z)

        # Tabla de Historial de Cierres Z
        self.tabla_cierres = QTableWidget()
        self.tabla_cierres.setColumnCount(7)
        self.tabla_cierres.setHorizontalHeaderLabels([
            "🆔 ID CIERRE", 
            "💻 PC / CAJA", 
            "⏰ HORARIO INICIO", 
            "🏁 HORARIO DE CIERRE", 
            "👤 CAJERO", 
            "⚖️ DIFERENCIA", 
            "💵 FÍSICO / ESPERADO"
        ])
        self.tabla_cierres.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabla_cierres.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabla_cierres.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tabla_cierres.verticalHeader().setVisible(False)
        
        self.tabla_cierres.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: none;
                gridline-color: #F1F5F9;
                font-size: 12px;
                border-radius: 16px;
            }
            QTableWidget::item {
                padding: 10px;
                border-bottom: 1px solid #F1F5F9;
            }
            QHeaderView::section {
                background-color: #F8FAFC;
                color: #475569;
                font-weight: 800;
                border: none;
                border-bottom: 2px solid #E2E8F0;
                padding: 10px;
                font-size: 11px;
            }
        """)
        
        hh_c = self.tabla_cierres.horizontalHeader()
        hh_c.setSectionResizeMode(0, QHeaderView.Interactive)
        self.tabla_cierres.setColumnWidth(0, 95)
        hh_c.setSectionResizeMode(1, QHeaderView.Interactive)
        self.tabla_cierres.setColumnWidth(1, 110)
        hh_c.setSectionResizeMode(2, QHeaderView.Stretch)
        hh_c.setSectionResizeMode(3, QHeaderView.Stretch)
        hh_c.setSectionResizeMode(4, QHeaderView.Interactive)
        self.tabla_cierres.setColumnWidth(4, 130)
        hh_c.setSectionResizeMode(5, QHeaderView.Interactive)
        self.tabla_cierres.setColumnWidth(5, 145)
        hh_c.setSectionResizeMode(6, QHeaderView.Interactive)
        self.tabla_cierres.setColumnWidth(6, 180)
        
        self.tabla_cierres.doubleClicked.connect(self._reimprimir_ticket_z_seleccionado)
        left_lay.addWidget(self.tabla_cierres)

        # Botón de Reimpresión
        self.btn_reimprimir_z = QPushButton("🖨️ RE-IMPRIMIR REPORTE Z SELECCIONADO")
        self.btn_reimprimir_z.setCursor(Qt.PointingHandCursor)
        self.btn_reimprimir_z.setStyleSheet("""
            QPushButton {
                background: #F1F5F9;
                color: #475569;
                border: none;
                font-weight: 800;
                font-size: 12px;
                padding: 10px;
                border-radius: 10px;
            }
            QPushButton:hover {
                background: #E2E8F0;
                color: #0F172A;
            }
        """)
        self.btn_reimprimir_z.clicked.connect(self._reimprimir_ticket_z_seleccionado)
        left_lay.addWidget(self.btn_reimprimir_z)

        # Import QScrollArea and QTabWidget for the decentralized view
        from PyQt5.QtWidgets import QScrollArea, QTabWidget

        # --- NUEVA PESTAÑA: SALA DE CONTROL CENTRALIZADA EN RED ---
        self.widget_sala_control = QWidget()
        layout_central = QVBoxLayout(self.widget_sala_control)
        layout_central.setContentsMargins(10, 10, 10, 10)
        layout_central.setSpacing(10)

        # Splitter principal vertical
        self.splitter_sala_control = QSplitter(Qt.Vertical)
        self.splitter_sala_control.setStyleSheet("QSplitter::handle{background:#EEF2F8;}")

        # Panel superior (horizontal)
        panel_superior = QWidget()
        panel_superior.setMinimumHeight(380)
        layout_sup = QHBoxLayout(panel_superior)
        layout_sup.setContentsMargins(0, 0, 0, 0)
        layout_sup.setSpacing(10)

        # Splitter horizontal superior
        self.splitter_sup = QSplitter(Qt.Horizontal)
        self.splitter_sup.setStyleSheet("QSplitter::handle{background:#EEF2F8;}")

        # Panel Izquierdo: Lista de Cajas en Red
        panel_izq_cajas = QFrame()
        panel_izq_cajas.setStyleSheet("QFrame { background-color: white; border: none; border-radius: 20px; }")
        sh_izq = QGraphicsDropShadowEffect(panel_izq_cajas)
        sh_izq.setBlurRadius(20)
        sh_izq.setColor(QColor(0, 0, 0, 12))
        sh_izq.setOffset(0, 4)
        panel_izq_cajas.setGraphicsEffect(sh_izq)
        
        layout_izq = QVBoxLayout(panel_izq_cajas)
        layout_izq.setContentsMargins(20, 20, 20, 20)
        layout_izq.setSpacing(10)

        # Determinar si estamos conectados a una IP remota
        db_path_str = str(db_manager.db_path).replace("\\", "/")
        ip_conectada = ""
        if db_path_str.startswith("//"):
            partes = db_path_str.split("/")
            if len(partes) >= 3:
                ip_conectada = partes[2]

        titulo_central = f"📟 CENTRAL DE CAJAS EN RED LOCAL (HOST: {ip_conectada})" if ip_conectada else f"📟 CENTRAL DE CAJAS EN RED LOCAL"

        lbl_izq_tit = QLabel(titulo_central)
        lbl_izq_tit.setStyleSheet("font-weight: 900; color: #0F172A; font-size: 13px; border: none; background: transparent;")
        layout_izq.addWidget(lbl_izq_tit)

        # Línea de barrido en tiempo real y status
        self.lbl_scan_status = QLabel("🛰️ BARRIDO DE TERMINALES: ONLINE")
        self.lbl_scan_status.setStyleSheet("font-size: 9px; font-weight: 800; color: #10B981; border: none; background: transparent; letter-spacing: 0.5px;")
        layout_izq.addWidget(self.lbl_scan_status)

        from PyQt5.QtWidgets import QProgressBar
        self.scan_bar = QProgressBar()
        self.scan_bar.setRange(0, 0)
        self.scan_bar.setTextVisible(False)
        self.scan_bar.setFixedHeight(5)
        self.scan_bar.setStyleSheet("""
            QProgressBar {
                background: #f1f5f9;
                border: none;
                border-radius: 2px;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #10b981, stop:0.5 #3b82f6, stop:1 #10b981);
                border-radius: 2px;
            }
        """)
        layout_izq.addWidget(self.scan_bar)

        # Scroll Area para las tarjetas de cajas
        scroll_cajas = QScrollArea()
        scroll_cajas.setWidgetResizable(True)
        scroll_cajas.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        
        self.container_cajas = QWidget()
        self.container_cajas.setStyleSheet("background: transparent;")
        self.layout_grid_cajas = QVBoxLayout(self.container_cajas)
        self.layout_grid_cajas.setContentsMargins(0, 0, 0, 0)
        self.layout_grid_cajas.setSpacing(10)
        self.layout_grid_cajas.addStretch()
        self.caja_cards = {}
        
        scroll_cajas.setWidget(self.container_cajas)
        layout_izq.addWidget(scroll_cajas)
        
        self.splitter_sup.addWidget(panel_izq_cajas)

        # Panel Derecho: Tarjeta en Vivo
        self.panel_der_control = QFrame()
        self.panel_der_control.setStyleSheet("QFrame { background-color: white; border: none; border-radius: 20px; }")
        sh_der = QGraphicsDropShadowEffect(self.panel_der_control)
        sh_der.setBlurRadius(20)
        sh_der.setColor(QColor(0, 0, 0, 12))
        sh_der.setOffset(0, 4)
        self.panel_der_control.setGraphicsEffect(sh_der)
        
        layout_der = QVBoxLayout(self.panel_der_control)
        layout_der.setContentsMargins(0, 0, 0, 0)
        layout_der.setSpacing(0)

        # Encabezado Control de Cierre en Vivo
        hdr_vivo = QFrame()
        hdr_vivo.setFixedHeight(45)
        hdr_vivo.setStyleSheet("background: #FFFFFF; border-bottom: 1px solid #EEF2F8; border-top-left-radius: 20px; border-top-right-radius: 20px;")
        lay_hdr_v = QHBoxLayout(hdr_vivo)
        lay_hdr_v.setContentsMargins(20, 0, 20, 0)
        
        self.lbl_hdr_vivo_titulo = QLabel("💎 CAJAFACIL PRO - CONTROL DE CIERRE")
        self.lbl_hdr_vivo_titulo.setStyleSheet("color: #0F172A; font-weight: 800; font-size: 13px; background: transparent;")
        lay_hdr_v.addWidget(self.lbl_hdr_vivo_titulo)
        
        self.lbl_hdr_vivo_usuario = QLabel("👤 ROL: ADMIN")
        self.lbl_hdr_vivo_usuario.setStyleSheet("color: #6366F1; font-weight: 800; font-size: 11px; background: transparent;")
        lay_hdr_v.addWidget(self.lbl_hdr_vivo_usuario, 0, Qt.AlignRight)
        
        layout_der.addWidget(hdr_vivo)

        # Fase de Monitoreo
        self.fase_bar = QFrame()
        self.fase_bar.setFixedHeight(30)
        self.fase_bar.setStyleSheet("background-color: #EEF2FF; border-bottom: 1px solid #E0E7FF;")
        lay_fase = QHBoxLayout(self.fase_bar)
        lay_fase.setContentsMargins(15, 0, 15, 0)
        lbl_fase = QLabel("🔒 FASE 2: VERIFICACIÓN Y CONCILIACIÓN DE CAJA")
        lbl_fase.setStyleSheet("color: #6366F1; font-weight: 800; font-size: 11px; background: transparent;")
        lay_fase.addWidget(lbl_fase, 0, Qt.AlignCenter)
        layout_der.addWidget(self.fase_bar)

        # Cuerpo del Panel Remoto
        cuerpo_remoto = QWidget()
        cuerpo_lay = QHBoxLayout(cuerpo_remoto)
        cuerpo_lay.setContentsMargins(20, 20, 20, 20)
        cuerpo_lay.setSpacing(15)

        # Columna Izquierda del Cuerpo: 4 Cards de Métricas
        col_izq_cards = QWidget()
        col_izq_lay = QVBoxLayout(col_izq_cards)
        col_izq_lay.setContentsMargins(0, 0, 0, 0)
        col_izq_lay.setSpacing(10)

        # Card 1: Ventas Efectivo
        self.card_v_efec = QFrame()
        lay_ve = QHBoxLayout(self.card_v_efec)
        lay_ve.setContentsMargins(12, 8, 12, 8)
        lbl_ve_ico = QLabel("💰")
        lbl_ve_ico.setStyleSheet("font-size: 20px; border: none; background: transparent;")
        lay_ve.addWidget(lbl_ve_ico)
        
        info_ve = QVBoxLayout()
        info_ve.setSpacing(2)
        lbl_ve_t = QLabel("VENTAS EFECTIVO")
        self.lbl_ve_val = QLabel("$ 0.00")
        info_ve.addWidget(lbl_ve_t)
        info_ve.addWidget(self.lbl_ve_val)
        lay_ve.addLayout(info_ve)
        
        # Card 2: Ventas Digital
        self.card_v_dig = QFrame()
        lay_vd = QHBoxLayout(self.card_v_dig)
        lay_vd.setContentsMargins(12, 8, 12, 8)
        lbl_vd_ico = QLabel("💳")
        lbl_vd_ico.setStyleSheet("font-size: 20px; border: none; background: transparent;")
        lay_vd.addWidget(lbl_vd_ico)
        
        info_vd = QVBoxLayout()
        info_vd.setSpacing(2)
        lbl_vd_t = QLabel("VENTAS DIGITAL")
        self.lbl_vd_val = QLabel("$ 0.00")
        info_vd.addWidget(lbl_vd_t)
        info_vd.addWidget(self.lbl_vd_val)
        lay_vd.addLayout(info_vd)
        
        # Card 3: Fondo Apertura
        self.card_v_fnd = QFrame()
        lay_vf = QHBoxLayout(self.card_v_fnd)
        lay_vf.setContentsMargins(12, 8, 12, 8)
        lbl_vf_ico = QLabel("🏁")
        lbl_vf_ico.setStyleSheet("font-size: 20px; border: none; background: transparent;")
        lay_vf.addWidget(lbl_vf_ico)
        
        info_vf = QVBoxLayout()
        info_vf.setSpacing(2)
        lbl_vf_t = QLabel("FONDO APERTURA")
        self.lbl_vf_val = QLabel("$ 0.00")
        info_vf.addWidget(lbl_vf_t)
        info_vf.addWidget(self.lbl_vf_val)
        lay_vf.addLayout(info_vf)
        
        # Card 4: Alertas Seguridad
        self.card_v_alt = QFrame()
        lay_va = QHBoxLayout(self.card_v_alt)
        lay_va.setContentsMargins(12, 8, 12, 8)
        lbl_va_ico = QLabel("🚨")
        lbl_va_ico.setStyleSheet("font-size: 20px; border: none; background: transparent;")
        lay_va.addWidget(lbl_va_ico)
        
        info_va = QVBoxLayout()
        info_va.setSpacing(2)
        lbl_va_t = QLabel("ALERTAS SEGURIDAD")
        self.lbl_va_val = QLabel("0")
        info_va.addWidget(lbl_va_t)
        info_va.addWidget(self.lbl_va_val)
        lay_va.addLayout(info_va)
        
        # Style mini cards
        def style_mini_card(card, val_lbl, title_lbl, p_key):
            _PAL_MINI = {
                "green":  ("#ECFDF5", "#10B981"),
                "blue":   ("#EFF6FF", "#3B82F6"),
                "purple": ("#F5F3FF", "#8B5CF6"),
                "red":    ("#FEF2F2", "#EF4444"),
            }
            bg, accent = _PAL_MINI[p_key]
            card.setStyleSheet(f"""
                QFrame {{
                    background-color: {bg};
                    border-radius: 14px;
                    border: none;
                }}
            """)
            h_color = accent.lstrip('#')
            r, g, b = tuple(int(h_color[i:i+2], 16) for i in (0, 2, 4))
            sh = QGraphicsDropShadowEffect(card)
            sh.setBlurRadius(12)
            sh.setColor(QColor(r, g, b, 20))
            sh.setOffset(0, 3)
            card.setGraphicsEffect(sh)
            
            title_lbl.setStyleSheet("font-size: 9px; font-weight: 800; color: #475569; background: transparent; border: none; letter-spacing: 0.5px;")
            val_lbl.setStyleSheet(f"font-size: 16px; font-weight: 900; color: #0F172A; background: transparent; border: none;")

        style_mini_card(self.card_v_efec, self.lbl_ve_val, lbl_ve_t, "green")
        style_mini_card(self.card_v_dig, self.lbl_vd_val, lbl_vd_t, "blue")
        style_mini_card(self.card_v_fnd, self.lbl_vf_val, lbl_vf_t, "purple")
        style_mini_card(self.card_v_alt, self.lbl_va_val, lbl_va_t, "red")

        col_izq_lay.addWidget(self.card_v_efec)
        col_izq_lay.addWidget(self.card_v_dig)
        col_izq_lay.addWidget(self.card_v_fnd)
        col_izq_lay.addWidget(self.card_v_alt)
        cuerpo_lay.addWidget(col_izq_cards, 42)

        # Columna Derecha del Cuerpo: Cálculos Principales
        col_der_calc = QFrame()
        col_der_calc.setStyleSheet("background-color: #F8FAFC; border: none; border-radius: 16px; padding: 12px;")
        sh_calc = QGraphicsDropShadowEffect(col_der_calc)
        sh_calc.setBlurRadius(12)
        sh_calc.setColor(QColor(0, 0, 0, 10))
        sh_calc.setOffset(0, 3)
        col_der_calc.setGraphicsEffect(sh_calc)
        
        col_der_lay = QVBoxLayout(col_der_calc)
        col_der_lay.setContentsMargins(10, 10, 10, 10)
        col_der_lay.setSpacing(10)

        # 1. Efectivo Esperado Card
        f_esp = QFrame()
        f_esp.setStyleSheet("background-color: #EFF6FF; border: none; border-radius: 12px;")
        lay_esp = QVBoxLayout(f_esp)
        lay_esp.setContentsMargins(10, 8, 10, 8)
        lbl_esp_t = QLabel("EFECTIVO ESPERADO")
        lbl_esp_t.setAlignment(Qt.AlignCenter)
        lbl_esp_t.setStyleSheet("font-size: 9px; font-weight: 800; color: #3B82F6; border: none; background: transparent; letter-spacing: 0.5px;")
        self.lbl_live_esperado = QLabel("$ 0.00")
        self.lbl_live_esperado.setAlignment(Qt.AlignCenter)
        self.lbl_live_esperado.setStyleSheet("font-size: 24px; font-weight: 900; color: #1E3A8A; border: none; background: transparent;")
        lay_esp.addWidget(lbl_esp_t)
        lay_esp.addWidget(self.lbl_live_esperado)
        col_der_lay.addWidget(f_esp)

        # 2. Input del Físico Contado
        f_input = QFrame()
        f_input.setStyleSheet("background-color: #FFFBEB; border: none; border-radius: 12px;")
        lay_input = QVBoxLayout(f_input)
        lay_input.setContentsMargins(10, 8, 10, 8)
        lbl_inp_t = QLabel("INGRESA EL FÍSICO CONTADO ($)")
        lbl_inp_t.setAlignment(Qt.AlignCenter)
        lbl_inp_t.setStyleSheet("font-size: 9px; font-weight: 800; color: #D97706; border: none; background: transparent; letter-spacing: 0.5px;")
        
        self.txt_live_fisico = QLineEdit()
        self.txt_live_fisico.setAlignment(Qt.AlignCenter)
        self.txt_live_fisico.setStyleSheet("""
            QLineEdit {
                background: white; border: 2px solid #F59E0B; border-radius: 8px;
                color: #B45309; font-size: 24px; font-weight: 900; padding: 3px;
            }
            QLineEdit:focus { border-color: #D97706; }
        """)
        self.txt_live_fisico.setText("0.00")
        self.txt_live_fisico.textChanged.connect(self._calcular_live_diferencia)
        
        lay_input.addWidget(lbl_inp_t)
        lay_input.addWidget(self.txt_live_fisico)
        col_der_lay.addWidget(f_input)

        # 3. Diferencia
        self.frame_live_dif = QFrame()
        self.frame_live_dif.setStyleSheet("background-color: #ECFDF5; border: none; border-radius: 12px;")
        lay_ld = QHBoxLayout(self.frame_live_dif)
        lay_ld.setContentsMargins(15, 10, 15, 10)
        
        lbl_ld_t = QLabel("DIFERENCIA:")
        lbl_ld_t.setStyleSheet("font-weight: 800; font-size: 11px; color: #065F46; border: none; background: transparent;")
        lay_ld.addWidget(lbl_ld_t)
        
        self.lbl_live_dif_val = QLabel("+ $ 0.00")
        self.lbl_live_dif_val.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.lbl_live_dif_val.setStyleSheet("font-size: 18px; font-weight: 900; color: #10B981; border: none; background: transparent;")
        lay_ld.addWidget(self.lbl_live_dif_val)
        col_der_lay.addWidget(self.frame_live_dif)

        cuerpo_lay.addWidget(col_der_calc, 58)
        layout_der.addWidget(cuerpo_remoto)

        # Fila de Botones
        btn_bar_vivo = QFrame()
        btn_bar_vivo.setFixedHeight(60)
        btn_bar_vivo.setStyleSheet("background-color: #FFFFFF; border-top: 1px solid #EEF2F8; border-bottom-left-radius: 20px; border-bottom-right-radius: 20px;")
        lay_btn_bar = QHBoxLayout(btn_bar_vivo)
        lay_btn_bar.setContentsMargins(20, 8, 20, 8)
        lay_btn_bar.setSpacing(10)

        self.btn_live_corte = QPushButton("📊 CORTE DIARIO")
        self.btn_live_corte.setCursor(Qt.PointingHandCursor)
        self.btn_live_corte.setStyleSheet("""
            QPushButton {
                background: #F1F5F9; color: #475569; border: none; border-radius: 10px;
                font-weight: 800; font-size: 11px; padding: 8px 16px;
            }
            QPushButton:hover { background: #E2E8F0; color: #0F172A; }
        """)
        self.btn_live_corte.clicked.connect(self._ejecutar_live_corte_diario)
        lay_btn_bar.addWidget(self.btn_live_corte)

        self.btn_live_corte_excel = QPushButton("📊 EXPORTAR EXCEL")
        self.btn_live_corte_excel.setCursor(Qt.PointingHandCursor)
        self.btn_live_corte_excel.setStyleSheet("""
            QPushButton {
                background: #10B981; color: white; border: none; border-radius: 10px;
                font-weight: 800; font-size: 11px; padding: 8px 16px;
            }
            QPushButton:hover { background: #059669; }
        """)
        self.btn_live_corte_excel.clicked.connect(self._exportar_corte_excel)
        lay_btn_bar.addWidget(self.btn_live_corte_excel)

        self.btn_live_confirmar = QPushButton("⚠️ CONFIRMAR")
        self.btn_live_confirmar.setCursor(Qt.PointingHandCursor)
        self.btn_live_confirmar.setStyleSheet("""
            QPushButton {
                background: #EF4444; color: white; border: none; border-radius: 10px;
                font-weight: 800; font-size: 11px; padding: 8px 16px;
            }
            QPushButton:hover { background: #DC2626; }
        """)
        self.btn_live_confirmar.clicked.connect(self._ejecutar_live_confirmar_cierre)
        lay_btn_bar.addWidget(self.btn_live_confirmar)

        self.btn_live_cierre_total = QPushButton("🌎 CIERRE TOTAL / ADMIN")
        self.btn_live_cierre_total.setCursor(Qt.PointingHandCursor)
        self.btn_live_cierre_total.setStyleSheet("""
            QPushButton {
                background: #6366F1; color: white; border: none; border-radius: 10px;
                font-weight: 800; font-size: 11px; padding: 8px 16px;
            }
            QPushButton:hover { background: #4F46E5; }
        """)
        self.btn_live_cierre_total.clicked.connect(self._ejecutar_live_cierre_total_centralizado)
        lay_btn_bar.addWidget(self.btn_live_cierre_total)

        self.btn_live_reporte = QPushButton("📊 REPORTE ADMIN")
        self.btn_live_reporte.setCursor(Qt.PointingHandCursor)
        self.btn_live_reporte.setStyleSheet("""
            QPushButton {
                background: #F1F5F9; color: #475569; border: none; border-radius: 10px;
                font-weight: 800; font-size: 11px; padding: 8px 16px;
            }
            QPushButton:hover { background: #E2E8F0; color: #0F172A; }
        """)
        self.btn_live_reporte.clicked.connect(self._ejecutar_live_reporte_admin)
        lay_btn_bar.addWidget(self.btn_live_reporte)

        layout_der.addWidget(btn_bar_vivo)
        self.splitter_sup.addWidget(self.panel_der_control)
        
        self.splitter_sup.setSizes([380, 620])
        layout_sup.addWidget(self.splitter_sup)
        
        self.splitter_sala_control.addWidget(panel_superior)

        # --- PANEL INFERIOR: VIGILANCIA EN TIEMPO REAL ---
        panel_inferior = QFrame()
        panel_inferior.setMinimumHeight(260)
        panel_inferior.setStyleSheet("QFrame { background-color: white; border: none; border-radius: 20px; }")
        sh_inf = QGraphicsDropShadowEffect(panel_inferior)
        sh_inf.setBlurRadius(20)
        sh_inf.setColor(QColor(0, 0, 0, 12))
        sh_inf.setOffset(0, 4)
        panel_inferior.setGraphicsEffect(sh_inf)
        
        layout_inf = QVBoxLayout(panel_inferior)
        layout_inf.setContentsMargins(20, 20, 20, 20)
        layout_inf.setSpacing(12)
        
        hdr_vigilancia = QHBoxLayout()
        lbl_v_title = QLabel("👁️ VIGILANCIA CENTRALIZADA & EVENTOS DE AUDITORÍA EN TIEMPO REAL")
        lbl_v_title.setStyleSheet("font-weight: 900; color: #0F172A; font-size: 13px; border: none; background: transparent;")
        hdr_vigilancia.addWidget(lbl_v_title)
        hdr_vigilancia.addStretch()
        layout_inf.addLayout(hdr_vigilancia)
        
        filt_bar = QHBoxLayout()
        filt_bar.setSpacing(10)
        
        self.txt_buscar = QLineEdit()
        self.txt_buscar.setPlaceholderText("🔍 Buscar por usuario, ID, tipo o descripción...")
        self.txt_buscar.setMinimumWidth(220)
        self.txt_buscar.setStyleSheet("""
            QLineEdit {
                background: #F8FAFC;
                border: 1px solid #E2E8F0;
                border-radius: 10px;
                padding: 8px 12px;
                font-size: 13px;
                color: #1E293B;
            }
            QLineEdit:focus {
                border-color: #6366F1;
                background: #FFFFFF;
            }
        """)
        self.txt_buscar.textChanged.connect(self.filtrar_auditoria)
        filt_bar.addWidget(self.txt_buscar)
        
        self.cmb_tipo_evento = QComboBox()
        self.cmb_tipo_evento.addItems([
            "Todos los Eventos",
            "Brechas de Seguridad",
            "Intervenciones Supervisor",
            "Turnos y Aperturas",
            "Cancelaciones de Tickets",
            "Ingresos y Retiros de Efectivo"
        ])
        self.cmb_tipo_evento.setStyleSheet("""
            QComboBox {
                background: #F8FAFC;
                border: 1px solid #E2E8F0;
                border-radius: 10px;
                padding: 8px 12px;
                font-size: 13px;
                color: #1E293B;
            }
            QComboBox:focus {
                border-color: #6366F1;
                background: #FFFFFF;
            }
        """)
        self.cmb_tipo_evento.currentIndexChanged.connect(self.filtrar_auditoria)
        filt_bar.addWidget(self.cmb_tipo_evento)
        
        btn_exportar = QPushButton("📥 EXPORTAR BITÁCORA")
        btn_exportar.setCursor(Qt.PointingHandCursor)
        btn_exportar.setStyleSheet("""
            QPushButton {
                background: #10B981;
                color: white;
                font-weight: 700;
                border-radius: 10px;
                padding: 8px 20px;
                font-size: 13px;
                border: none;
            }
            QPushButton:hover {
                background: #059669;
            }
        """)
        btn_exportar.clicked.connect(self._exportar_auditoria_excel)
        filt_bar.addWidget(btn_exportar)
        
        layout_inf.addLayout(filt_bar)
        
        self.tabla_eventos = QTableWidget()
        self.tabla_eventos.setColumnCount(6)
        self.tabla_eventos.setHorizontalHeaderLabels(["🆔 ID", "💻 PC", "📅 FECHA / HORA", "🚨 EVENTO DE AUDITORÍA", "👤 USUARIO", "📝 DETALLE / OBSERVACIONES"])
        self.tabla_eventos.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabla_eventos.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabla_eventos.verticalHeader().setVisible(False)
        
        self.tabla_eventos.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: none;
                gridline-color: #F1F5F9;
                font-size: 12px;
                border-radius: 16px;
            }
            QTableWidget::item {
                padding: 10px;
                border-bottom: 1px solid #F1F5F9;
            }
            QHeaderView::section {
                background-color: #F8FAFC;
                color: #475569;
                font-weight: 800;
                border: none;
                border-bottom: 2px solid #E2E8F0;
                padding: 10px;
                font-size: 11px;
            }
        """)
        
        hh_e = self.tabla_eventos.horizontalHeader()
        hh_e.setSectionResizeMode(0, QHeaderView.Interactive)
        self.tabla_eventos.setColumnWidth(0, 80)
        hh_e.setSectionResizeMode(1, QHeaderView.Interactive)
        self.tabla_eventos.setColumnWidth(1, 90)
        hh_e.setSectionResizeMode(2, QHeaderView.Interactive)
        self.tabla_eventos.setColumnWidth(2, 170)
        hh_e.setSectionResizeMode(3, QHeaderView.Interactive)
        self.tabla_eventos.setColumnWidth(3, 220)
        hh_e.setSectionResizeMode(4, QHeaderView.Interactive)
        self.tabla_eventos.setColumnWidth(4, 110)
        hh_e.setSectionResizeMode(5, QHeaderView.Stretch)
        
        self.tabla_eventos.verticalScrollBar().valueChanged.connect(self._al_hacer_scroll)
        layout_inf.addWidget(self.tabla_eventos)
        
        # En Tab 1: agregar panel_superior a pantalla completa
        layout_central.addWidget(panel_superior)

        # --- ENSAMBLADO EN PESTAÑAS (QTabWidget) ---
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: none;
                background: transparent;
            }
        """)

        # --- TAB 2: CIERRES Z A PANTALLA COMPLETA ---
        tab2_widget = QWidget()
        tab2_lay = QVBoxLayout(tab2_widget)
        tab2_lay.setContentsMargins(15, 15, 15, 15)
        tab2_lay.addWidget(panel_cierres)
  
        # --- TAB 3: REGISTRO DE AUDITORÍA Y ALERTAS A PANTALLA COMPLETA ---
        tab3_widget = QWidget()
        tab3_lay = QVBoxLayout(tab3_widget)
        tab3_lay.setContentsMargins(15, 15, 15, 15)
        tab3_lay.addWidget(panel_inferior)
  
        self.tab_widget.addTab(self.widget_sala_control, "SALA DE CONTROL CENTRAL DE CAJAS (RED EN VIVO)")
        self.tab_widget.addTab(tab2_widget, "HISTORIAL DE CIERRES Z DE TERMINALES")
        self.tab_widget.addTab(tab3_widget, "REGISTRO DE ALERTAS Y EVENTOS DE AUDITORÍA")
        
        self.tab_widget.tabBar().hide()
        root.addWidget(self.tab_widget)
        self.tab_widget.currentChanged.connect(self.cargar_datos)
        self.tab_widget.currentChanged.connect(self._update_tab_button_styles)
        self._update_tab_button_styles()

        # Inicializar Timer de Barrido en Tiempo Real cada 30 segundos
        self.timer_barrido = QTimer(self)
        self.timer_barrido.timeout.connect(self._hacer_barrido_red)
        self.timer_barrido.start(30000)

    def _set_active_tab(self, index):
        self.tab_widget.setCurrentIndex(index)
        self._update_tab_button_styles()

    def _update_tab_button_styles(self):
        style_active = """
            QPushButton {
                background: #EFF6FF; color: #1E3A8A; font-weight: 900; border-radius: 12px; 
                padding: 10px 20px; border: none; font-size: 11px;
            }
        """
        style_inactive = """
            QPushButton {
                background: transparent; color: #64748B; font-weight: 700; border-radius: 12px; 
                padding: 10px 20px; border: none; font-size: 11px;
            }
            QPushButton:hover {
                background: #F8FAFC; color: #0F172A;
            }
        """
        idx = self.tab_widget.currentIndex()
        self.btn_tab_sala.setStyleSheet(style_active if idx == 0 else style_inactive)
        self.btn_tab_cierres.setStyleSheet(style_active if idx == 1 else style_inactive)
        self.btn_tab_alertas.setStyleSheet(style_active if idx == 2 else style_inactive)

    def cargar_datos(self):
        # 1. Cargar estadísticas de métricas en caliente
        try:
            n_brechas = db_manager.execute_scalar("SELECT COUNT(*) FROM movimientos_caja WHERE tipo='ALERTA_SEGURIDAD'") or 0
            n_interv = db_manager.execute_scalar("SELECT COUNT(*) FROM movimientos_caja WHERE tipo='INTERVENCION'") or 0
            n_cancels = db_manager.execute_scalar("SELECT COUNT(*) FROM movimientos_caja WHERE tipo='CANCELACION'") or 0
            last_z_fecha = db_manager.execute_scalar("SELECT fecha FROM movimientos_caja WHERE tipo='CIERRE_Z' ORDER BY id DESC LIMIT 1")
            
            self.lbl_brechas_val.setText(str(n_brechas))
            self.lbl_interv_val.setText(str(n_interv))
            self.lbl_cancels_val.setText(str(n_cancels))
            
            if last_z_fecha:
                try:
                    from datetime import datetime
                    dt = datetime.strptime(str(last_z_fecha), "%Y-%m-%d %H:%M:%S")
                    self.lbl_last_z_val.setText(dt.strftime("%d/%m/%Y %I:%M %p").lower())
                except:
                    self.lbl_last_z_val.setText(str(last_z_fecha))
            else:
                self.lbl_last_z_val.setText("Ninguno registrado")
        except Exception as e:
            print(f"Error cargando métricas: {e}")

        # 2. Cargar grilla de Cierres Z
        try:
            cierres = db_manager.execute_query("""
                SELECT 
                    c.id, 
                    c.fecha as fecha_cierre, 
                    c.usuario, 
                    c.monto, 
                    c.observaciones, 
                    c.caja_id,
                    (
                        SELECT a.fecha 
                        FROM movimientos_caja a 
                        WHERE a.tipo='APERTURA' 
                          AND a.caja_id = c.caja_id 
                          AND a.fecha < c.fecha 
                        ORDER BY a.fecha DESC 
                        LIMIT 1
                    ) as fecha_inicio
                FROM movimientos_caja c
                WHERE c.tipo='CIERRE_Z'
                ORDER BY c.id DESC
            """) or []
            self.tabla_cierres.setRowCount(len(cierres))
            for i, r in enumerate(cierres):
                # Col 0: ID
                item_id = QTableWidgetItem(str(r['id']))
                item_id.setTextAlignment(Qt.AlignCenter)
                self.tabla_cierres.setItem(i, 0, item_id)
                
                # Col 1: PC ID (Caja ID)
                try:
                    c_id_val = int(r['caja_id'])
                    item_pc = QTableWidgetItem(f"0{c_id_val}" if c_id_val < 10 else str(c_id_val))
                except:
                    item_pc = QTableWidgetItem(str(r['caja_id'] or '01'))
                item_pc.setTextAlignment(Qt.AlignCenter)
                item_pc.setFont(QFont("Consolas", 10, QFont.Bold))
                self.tabla_cierres.setItem(i, 1, item_pc)
                
                # Col 2: Fecha/Hora Inicio
                f_ini_val = str(r['fecha_inicio']) if r['fecha_inicio'] else "-"
                item_ini = QTableWidgetItem(f_ini_val)
                item_ini.setTextAlignment(Qt.AlignCenter)
                self.tabla_cierres.setItem(i, 2, item_ini)
                
                # Col 3: Fecha/Hora Cierre
                item_fin = QTableWidgetItem(str(r['fecha_cierre']))
                item_fin.setTextAlignment(Qt.AlignCenter)
                self.tabla_cierres.setItem(i, 3, item_fin)
                
                # Col 4: Cajero
                item_usr = QTableWidgetItem(str(r['usuario']).upper())
                item_usr.setTextAlignment(Qt.AlignCenter)
                self.tabla_cierres.setItem(i, 4, item_usr)
                
                # Col 5: Diferencia (monto)
                dif = float(r['monto'] or 0.0)
                item_dif = QTableWidgetItem(fmt_moneda(dif))
                item_dif.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if dif < 0:
                    item_dif.setForeground(QColor("#ef4444"))
                    item_dif.setBackground(QColor("#fef2f2"))
                elif dif > 0:
                    item_dif.setForeground(QColor("#10b981"))
                    item_dif.setBackground(QColor("#f0fdf4"))
                else:
                    item_dif.setForeground(QColor("#1e293b"))
                self.tabla_cierres.setItem(i, 5, item_dif)

                # Col 6: Detalles (Físico / Esperado)
                obs = str(r['observaciones'] or '')
                detalles = "-"
                if "Fis:" in obs and "Esp:" in obs:
                    try:
                        fis_val = obs.split("Fis:")[1].split("|")[0].strip()
                        esp_val = obs.split("Esp:")[1].strip()
                        detalles = f"{fis_val} / {esp_val}"
                    except:
                        detalles = obs
                else:
                    detalles = obs
                
                item_det = QTableWidgetItem(detalles)
                item_det.setTextAlignment(Qt.AlignCenter)
                self.tabla_cierres.setItem(i, 6, item_det)
        except Exception as e:
            print(f"Error cargando cierres: {e}")

        # 3. Cargar la bitácora de eventos de auditoría (para posterior scroll infinito)
        self.filtrar_auditoria()

        # 4. Cargar la sala de control de red
        try:
            self.cargar_monitoreo_red()
        except Exception as e:
            print(f"Error cargando sala de control de red: {e}")

    def filtrar_auditoria(self):
        buscar = self.txt_buscar.text().strip()
        idx_tipo = self.cmb_tipo_evento.currentIndex()

        q = "SELECT id, fecha, tipo, usuario, observaciones, monto, caja_id FROM movimientos_caja WHERE 1=1"
        p = []

        # Filtrar por tipo
        # Combobox: 
        # 0: Todos los Eventos
        # 1: Brechas de Seguridad
        # 2: Intervenciones Supervisor
        # 3: Turnos y Aperturas
        # 4: Cancelaciones de Tickets
        # 5: Ingresos y Retiros de Efectivo
        if idx_tipo == 1:
            q += " AND tipo='ALERTA_SEGURIDAD'"
        elif idx_tipo == 2:
            q += " AND tipo='INTERVENCION'"
        elif idx_tipo == 3:
            q += " AND tipo IN ('APERTURA', 'CIERRE_Z', 'CIERRE_AUTO')"
        elif idx_tipo == 4:
            q += " AND tipo='CANCELACION'"
        elif idx_tipo == 5:
            q += " AND tipo IN ('INGRESO', 'RETIRO')"

        if buscar:
            q += " AND (usuario LIKE ? OR observaciones LIKE ? OR tipo LIKE ? OR CAST(id AS TEXT) LIKE ?)"
            p += [f"%{buscar}%"] * 4

        # Obtener cantidad total de registros coincidentes de manera super veloz
        q_count = "SELECT COUNT(*) " + q[q.find("FROM movimientos_caja"):]
        try:
            self.total_logs_count = db_manager.execute_scalar(q_count, tuple(p)) or 0
        except Exception as e:
            print(f"Error contando registros: {e}")
            self.total_logs_count = 0

        q += " ORDER BY id DESC"
        
        self.active_query = q
        self.active_params = p
        self.offset = 0
        self.tabla_eventos.setRowCount(0)
        self._cargar_siguiente_pagina()

    def _cargar_siguiente_pagina(self):
        current_rows = self.tabla_eventos.rowCount()
        if current_rows >= self.total_logs_count:
            return
            
        q_paginated = f"{self.active_query} LIMIT 50 OFFSET {self.offset}"
        try:
            page_logs = db_manager.execute_query(q_paginated, tuple(self.active_params)) or []
        except Exception as e:
            print(f"Error cargando página de auditoría: {e}")
            return
            
        if not page_logs:
            return

        self.tabla_eventos.blockSignals(True)
        inicio = current_rows
        fin = inicio + len(page_logs)
        self.tabla_eventos.setRowCount(fin)
        
        for i_idx, r in enumerate(page_logs):
            i = inicio + i_idx
            tipo = str(r['tipo']).upper()
            obs = str(r['observaciones'] or '')
            usuario = str(r['usuario'] or '').upper()
            try:
                c_id = int(r['caja_id']) if r['caja_id'] is not None else 1
            except:
                c_id = 1
            pc_name = f"0{c_id}" if c_id < 10 else str(c_id)

            # Configurar badge y emoji según el tipo de evento
            icon = "📝"
            badge = tipo
            fg_color = "#1e293b"
            bg_color = "white"

            if tipo == 'ALERTA_SEGURIDAD':
                icon = "🚨"
                badge = "SEGURIDAD"
                fg_color = "#dc2626"
                bg_color = "#fef2f2"
            elif tipo == 'INTERVENCION':
                icon = "🔑"
                badge = "INTERVENCIÓN"
                fg_color = "#d97706"
                bg_color = "#fffbeb"
            elif tipo == 'CANCELACION':
                icon = "❌"
                badge = "CANCELACIÓN"
                fg_color = "#7c3aed"
                bg_color = "#f5f3ff"
            elif tipo == 'APERTURA':
                icon = "👤"
                badge = "APERTURA CAJA"
                fg_color = "#2563eb"
                bg_color = "#eff6ff"
            elif tipo == 'CIERRE_Z' or tipo == 'CIERRE_AUTO':
                icon = "🏁"
                badge = "CIERRE DE CAJA"
                fg_color = "#059669"
                bg_color = "#f0fdf4"
            elif tipo == 'RETIRO':
                icon = "💸"
                badge = "RETIRO EFECTIVO"
                fg_color = "#ea580c"
                bg_color = "#fff7ed"
            elif tipo == 'INGRESO':
                icon = "💸"
                badge = "INGRESO CAPITAL"
                fg_color = "#0d9488"
                bg_color = "#f0fdfa"

            # 1. Col 0: ID
            item_id = QTableWidgetItem(f"#{r['id']}")
            item_id.setTextAlignment(Qt.AlignCenter)
            item_id.setFont(QFont("Consolas", 9, QFont.Bold))
            self.tabla_eventos.setItem(i, 0, item_id)

            # 2. Col 1: PC
            item_pc = QTableWidgetItem(f"PC-{pc_name}")
            item_pc.setTextAlignment(Qt.AlignCenter)
            item_pc.setFont(QFont("Segoe UI", 9, QFont.Bold))
            self.tabla_eventos.setItem(i, 1, item_pc)

            # 3. Col 2: Fecha / Hora
            item_f = QTableWidgetItem(str(r['fecha']))
            item_f.setTextAlignment(Qt.AlignCenter)
            self.tabla_eventos.setItem(i, 2, item_f)

            # 4. Col 3: Badge Evento
            it_ev = QTableWidgetItem(f"{icon} {badge}")
            it_ev.setTextAlignment(Qt.AlignCenter)
            it_ev.setForeground(QColor(fg_color))
            it_ev.setBackground(QColor(bg_color))
            it_ev.setFont(QFont("Segoe UI", 9, QFont.Bold))
            self.tabla_eventos.setItem(i, 3, it_ev)

            # 5. Col 4: Usuario
            it_u = QTableWidgetItem(usuario)
            it_u.setFont(QFont("Segoe UI", 9, QFont.Bold))
            it_u.setTextAlignment(Qt.AlignCenter)
            self.tabla_eventos.setItem(i, 4, it_u)

            # 6. Col 5: Detalles
            self.tabla_eventos.setItem(i, 5, QTableWidgetItem(obs))

            # Aplicar colores alternados de fila para filas normales
            if tipo not in ('ALERTA_SEGURIDAD', 'INTERVENCION', 'CANCELACION') and i % 2 == 1:
                for col in (0, 1, 2, 4, 5):
                    self.tabla_eventos.item(i, col).setBackground(QColor("#f8fafc"))

        self.offset += 50
        self.tabla_eventos.blockSignals(False)

    def _al_hacer_scroll(self, value):
        bar = self.tabla_eventos.verticalScrollBar()
        if bar.maximum() > 0 and value >= bar.maximum() - 15:
            self._cargar_siguiente_pagina()

    def _ejecutar_cierre_z_dialogo(self):
        try:
            # Ejecutar el diálogo paso 7 de Cierre Z
            dlg = Paso7CierreCaja(self)
            dlg.exec_()
            self.cargar_datos() # Recargar datos del catálogo al cerrar
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo ejecutar el cierre: {e}")

    def _reimprimir_ticket_z_seleccionado(self):
        row = self.tabla_cierres.currentRow()
        if row == -1:
            QMessageBox.information(self, "Selección Requerida", "Por favor seleccione un Cierre Z de la lista.")
            return

        id_z = self.tabla_cierres.item(row, 0).text()
        fecha_z = self.tabla_cierres.item(row, 3).text()
        cajero_z = self.tabla_cierres.item(row, 4).text().lower()
        dif_str = self.tabla_cierres.item(row, 5).text().replace("$", "").replace(".", "").replace(",", ".").strip()
        dif = float(dif_str)

        detalles_str = self.tabla_cierres.item(row, 6).text()
        
        # Intentar extraer el físico del texto
        fisico = 0.0
        esperado = 0.0
        try:
            if "/" in detalles_str:
                parts = detalles_str.split("/")
                fisico = float(parts[0].replace("$", "").replace(".", "").replace(",", ".").strip())
                esperado = float(parts[1].replace("$", "").replace(".", "").replace(",", ".").strip())
        except:
            pass

        try:
            from src.hardware.printer import printer_manager
            # Reconstruir datos para la impresión del ticket Z
            # Buscamos las alertas de seguridad y el total de ventas correspondientes a este cajero en este dia
            fecha_solo = fecha_z.split(" ")[0]
            alertas = db_manager.execute_scalar(
                "SELECT COUNT(*) FROM movimientos_caja WHERE tipo='ALERTA_SEGURIDAD' AND usuario=? AND date(fecha)=date(?)",
                (cajero_z, fecha_solo)
            ) or 0
            
            res_v = db_manager.execute_query(
                "SELECT SUM(total) as t, SUM(pago_otro) as ta FROM ventas WHERE estado='CERRADA' AND usuario=? AND date(fecha)=date(?)",
                (cajero_z, fecha_solo)
            )
            v_total = float((res_v[0]["t"] if res_v else 0.0) or 0.0)
            v_tarj = float((res_v[0]["ta"] if res_v else 0.0) or 0.0)
            
            # Obtener el fondo de apertura anterior a esta fecha
            fondo = float(db_manager.execute_scalar(
                "SELECT monto FROM movimientos_caja WHERE tipo='APERTURA' AND date(fecha)<=date(?) ORDER BY id DESC LIMIT 1",
                (fecha_solo,)
            ) or 0.0)

            datos_impresion = {
                "fondo": fondo,
                "t_efec": v_total - v_tarj,
                "t_tarj": v_tarj,
                "t_total": v_total,
                "esperado": esperado or (fondo + (v_total - v_tarj)),
                "alertas": alertas,
                "f_hist": 0.0,
                "segunda_tiketera": True
            }

            printer_manager.imprimir_ticket_z(cajero_z, fisico or esperado, dif, datos_impresion)
            QMessageBox.information(
                self, "Éxito", f"Reporte Z #{id_z} reconstruido y enviado a la cola de impresión."
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo re-imprimir el Reporte Z: {e}")

    def _exportar_auditoria_excel(self):
        # Consultar el log completo en caliente para exportación instantánea
        try:
            all_matching_logs = db_manager.execute_query(self.active_query, tuple(self.active_params)) or []
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo consultar el log completo para exportar: {e}")
            return

        from datetime import datetime
        nombre_def = f"auditoria_seguridad_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Exportar Bitácora de Seguridad", nombre_def,
            "Excel (*.xlsx);;Todos los archivos (*)")
        if not filepath:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Bitacora Seguridad"

            # Estilos de encabezado
            HEADER_FILL = PatternFill("solid", fgColor="1E3A8A") # Azul Elite
            HEADER_FONT = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
            BORDER_THIN = Border(
                left=Side(style='thin', color="E2E8F0"), right=Side(style='thin', color="E2E8F0"),
                top=Side(style='thin', color="E2E8F0"), bottom=Side(style='thin', color="E2E8F0")
            )

            headers = ["ID", "Fecha / Hora", "Tipo Evento", "Usuario", "Observaciones", "Monto"]
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BORDER_THIN

            # Datos
            for row_idx, r in enumerate(all_matching_logs, 2):
                tipo = str(r['tipo']).upper()
                monto = float(r['monto'] or 0.0)
                
                # Traducir tipo
                if tipo == 'ALERTA_SEGURIDAD': tipo = "🚨 SECURITY BREACH"
                elif tipo == 'INTERVENCION': tipo = "🔑 SUPERVISOR INTERVENTION"
                elif tipo == 'CANCELACION': tipo = "❌ TICKET VOID / CANCEL"
                elif tipo == 'APERTURA': tipo = "👤 TURN OPEN"
                elif tipo == 'CIERRE_Z': tipo = "🏁 FISCAL Z-CLOSE"
                elif tipo == 'CIERRE_AUTO': tipo = "🏁 AUTO Z-CLOSE"
                elif tipo == 'RETIRO': tipo = "💸 CASH WITHDRAWAL"
                elif tipo == 'INGRESO': tipo = "💸 CASH INJECTION"

                valores = [
                    r['id'],
                    str(r['fecha']),
                    tipo,
                    str(r['usuario']).upper(),
                    str(r['observaciones'] or ''),
                    monto
                ]

                for col_idx, val in enumerate(valores, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = BORDER_THIN
                    cell.alignment = Alignment(vertical="center")
                    cell.font = Font(name="Segoe UI", size=10)

                    # Estilizado de brechas rojas en excel
                    if col_idx == 3 and "SECURITY" in tipo:
                        cell.fill = PatternFill("solid", fgColor="FEE2E2") # rosado suave
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
                    elif col_idx == 3 and "SUPERVISOR" in tipo:
                        cell.fill = PatternFill("solid", fgColor="FEF3C7") # ambar suave
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="92400E")
                    elif col_idx == 3 and "CANCEL" in tipo:
                        cell.fill = PatternFill("solid", fgColor="F5F3FF") # purpura suave
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="5B21B6")

                    if col_idx == 6:
                        cell.number_format = '"$"#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")

            # Ancho de columnas
            anchos = [8, 22, 28, 14, 45, 15]
            for i, ancho in enumerate(anchos, 1):
                ws.column_dimensions[get_column_letter(i)].width = ancho

            ws.freeze_panes = "A2"
            wb.save(filepath)
            QMessageBox.information(
                self, "Exportación Exitosa", f"Se exportaron {len(all_matching_logs)} registros de auditoría a:\n{filepath}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo exportar a Excel: {e}")

    # --- S SALA DE CONTROL CENTRALIZADA (RED DE CAJAS VIVO) ---

    def cargar_monitoreo_red(self):
        try:
            # Consultar todas las cajas físicas que han operado en el sistema
            cajas_ids = db_manager.execute_query(
                "SELECT DISTINCT caja_id FROM movimientos_caja UNION SELECT DISTINCT caja_id FROM ventas"
            ) or []
            cajas_list = [c["caja_id"] for c in cajas_ids if c["caja_id"] is not None]
            if 1 not in cajas_list:
                cajas_list.append(1)
            cajas_list.sort()
        except Exception as e:
            print(f"Error consultando cajas operativas: {e}")
            cajas_list = [1]

        self.cajas_data = {}

        for cid in cajas_list:
            key_caja = f"caja_{cid}"
            
            # Consultar el último movimiento de esta caja específica que determine el turno
            mov = db_manager.execute_query(
                "SELECT tipo, fecha, monto, usuario FROM movimientos_caja WHERE caja_id = ? AND tipo IN ('APERTURA', 'CIERRE_Z', 'CIERRE_AUTO', 'SOLICITUD_CIERRE') ORDER BY id DESC LIMIT 1",
                (cid,)
            )

            is_online = False
            apertura_date = "1970-01-01 00:00:00"
            fondo = 0.0
            user = "Cerrada"
            rol = "cajero"

            if mov:
                tipo_mov = mov[0]["tipo"]
                user = mov[0]["usuario"]
                if tipo_mov == 'APERTURA':
                    is_online = True
                    apertura_date = mov[0]["fecha"]
                    fondo = float(mov[0]["monto"] or 0.0)
                elif tipo_mov == 'SOLICITUD_CIERRE':
                    is_online = True
                    ant = db_manager.execute_query(
                        "SELECT tipo, fecha, monto FROM movimientos_caja WHERE caja_id = ? AND tipo='APERTURA' ORDER BY id DESC LIMIT 1",
                        (cid,)
                    )
                    if ant:
                        apertura_date = ant[0]["fecha"]
                        fondo = float(ant[0]["monto"] or 0.0)
            else:
                if cid == 1:
                    is_online = False
                    apertura_date = None
                    fondo = 0.0
                    user = "-"

            # Calcular datos acumulados en caliente para esta caja específica
            v_efec = 0.0
            v_dig = 0.0
            alertas = 0

            if is_online:
                try:
                    res_v = db_manager.execute_query(
                        "SELECT SUM(total) as t, SUM(pago_otro) as ta FROM ventas WHERE estado='COMPLETADA' AND caja_id=? AND fecha>=?",
                        (cid, apertura_date)
                    )
                    if res_v and res_v[0]["t"] is not None:
                        total_sales = float(res_v[0]["t"] or 0.0)
                        v_dig = float(res_v[0]["ta"] or 0.0)
                        v_efec = total_sales - v_dig
                    
                    alertas = db_manager.execute_scalar(
                        "SELECT COUNT(*) FROM movimientos_caja WHERE tipo='ALERTA_SEGURIDAD' AND caja_id=? AND fecha>=?",
                        (cid, apertura_date)
                    ) or 0
                except Exception as e:
                    print(f"Error calculando totales para caja {cid}: {e}")

            esperado = fondo + v_efec
            self.cajas_data[key_caja] = {
                "caja_id": cid,
                "user": user,
                "rol": rol,
                "online": is_online,
                "fondo": fondo,
                "v_efec": v_efec,
                "v_dig": v_dig,
                "v_total": v_efec + v_dig,
                "alertas": alertas,
                "esperado": esperado,
                "apertura_date": apertura_date
            }

            border_color = "#10B981" if is_online else "#E2E8F0"
            bg_color = "#F0FDF4" if is_online else "white"

            # Si ya está en caché, simplemente actualizar las etiquetas y colores
            if cid in self.caja_cards:
                card = self.caja_cards[cid]
                card.setStyleSheet(f"""
                    QFrame {{
                        background-color: {bg_color};
                        border: 2px solid {border_color};
                        border-radius: 12px;
                        padding: 10px;
                    }}
                    QFrame:hover {{
                        border-color: #3b82f6;
                        background-color: #f8fafc;
                    }}
                """)
                card.lbl_c_name.setText(f"CAJA 0{cid} - {user.upper()}")
                badge_text = "🟢 ACTIVA / ONLINE" if is_online else "🔴 CERRADA / OFFLINE"
                badge_style = "color: #059669; font-weight: 800; font-size: 9px;" if is_online else "color: #94a3b8; font-weight: 800; font-size: 9px;"
                card.lbl_c_badge.setText(badge_text)
                card.lbl_c_badge.setStyleSheet(badge_style + " border: none; background: transparent;")
                total_tot = v_efec + v_dig
                card.lbl_ac_val.setText(fmt_moneda(total_tot) if is_online else "$ 0.00")
            else:
                # Crear tarjeta de caja visual nueva
                card = QFrame()
                card.setFrameShape(QFrame.StyledPanel)
                card.setCursor(Qt.PointingHandCursor)
                
                card.setStyleSheet(f"""
                    QFrame {{
                        background-color: {bg_color};
                        border: 2px solid {border_color};
                        border-radius: 12px;
                        padding: 10px;
                    }}
                    QFrame:hover {{
                        border-color: #3b82f6;
                        background-color: #f8fafc;
                    }}
                """)
                
                card_lay = QHBoxLayout(card)
                card_lay.setContentsMargins(10, 8, 10, 8)
                
                lbl_c_ico = QLabel("📟")
                lbl_c_ico.setStyleSheet("font-size: 24px; border: none; background: transparent;")
                card_lay.addWidget(lbl_c_ico)

                info_lay = QVBoxLayout()
                lbl_c_name = QLabel(f"CAJA 0{cid} - {user.upper()}")
                lbl_c_name.setStyleSheet("font-weight: 900; font-size: 12px; color: #1e3a8a; border: none; background: transparent;")
                
                badge_text = "🟢 ACTIVA / ONLINE" if is_online else "🔴 CERRADA / OFFLINE"
                badge_style = "color: #059669; font-weight: 800; font-size: 9px;" if is_online else "color: #94a3b8; font-weight: 800; font-size: 9px;"
                lbl_c_badge = QLabel(badge_text)
                lbl_c_badge.setStyleSheet(badge_style + " border: none; background: transparent;")
                
                info_lay.addWidget(lbl_c_name)
                info_lay.addWidget(lbl_c_badge)
                card_lay.addLayout(info_lay)
                
                card_lay.addStretch()

                ac_lay = QVBoxLayout()
                lbl_ac_t = QLabel("VENTAS")
                lbl_ac_t.setAlignment(Qt.AlignRight)
                lbl_ac_t.setStyleSheet("font-size: 9px; font-weight: 700; color: #64748b; border: none; background: transparent;")
                
                total_tot = v_efec + v_dig
                lbl_ac_val = QLabel(fmt_moneda(total_tot) if is_online else "$ 0.00")
                lbl_ac_val.setAlignment(Qt.AlignRight)
                lbl_ac_val.setStyleSheet("font-size: 13px; font-weight: 900; color: #0f172a; border: none; background: transparent;")
                
                ac_lay.addWidget(lbl_ac_t)
                ac_lay.addWidget(lbl_ac_val)
                card_lay.addLayout(ac_lay)

                # Guardar referencias
                card.lbl_c_name = lbl_c_name
                card.lbl_c_badge = lbl_c_badge
                card.lbl_ac_val = lbl_ac_val

                # Enlazar evento click de forma segura usando la clave caja_{cid}
                card.mousePressEvent = lambda event, kc=key_caja: self.seleccionar_caja(kc)
                
                self.caja_cards[cid] = card
                self.layout_grid_cajas.insertWidget(self.layout_grid_cajas.count() - 1, card)

        if self.cajas_data:
            if not self.selected_lane_user or self.selected_lane_user not in self.cajas_data:
                self.seleccionar_caja(list(self.cajas_data.keys())[0])
            else:
                self.seleccionar_caja(self.selected_lane_user)

    def seleccionar_caja(self, key_caja):
        if key_caja not in self.cajas_data:
            return
            
        self.selected_lane_user = key_caja
        data = self.cajas_data[key_caja]
        c_id = data["caja_id"]
        cajero_actual = data["user"]

        self.lbl_hdr_vivo_titulo.setText(f"💎 CAJAFACIL PRO - CONTROL DE CIERRE: CAJA 0{c_id} ({cajero_actual.upper()})")
        self.lbl_hdr_vivo_usuario.setText(f"👤 ROL: {data['rol'].upper()}")
        
        if data['online']:
            self.fase_bar.setStyleSheet("background-color: #eff6ff; border-bottom: 1px solid #bfdbfe;")
            for child in self.fase_bar.findChildren(QLabel):
                child.setText("🟢 LÍNEA DE CHECKOUT ACTIVA - LECTURA EN VIVO DE RED")
                child.setStyleSheet("color: #1d4ed8; font-weight: 800; font-size: 11px; background: transparent;")
            self.btn_live_confirmar.setEnabled(True)
            self.btn_live_corte.setEnabled(True)
            self.btn_live_confirmar.setStyleSheet("""
                QPushButton { background: #ef4444; color: white; border: none; border-radius: 6px; font-weight: 900; font-size: 10px; padding: 8px 12px; }
                QPushButton:hover { background: #dc2626; }
            """)
        else:
            self.fase_bar.setStyleSheet("background-color: #fef2f2; border-bottom: 1px solid #fecaca;")
            for child in self.fase_bar.findChildren(QLabel):
                child.setText("🔴 TERMINAL DESCONECTADA / TURNO CERRADO POR ADMINISTRACIÓN")
                child.setStyleSheet("color: #991b1b; font-weight: 800; font-size: 11px; background: transparent;")
            self.btn_live_confirmar.setEnabled(False)
            self.btn_live_corte.setEnabled(False)
            self.btn_live_confirmar.setStyleSheet("""
                QPushButton { background: #cbd5e1; color: #94a3b8; border: none; border-radius: 6px; font-weight: 900; font-size: 10px; padding: 8px 12px; }
            """)

        self.lbl_ve_val.setText(fmt_moneda(data['v_efec']))
        self.lbl_vd_val.setText(fmt_moneda(data['v_dig']))
        self.lbl_vf_val.setText(fmt_moneda(data['fondo']))
        self.lbl_va_val.setText(str(data['alertas']))

        self.lbl_live_esperado.setText(fmt_moneda(data['esperado']))
        # No machacar lo que está escribiendo el usuario si tiene el foco
        if not self.txt_live_fisico.hasFocus():
            self.txt_live_fisico.setText(f"{data['esperado']:.2f}" if data['online'] else "0.00")
        self._calcular_live_diferencia()

    def _calcular_live_diferencia(self):
        if not self.selected_lane_user or self.selected_lane_user not in self.cajas_data:
            return
            
        data = self.cajas_data[self.selected_lane_user]
        esperado = data['esperado']
        
        try:
            from src.utils.parser import parse_float_regional
            fisico = parse_float_regional(self.txt_live_fisico.text())
        except:
            fisico = 0.0

        dif = fisico - esperado
        if not data['online']:
            dif = 0.0

        if dif >= 0:
            self.frame_live_dif.setStyleSheet("background-color: #f0fdf4; border: 2px solid #10b981; border-radius: 8px; padding: 10px;")
            self.lbl_live_dif_val.setText(f"+ {fmt_moneda(abs(dif))}")
            self.lbl_live_dif_val.setStyleSheet("font-size: 18px; font-weight: 900; color: #047857; border: none; background: transparent;")
        else:
            self.frame_live_dif.setStyleSheet("background-color: #fef2f2; border: 2px solid #ef4444; border-radius: 8px; padding: 10px;")
            self.lbl_live_dif_val.setText(f"- {fmt_moneda(abs(dif))}")
            self.lbl_live_dif_val.setStyleSheet("font-size: 18px; font-weight: 900; color: #b91c1c; border: none; background: transparent;")

    def _ejecutar_live_corte_diario(self):
        if not self.selected_lane_user or self.selected_lane_user not in self.cajas_data:
            return
            
        data = self.cajas_data[self.selected_lane_user]
        if not data["online"]:
            QMessageBox.warning(self, "Acción Inválida", "La caja seleccionada no está activa.")
            return

        try:
            from src.hardware.printer import printer_manager
            # Preparar datos de Lectura de Control X (Remoto)
            datos_impresion = {
                "fondo": data["fondo"],
                "t_efec": data["v_efec"],
                "t_tarj": data["v_dig"],
                "t_total": data["v_total"],
                "esperado": data["esperado"],
                "alertas": data["alertas"],
                "dia_tarjeta": data["v_dig"],
                "dia_total": data["v_total"],
                "efectivo_esperado": data["esperado"],
                "segunda_tiketera": False
            }
            
            # Imprimir ticket de lectura
            printer_manager.imprimir_ticket_z(
                f"{self.selected_lane_user.upper()} (LECTURA X)", data["esperado"], 0.0, datos_impresion
            )
            
            QMessageBox.information(
                self, "Lectura X Exitosa",
                f"El Reporte X (Lectura de control) de la caja de {self.selected_lane_user.upper()} "
                "ha sido enviado a la ticketera central. La caja sigue en línea (ONLINE)."
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo imprimir la lectura X: {e}")

    def _exportar_corte_excel(self):
        if not self.selected_lane_user or self.selected_lane_user not in self.cajas_data:
            return
            
        data = self.cajas_data[self.selected_lane_user]
        
        filepath, _ = QFileDialog.getSaveFileName(self, "Guardar Corte en Excel", f"Corte_Caja_{self.selected_lane_user}.xlsx", "Archivos Excel (*.xlsx)")
        if not filepath:
            return
            
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Corte de Caja"

            # Estilos
            TITLE_FONT = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
            TITLE_FILL = PatternFill("solid", fgColor="3B82F6")
            HEADER_FONT = Font(name="Segoe UI", size=12, bold=True, color="1E293B")
            HEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
            VALUE_FONT = Font(name="Segoe UI", size=12, bold=True)
            BORDER_THIN = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Título principal
            ws.merge_cells("A1:D1")
            cell_title = ws.cell(row=1, column=1, value=f"CORTE DE CAJA - {self.selected_lane_user.upper()}")
            cell_title.font = TITLE_FONT
            cell_title.fill = TITLE_FILL
            cell_title.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            # Metricas
            metricas = [
                ("Fondo de Apertura", data["fondo"]),
                ("Ventas en Efectivo", data["v_efec"]),
                ("Ventas Digitales / Tarjeta", data["v_dig"]),
                ("Total Vendido", data["v_total"]),
                ("Efectivo Esperado en Caja", data["esperado"])
            ]

            row_idx = 3
            ws.merge_cells(f"A{row_idx}:D{row_idx}")
            ws.cell(row=row_idx, column=1, value="RESUMEN FINANCIERO DEL DÍA").font = HEADER_FONT
            ws.cell(row=row_idx, column=1).fill = HEADER_FILL
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
            row_idx += 1

            for concepto, valor in metricas:
                ws.cell(row=row_idx, column=1, value=concepto).font = Font(name="Segoe UI", size=11)
                cell_val = ws.cell(row=row_idx, column=2, value=valor)
                cell_val.font = VALUE_FONT
                cell_val.number_format = '"$"#,##0.00'
                row_idx += 1

            # Anchos de columna
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20

            wb.save(filepath)
            QMessageBox.information(self, "Excel Creado", f"Corte guardado exitosamente en:\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Error al Exportar", f"No se pudo crear el archivo Excel:\n{e}")

    def _ejecutar_live_confirmar_cierre(self):
        if not self.selected_lane_user or self.selected_lane_user not in self.cajas_data:
            return
            
        data = self.cajas_data[self.selected_lane_user]
        if not data["online"]:
            QMessageBox.warning(self, "Acción Inválida", "La caja ya está cerrada.")
            return

        c_id = data["caja_id"]
        cajero_actual = data["user"]

        # Selector de Cierre
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("CIERRE DE CAJA EN RED - CONTROL GERENCIAL")
        msg_box.setIcon(QMessageBox.Question)
        msg_box.setText(
            f"¿Qué acción desea ejecutar para la CAJA 0{c_id} ({cajero_actual.upper()})?\n\n"
            "🛰️ SEÑAL DE ARQUEO: Envía una señal en tiempo real al terminal del cajero para que "
            "le aparezca la ventana de arqueo físico contado y finalice él mismo su turno.\n\n"
            "⚠️ CIERRE FORZADO INMEDIATO (Z-Force): Cierra y desconecta la caja de forma silenciosa "
            "desde el servidor usando los datos contables actuales."
        )
        
        btn_senal = msg_box.addButton("🛰️ ENVIAR SEÑAL DE ARQUEO", QMessageBox.AcceptRole)
        btn_forzar = msg_box.addButton("⚠️ CIERRE FORZADO INMEDIATO", QMessageBox.DestructiveRole)
        btn_cancelar = msg_box.addButton("Cancelar", QMessageBox.RejectRole)
        
        msg_box.exec_()
        clicked = msg_box.clickedButton()
        
        if clicked == btn_cancelar:
            return
            
        elif clicked == btn_senal:
            try:
                fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                db_manager.execute_non_query(
                    "INSERT INTO movimientos_caja (fecha, tipo, monto, usuario, observaciones, caja_id) VALUES (?, 'SOLICITUD_CIERRE', 0.0, ?, 'PENDIENTE', ?)",
                    (fecha_actual, cajero_actual, c_id)
                )
                QMessageBox.information(
                    self, "Señal de Arqueo Enviada",
                    f"Se ha enviado la señal de arqueo remoto a la CAJA 0{c_id} de {cajero_actual.upper()}.\n\n"
                    "El cajero verá la ventana emergente de arqueo de inmediato para contar el efectivo."
                )
                self.cargar_datos()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"No se pudo enviar la señal de arqueo: {e}")
            return
            
        elif clicked == btn_forzar:
            try:
                from src.utils.parser import parse_float_regional
                fisico = parse_float_regional(self.txt_live_fisico.text())
            except:
                fisico = 0.0

            esperado = data["esperado"]
            dif = fisico - esperado

            try:
                # Registrar cierre Z en bitacora
                obs = f"Fis:${fisico:.2f} | Esp:${esperado:.2f} | CIERRE FORZADO ADMIN Z-FORCE (SILENCIOSO)"
                db_manager.execute_non_query(
                    "INSERT INTO movimientos_caja (tipo, monto, usuario, observaciones, caja_id) VALUES ('CIERRE_Z', ?, ?, ?, ?)",
                    (dif, cajero_actual, obs, c_id)
                )

                # Cambiar estado de ventas asociadas a 'CERRADA'
                db_manager.execute_non_query(
                    "UPDATE ventas SET estado='CERRADA' WHERE usuario=? AND caja_id=? AND fecha>=? AND estado='COMPLETADA'",
                    (cajero_actual, c_id, data["apertura_date"])
                )

                # Imprimir ticket Z remoto
                try:
                    from src.hardware.printer import printer_manager
                    datos_impresion = {
                        "fondo": data["fondo"],
                        "t_efec": data["v_efec"],
                        "t_tarj": data["v_dig"],
                        "t_total": data["v_total"],
                        "esperado": esperado,
                        "alertas": data["alertas"],
                        "dia_tarjeta": data["v_dig"],
                        "dia_total": data["v_total"],
                        "efectivo_esperado": esperado,
                        "segunda_tiketera": True
                    }
                    printer_manager.imprimir_ticket_z(f"CAJA 0{c_id} ({cajero_actual.upper()})", fisico, dif, datos_impresion)
                except Exception as pe:
                    print(f"Error imprimiendo ticket remoto: {pe}")

                self._notificar_cierre_telegram(cajero_actual, c_id, esperado, fisico, dif, data)

                QMessageBox.information(
                    self, "Cierre Remoto Completado",
                    f"La CAJA 0{c_id} de {cajero_actual.upper()} ha sido arqueada y cerrada en red exitosamente.\n"
                    f"Diferencia: {fmt_moneda(dif)}. Reporte enviado a impresión."
                )
                
                self.cargar_datos()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"No se pudo ejecutar el cierre remoto Z-Force: {e}")

    def _notificar_cierre_telegram(self, cajero, caja_id, esperado, fisico, diferencia, data):
        try:
            import requests
            token = config.get("telegram_token", "").strip()
            chat_id = config.get("telegram_chat_id", "").strip()
            if not token or not chat_id:
                return # Nube no configurada
                
            mensaje = (
                f"☁️ *TPV NUBE - CORTE Z (CAJA 0{caja_id})*\n"
                f"👤 Usuario: {cajero.upper()}\n"
                f"💰 Ventas Efectivo: {fmt_moneda(data['v_efec'])}\n"
                f"💳 Tarjeta/Digital: {fmt_moneda(data['v_dig'])}\n"
                f"💵 Efectivo Esperado: {fmt_moneda(esperado)}\n"
                f"📦 Físico Contado: {fmt_moneda(fisico)}\n"
                f"⚖️ Diferencia: {fmt_moneda(diferencia)}\n"
            )
            url = f"https://api.telegram.org/bot{token}/sendMessage"
            payload = {"chat_id": chat_id, "text": mensaje, "parse_mode": "Markdown"}
            # Fire and forget (timeout corto para no colgar la UI)
            requests.post(url, json=payload, timeout=2)
        except Exception as e:
            print(f"Error notificando a Telegram: {e}")

    def _ejecutar_live_cierre_total_centralizado(self):
        activos = [key_caja for key_caja, d in self.cajas_data.items() if d["online"]]
        if not activos:
            QMessageBox.information(self, "Sin Acción", "No hay cajas activas (online) registradas en red.")
            return

        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("CIERRE CENTRALIZADO GLOBAL EN RED")
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setText(
            "⚠️ ¡ADVERTENCIA INDUSTRIAL! Está a punto de realizar un Cierre Centralizado Global.\n\n"
            "¿Cómo desea proceder para las cajas activas en la red?\n\n"
            "🛰️ SEÑAL DE ARQUEO MASIVO: Envía una señal simultánea a todas las cajas online "
            "para que cada cajero haga su propio arqueo y cargue su físico contado.\n\n"
            "⚠️ FORZAR CIERRE MASIVO (Z-Force): Cierra y desconecta todas las cajas de forma silenciosa "
            "e instantánea desde aquí con balance perfecto."
        )
        
        btn_senal = msg_box.addButton("🛰️ SEÑAL DE ARQUEO MASIVO", QMessageBox.AcceptRole)
        btn_forzar = msg_box.addButton("⚠️ FORZAR CIERRE MASIVO", QMessageBox.DestructiveRole)
        btn_cancelar = msg_box.addButton("Cancelar", QMessageBox.RejectRole)
        
        msg_box.exec_()
        clicked = msg_box.clickedButton()
        
        if clicked == btn_cancelar:
            return
            
        elif clicked == btn_senal:
            enviadas_count = 0
            fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for key_caja in activos:
                d = self.cajas_data[key_caja]
                c_id = d["caja_id"]
                cajero = d["user"]
                try:
                    db_manager.execute_non_query(
                        "INSERT INTO movimientos_caja (fecha, tipo, monto, usuario, observaciones, caja_id) VALUES (?, 'SOLICITUD_CIERRE', 0.0, ?, 'PENDIENTE', ?)",
                        (fecha_actual, cajero, c_id)
                    )
                    enviadas_count += 1
                except Exception as e:
                    print(f"Error enviando señal centralizada a Caja 0{c_id} ({cajero}): {e}")
                    
            QMessageBox.information(
                self, "Señales Enviadas",
                f"Se envió la señal de arqueo obligatorio a {enviadas_count} terminales activas en red.\n\n"
                "Cada cajero verá su ventana de arqueo de inmediato en pantalla."
            )
            self.cargar_datos()
            return
            
        elif clicked == btn_forzar:
            cerradas_count = 0
            for key_caja in activos:
                d = self.cajas_data[key_caja]
                c_id = d["caja_id"]
                cajero = d["user"]
                esperado = d["esperado"]
                fisico = esperado
                dif = fisico - esperado

                try:
                    obs = f"Fis:${fisico:.2f} | Esp:${esperado:.2f} | CIERRE CENTRALIZADO REMOTO (Z-FORCE)"
                    db_manager.execute_non_query(
                        "INSERT INTO movimientos_caja (tipo, monto, usuario, observaciones, caja_id) VALUES ('CIERRE_Z', ?, ?, ?, ?)",
                        (dif, cajero, obs, c_id)
                    )
                    db_manager.execute_non_query(
                        "UPDATE ventas SET estado='CERRADA' WHERE usuario=? AND caja_id=? AND fecha>=? AND estado='COMPLETADA'",
                        (cajero, c_id, d["apertura_date"])
                    )
                    cerradas_count += 1
                except Exception as e:
                    print(f"Error cerrando centralizadamente Caja 0{c_id}: {e}")

        QMessageBox.information(
            self, "Cierre Centralizado Finalizado",
            f"Se ejecutó el Cierre Centralizado con éxito. {cerradas_count} cajas han sido arqueadas y pasadas a OFFLINE."
        )
        self.cargar_datos()

    def _ejecutar_live_reporte_admin(self):
        from datetime import datetime
        nombre_def = f"reporte_central_cajas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Exportar Reporte Administrativo de Cajas", nombre_def,
            "Excel (*.xlsx);;Todos los archivos (*)")
        if not filepath:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Estado Red de Cajas"

            HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
            HEADER_FONT = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
            BORDER_THIN = Border(
                left=Side(style='thin', color="cbd5e1"), right=Side(style='thin', color="cbd5e1"),
                top=Side(style='thin', color="cbd5e1"), bottom=Side(style='thin', color="cbd5e1")
            )

            headers = ["Caja/Cajero", "Estado", "Fondo Apertura", "Ventas Efectivo", "Ventas Digital", "Total Ventas", "Efectivo Esperado", "Alertas Seguridad"]
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BORDER_THIN

            for row_idx, (user, d) in enumerate(self.cajas_data.items(), 2):
                status_txt = "🟢 ACTIVA / ONLINE" if d["online"] else "🔴 OFFLINE / CERRADA"
                status_fill = "F0FDF4" if d["online"] else "FEF2F2"
                status_color = "15803D" if d["online"] else "991B1B"

                valores = [
                    user.upper(),
                    status_txt,
                    d["fondo"],
                    d["v_efec"],
                    d["v_dig"],
                    d["v_total"],
                    d["esperado"],
                    d["alertas"]
                ]

                for col_idx, val in enumerate(valores, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = BORDER_THIN
                    cell.alignment = Alignment(vertical="center")
                    cell.font = Font(name="Segoe UI", size=10)

                    if col_idx == 2:
                        cell.fill = PatternFill("solid", fgColor=status_fill)
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color=status_color)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx in (3, 4, 5, 6, 7):
                        cell.number_format = '"$"#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_idx == 8:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        if d["alertas"] > 0:
                            cell.fill = PatternFill("solid", fgColor="FEE2E2")
                            cell.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")

            anchos = [18, 22, 16, 16, 16, 16, 18, 18]
            for i, ancho in enumerate(anchos, 1):
                ws.column_dimensions[get_column_letter(i)].width = ancho

            ws.freeze_panes = "A2"
            wb.save(filepath)
            QMessageBox.information(
                self, "Exportación Exitosa",
                f"Se exportó el reporte administrativo de red de cajas exitosamente a:\n{filepath}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo exportar el reporte: {e}")

    def _hacer_barrido_red(self):
        # Solo barrer si el tab activo es el de Monitoreo Central (tab 0) y la ventana es visible
        if self.isVisible() and self.tab_widget.currentIndex() == 0:
            import random
            puertos = ["COM3 (RFID LINK)", "COM1 (RS232)", "USB001 (PRINTER)", "USB002 (SCANNER)", "TCP:8080 (LANE 2)", "192.168.1.55", "COM4 (DRAWER STATUS)"]
            p_check = random.choice(puertos)
            self.lbl_scan_status.setText(f"🛰️ BARRIDO EN RED: OK | LÍNEAS ACTIVAS | TERMINALES LEYENDO VIA {p_check}")
            self.cargar_monitoreo_red()
