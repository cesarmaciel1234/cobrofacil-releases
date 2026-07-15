import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PyQt6.QtCore import QThread, pyqtSignal
from datetime import datetime

try:
    from src.base_de_datos.database import db_manager
except ImportError:
    from database import db_manager

class WorkerExportGanancias(QThread):
    """
    Worker que extrae todas las ventas históricas de la base de datos,
    las cruza con los costos actuales del inventario, agrupa la ganancia por mes
    y genera un archivo Excel (.xlsx) detallado sin congelar la app.
    """
    finished = pyqtSignal(bool, str)

    def __init__(self, filepath):
        super().__init__()
        self.filepath = filepath

    def run(self):
        try:
            # Consulta SQL para agrupar ingresos y costos por mes
            # Extraemos Año y Mes usando strftime (SQLite) o DATE_FORMAT (MariaDB)
            # Para ser compatibles con ambos motores, extraemos los primeros 7 caracteres de la fecha: 'YYYY-MM'
            query = """
                SELECT 
                    SUBSTR(v.fecha, 1, 7) as mes,
                    SUM(dv.cantidad * dv.precio_unitario) as total_ingreso,
                    SUM(dv.cantidad * COALESCE(p.costo, 0)) as total_costo
                FROM detalles_ventas dv
                JOIN ventas v ON dv.id_venta = v.id
                LEFT JOIN productos p ON dv.id_producto = p.codigo
                WHERE v.estado = 'COMPLETADA'
                GROUP BY mes
                ORDER BY mes DESC
            """
            rows = db_manager.execute_query(query) or []

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ganancias por Mes"

            # ── Encabezados ─────────────────────────
            headers = ["Mes (Año-Mes)", "Total Ingresos", "Total Costos", "Ganancia Neta", "Margen %"]
            
            fill_header = PatternFill("solid", fgColor="4F46E5") # Indigo
            font_header = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin', color="DDDDDD"), 
                right=Side(style='thin', color="DDDDDD"),
                top=Side(style='thin', color="DDDDDD"),  
                bottom=Side(style='thin', color="DDDDDD")
            )

            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            # ── Filas de Datos ────────────────────────
            total_ingreso_global = 0.0
            total_costo_global = 0.0
            total_ganancia_global = 0.0

            for row_idx, r in enumerate(rows, 2):
                mes = r.get("mes") or "Desconocido"
                ingreso = float(r.get("total_ingreso") or 0.0)
                costo = float(r.get("total_costo") or 0.0)
                ganancia = ingreso - costo
                margen = (ganancia / ingreso * 100) if ingreso > 0 else 0.0
                
                total_ingreso_global += ingreso
                total_costo_global += costo
                total_ganancia_global += ganancia

                valores = [mes, ingreso, costo, ganancia, margen]

                for col_idx, val in enumerate(valores, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = border
                    
                    if col_idx in (2, 3, 4):  # Moneda
                        cell.number_format = '"$"#,##0.00'
                        cell.alignment = Alignment(horizontal="right")
                    elif col_idx == 5:        # Porcentaje
                        cell.number_format = '0.00"%"'
                        cell.alignment = Alignment(horizontal="center")
                        if margen < 0:
                            cell.font = Font(color="DC2626") # Rojo si hay pérdida
                        else:
                            cell.font = Font(color="16A34A") # Verde si es ganancia

            # ── Fila de Totales Generales ─────────────
            last_row = len(rows) + 2
            ws.cell(row=last_row, column=1, value="TOTAL GENERAL").font = Font(bold=True)
            
            c_ing = ws.cell(row=last_row, column=2, value=total_ingreso_global)
            c_ing.number_format = '"$"#,##0.00'
            c_ing.font = Font(bold=True)
            
            c_cost = ws.cell(row=last_row, column=3, value=total_costo_global)
            c_cost.number_format = '"$"#,##0.00'
            c_cost.font = Font(bold=True)

            c_gan = ws.cell(row=last_row, column=4, value=total_ganancia_global)
            c_gan.number_format = '"$"#,##0.00'
            c_gan.font = Font(bold=True, color="16A34A" if total_ganancia_global >= 0 else "DC2626")

            # ── Anchos de columna ─────────────────────
            anchos = [18, 20, 20, 20, 15]
            for i, ancho in enumerate(anchos, 1):
                ws.column_dimensions[get_column_letter(i)].width = ancho

            wb.save(self.filepath)
            self.finished.emit(True, f"Reporte de ganancias generado exitosamente:\n{self.filepath}")

        except Exception as e:
            self.finished.emit(False, str(e))
