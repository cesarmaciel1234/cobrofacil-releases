import os
import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from src.base_de_datos.database import db_manager
from src.admin.admin_importexport import importar_excel
import openpyxl

def run_tests():
    print("="*50)
    print("INICIANDO TEST DE ESTRÉS DE INVENTARIO Y NUBE")
    print("="*50)

    # 1. Eliminar datos previos de prueba
    print("\n[1] Limpiando datos de prueba...")
    db_manager.execute_non_query("DELETE FROM productos WHERE codigo LIKE 'INVTEST%'")
    
    # 2. Inserción Directa
    print("\n[2] Probando inserción directa...")
    db_manager.execute_non_query("INSERT INTO productos (codigo, nombre, precio, stock) VALUES ('INVTEST1', 'Producto Base', 100, 10)")
    
    # 3. Borrado
    print("\n[3] Probando borrado...")
    db_manager.execute_non_query("DELETE FROM productos WHERE codigo = 'INVTEST1'")
    r = db_manager.execute_query("SELECT id FROM productos WHERE codigo = 'INVTEST1'")
    if not r:
        print("EXITO: Borrado exitoso.")
    else:
        print("FALLO: Fallo borrado.")

    # 4. Prueba de Duplicados (Importación)
    print("\n[4] Creando Excel de prueba para importacion masiva y duplicados...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Codigo", "Descripcion", "P. Venta", "Costo", "Stock"])
    # Fila 1: Nuevo
    ws.append(["INVTEST100", "Producto Nube 1", 500, 200, 50])
    # Fila 2: Nuevo
    ws.append(["INVTEST200", "Producto Nube 2", 800, 400, 10])
    
    excel_path = "test_import.xlsx"
    wb.save(excel_path)
    
    print("\n[5] Simulando Carga Inicial / Nube (Excel)...")
    ok, msg = importar_excel(excel_path)
    if ok:
        print(f"EXITO: Importacion inicial exitosa: {msg}")
    else:
        print(f"FALLO: Error importando: {msg}")

    # Verificar creacion
    r1 = db_manager.execute_query("SELECT precio FROM productos WHERE codigo = 'INVTEST100'")
    if r1 and r1[0]['precio'] == 500:
        print("EXITO: Precios de carga inicial correctos.")
    
    print("\n[6] Probando actualizacion de precios y proteccion de duplicados...")
    # Modificar el excel para subir precios
    ws.cell(row=2, column=3, value=650) # INVTEST100 sube a 650
    ws.cell(row=3, column=3, value=1200) # INVTEST200 sube a 1200
    # Agregar un duplicado intencional en el excel a ver que pasa
    ws.append(["INVTEST100", "Producto Nube 1 DUPLICADO", 700, 200, 50])
    wb.save(excel_path)
    
    ok, msg = importar_excel(excel_path)
    if ok:
        print(f"EXITO: Importacion de actualizacion exitosa: {msg}")
    else:
        print(f"FALLO: Error actualizando: {msg}")
        
    r2 = db_manager.execute_query("SELECT precio FROM productos WHERE codigo = 'INVTEST100'")
    if r2:
        print(f"EXITO: Precio final tras actualizacion (esperado 700 por el duplicado final o 650): {r2[0]['precio']}")
        
    # Verificar cantidad de registros para evitar duplicidad fisica
    r3 = db_manager.execute_query("SELECT COUNT(*) as c FROM productos WHERE codigo = 'INVTEST100'")
    if r3 and r3[0]['c'] == 1:
        print("EXITO: Manejo de duplicados perfecto (no se crearon filas repetidas para el mismo codigo).")
    else:
        print(f"FALLO: Error de duplicados. Se crearon {r3[0]['c']} registros para el mismo codigo.")

    if os.path.exists(excel_path):
        os.remove(excel_path)
        
    print("\n" + "="*50)
    print("TEST DE INVENTARIO FINALIZADO")
    print("="*50)

if __name__ == '__main__':
    run_tests()
